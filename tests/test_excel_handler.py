"""
Testes para o módulo excel_handler.py
"""
import pytest
import openpyxl
from pathlib import Path
from Src.infrastructure.exporters.excel_handler_pmpv import ExcelHandlerPMPV


class TestExcelHandlerPMPV:
    """Suite de testes para ExcelHandlerPMPV"""
    
    @pytest.fixture
    def dados_exemplo(self):
        """Dados de exemplo para testes"""
        return {
            'Janeiro': [
                {
                    'empresa': 'PETROBRAS',
                    'molecula': 1.5000,
                    'transporte': 0.3000,
                    'logistica': 0.2000,
                    'volume': 1000
                },
                {
                    'empresa': 'GALP',
                    'molecula': 1.6000,
                    'transporte': 0.3500,
                    'logistica': 0.2500,
                    'volume': 800
                }
            ],
            'Fevereiro': [
                {
                    'empresa': 'PETROBRAS',
                    'molecula': 1.5200,
                    'transporte': 0.3100,
                    'logistica': 0.2100,
                    'volume': 1050
                }
            ],
            'Março': [
                {
                    'empresa': 'GALP',
                    'molecula': 1.5800,
                    'transporte': 0.3400,
                    'logistica': 0.2400,
                    'volume': 900
                }
            ]
        }
    
    @pytest.fixture
    def resultado_exemplo(self):
        """Resultado de exemplo para testes"""
        return {
            'volume_total': 90000,
            'vp_mensal': 90000,
            'custo_total': 180000,
            'pmpv': 2.0000,
            'conta_grafica': -0.0210,
            'preco_final': 1.9790
        }
    
    def test_exportar_trimestre_cria_arquivo(self, dados_exemplo, resultado_exemplo, tmp_path):
        """Testa se o arquivo Excel é criado"""
        arquivo = tmp_path / "teste.xlsx"
        
        nome_criado = ExcelHandlerPMPV.exportar_trimestre(
            dados_exemplo,
            resultado_exemplo,
            str(arquivo)
        )
        
        assert Path(nome_criado).exists()
        assert nome_criado == str(arquivo)
    
    def test_exportar_sem_nome_gera_timestamp(self, dados_exemplo, resultado_exemplo, tmp_path, monkeypatch):
        """Testa se gera nome com timestamp quando não fornecido"""
        # Muda o diretório de trabalho para tmp_path
        monkeypatch.chdir(tmp_path)
        
        nome_criado = ExcelHandlerPMPV.exportar_trimestre(
            dados_exemplo,
            resultado_exemplo
        )
        
        assert Path(nome_criado).exists()
        assert nome_criado.startswith("Relatorio_PMPV_")
        assert nome_criado.endswith(".xlsx")
    
    # ── Constantes do layout atual (banner + cabeçalho + dados) ──────────────
    # O relatório foi redesenhado: cada aba mensal tem um banner (linhas 1-2),
    # um espaçador (linha 3), o cabeçalho (linha 4) e os dados a partir da
    # linha 5. O resumo virou cards de KPI + uma tabela de indicadores.
    _LINHA_CABECALHO = 4
    _LINHA_PRIMEIRO_DADO = 5
    _ABA_RESUMO = "📋 Resumo Executivo"

    @staticmethod
    def _achar_indicador(ws, rotulo_parcial):
        """Retorna o valor (coluna B) da linha de indicador cujo rótulo contém o texto.

        Procura apenas linhas onde a coluna B tem um número — assim ignora os
        cards de KPI (que guardam o valor na coluna A) e pega a tabela de
        indicadores complementares.
        """
        for r in range(1, ws.max_row + 1):
            rotulo = ws.cell(r, 1).value
            valor = ws.cell(r, 2).value
            if rotulo and rotulo_parcial in str(rotulo) and isinstance(valor, (int, float)):
                return valor
        return None

    def test_arquivo_contem_abas_corretas(self, dados_exemplo, resultado_exemplo, tmp_path):
        """Verifica se as abas são criadas corretamente"""
        arquivo = tmp_path / "teste.xlsx"

        ExcelHandlerPMPV.exportar_trimestre(dados_exemplo, resultado_exemplo, str(arquivo))

        wb = openpyxl.load_workbook(arquivo)

        # Verifica aba de resumo
        assert self._ABA_RESUMO in wb.sheetnames

        # Verifica abas mensais
        assert "Janeiro" in wb.sheetnames
        assert "Fevereiro" in wb.sheetnames
        assert "Março" in wb.sheetnames

        wb.close()

    def test_aba_mes_contem_cabecalhos(self, dados_exemplo, resultado_exemplo, tmp_path):
        """Verifica se os cabeçalhos estão corretos (na linha de cabeçalho atual)"""
        arquivo = tmp_path / "teste.xlsx"

        ExcelHandlerPMPV.exportar_trimestre(dados_exemplo, resultado_exemplo, str(arquivo))

        wb = openpyxl.load_workbook(arquivo)
        ws = wb["Janeiro"]

        # Cabeçalhos podem ter quebras de linha com a unidade (ex.: "Molécula\n(R$/m³)").
        # Comparamos só o "rótulo principal" (primeira linha de cada célula).
        rotulos_esperados = ["Empresa", "Molécula", "Transporte", "Logística",
                             "Preço Unit.", "Volume", "Custo"]
        rotulos_encontrados = [
            str(ws.cell(self._LINHA_CABECALHO, c).value or "").split("\n")[0]
            for c in range(1, 8)
        ]

        assert rotulos_encontrados == rotulos_esperados

        wb.close()

    def test_aba_mes_calcula_valores_corretamente(self, dados_exemplo, resultado_exemplo, tmp_path):
        """Verifica se os cálculos na aba estão corretos"""
        arquivo = tmp_path / "teste.xlsx"

        ExcelHandlerPMPV.exportar_trimestre(dados_exemplo, resultado_exemplo, str(arquivo))

        wb = openpyxl.load_workbook(arquivo)
        ws = wb["Janeiro"]

        # Primeira linha de dados (PETROBRAS)
        r = self._LINHA_PRIMEIRO_DADO
        empresa = ws.cell(r, 1).value
        molecula = ws.cell(r, 2).value
        transporte = ws.cell(r, 3).value
        logistica = ws.cell(r, 4).value
        preco_unit = ws.cell(r, 5).value
        volume = ws.cell(r, 6).value
        custo_total = ws.cell(r, 7).value

        assert empresa == "PETROBRAS"
        assert molecula == 1.5
        assert transporte == 0.3
        assert logistica == 0.2
        assert preco_unit == 2.0  # 1.5 + 0.3 + 0.2
        assert volume == 1000
        assert custo_total == 2000  # 2.0 * 1000

        wb.close()

    def test_aba_resumo_contem_totais(self, dados_exemplo, resultado_exemplo, tmp_path):
        """Verifica se a aba de resumo contém os totais corretos.

        O resumo agora é um layout de cards + uma tabela de indicadores.
        Buscamos os valores pelo rótulo, sem depender de linhas fixas.
        """
        arquivo = tmp_path / "teste.xlsx"

        ExcelHandlerPMPV.exportar_trimestre(dados_exemplo, resultado_exemplo, str(arquivo))

        wb = openpyxl.load_workbook(arquivo)
        ws = wb[self._ABA_RESUMO]

        # Tabela "Indicadores complementares": valores buscados pelo rótulo.
        assert self._achar_indicador(ws, "Volume Prospectivo") == 90000
        assert self._achar_indicador(ws, "Custo Total") == 180000
        assert self._achar_indicador(ws, "PMPV") == 2.0000
        assert self._achar_indicador(ws, "Conta Gráfica") == -0.0210
        assert self._achar_indicador(ws, "Preço Final") == 1.9790

        wb.close()
    
    def test_dados_vazios_nao_sao_adicionados(self, resultado_exemplo, tmp_path):
        """Verifica que linhas sem empresa não são adicionadas"""
        dados_com_vazios = {
            'Janeiro': [
                {
                    'empresa': 'PETROBRAS',
                    'molecula': 1.5,
                    'transporte': 0.3,
                    'logistica': 0.2,
                    'volume': 1000
                },
                {
                    'empresa': '',  # Vazio
                    'molecula': 1.6,
                    'transporte': 0.35,
                    'logistica': 0.25,
                    'volume': 800
                },
                {
                    'empresa': None,  # None
                    'molecula': 1.7,
                    'transporte': 0.4,
                    'logistica': 0.3,
                    'volume': 900
                }
            ]
        }
        
        arquivo = tmp_path / "teste.xlsx"
        ExcelHandlerPMPV.exportar_trimestre(dados_com_vazios, resultado_exemplo, str(arquivo))

        wb = openpyxl.load_workbook(arquivo)
        ws = wb["Janeiro"]

        # Conta as linhas de dados pelas empresas na coluna A, a partir da
        # primeira linha de dados. Só PETROBRAS deve aparecer (vazias ignoradas).
        empresas = []
        for r in range(self._LINHA_PRIMEIRO_DADO, ws.max_row + 1):
            valor = ws.cell(r, 1).value
            # Para de contar ao chegar nas linhas de total/rodapé.
            if valor and "PETROBRAS" in str(valor):
                empresas.append(valor)
        assert empresas == ["PETROBRAS"]

        wb.close()

    def test_formatacao_numerica(self, dados_exemplo, resultado_exemplo, tmp_path):
        """Verifica se a formatação numérica está aplicada na primeira linha de dados"""
        arquivo = tmp_path / "teste.xlsx"

        ExcelHandlerPMPV.exportar_trimestre(dados_exemplo, resultado_exemplo, str(arquivo))

        wb = openpyxl.load_workbook(arquivo)
        ws = wb["Janeiro"]

        # Verifica que as células numéricas têm formato de moeda/número (não 'General').
        r = self._LINHA_PRIMEIRO_DADO
        for col in (2, 3, 4, 5, 7):  # Molécula, Transporte, Logística, Preço, Custo
            fmt = ws.cell(r, col).number_format
            assert "R$" in fmt or "#,##0" in fmt, f"col {col} sem formato numérico: {fmt!r}"
        # Volume tem formato numérico próprio.
        assert "#,##0" in ws.cell(r, 6).number_format

        wb.close()
