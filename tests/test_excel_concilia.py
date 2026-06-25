from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from Src.Services.excel_concilia import ExcelConcilia
from Src.Services.servicos_concilia import PdfItem


def _item(file_name="a.pdf", category="Receita", amount=100.0, status="OK", method="m"):
    return PdfItem(
        file_name=file_name, file_path=f"/x/{file_name}",
        category=category, amount=amount, status=status, method=method,
    )


class TestGerarRelatorio:
    def test_gera_arquivo(self, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        ExcelConcilia.gerar_relatorio(out, [_item()])
        assert out.exists()

    def test_totais_receita_e_despesa(self, tmp_path):
        itens = [
            _item("r1.pdf", "Receita", 1000.0, "OK"),
            _item("r2.pdf", "Receita", 500.0, "OK"),
            _item("d1.pdf", "Despesa", 200.0, "OK"),
            _item("ignorada.pdf", "Receita", 999.0, "REVISAR"),  # não soma
        ]
        out = tmp_path / "out.xlsx"
        total_rec, total_desp = ExcelConcilia.gerar_relatorio(out, itens)
        assert total_rec == 1500.0
        assert total_desp == 200.0

    def test_layout_basico_estrutura(self, tmp_path):
        out = tmp_path / "x.xlsx"
        ExcelConcilia.gerar_relatorio(out, [_item()])
        wb = load_workbook(out)
        ws = wb.active
        # Header na linha 1 quando layout não-bonito
        assert ws.cell(1, 1).value == "Arquivo"
        assert ws.cell(1, 3).value == "Valor"

    def test_layout_bonito_com_titulo(self, tmp_path):
        out = tmp_path / "y.xlsx"
        ExcelConcilia.gerar_relatorio(out, [_item()], layout_bonito=True)
        wb = load_workbook(out)
        ws = wb.active
        # Linha 1 é o título mesclado
        assert "CONCILIACAO" in str(ws.cell(1, 1).value).upper()
        # Header desce pra linha 2
        assert ws.cell(2, 1).value == "Arquivo"

    def test_resumo_final_presente(self, tmp_path):
        out = tmp_path / "z.xlsx"
        ExcelConcilia.gerar_relatorio(
            out,
            [_item("r.pdf", "Receita", 300.0), _item("d.pdf", "Despesa", 100.0)],
        )
        wb = load_workbook(out)
        ws = wb.active
        valores = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("RESUMO FINAL" in v for v in valores)
        assert any("RECEITAS" in v for v in valores)
        assert any("DESPESAS" in v for v in valores)
        assert any("SALDO" in v for v in valores)

    def test_lista_vazia(self, tmp_path):
        out = tmp_path / "vazio.xlsx"
        total_rec, total_desp = ExcelConcilia.gerar_relatorio(out, [])
        assert total_rec == 0.0
        assert total_desp == 0.0
        assert out.exists()
