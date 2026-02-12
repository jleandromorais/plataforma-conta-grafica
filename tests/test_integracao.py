"""
Testes de integração entre módulos
"""
import pytest
from pathlib import Path
from database import DatabasePMPV
from excel_handler import ExcelHandlerPMPV
import openpyxl


class TestIntegracaoDatabaseExcel:
    """Testes de integração entre Database e ExcelHandler"""
    
    @pytest.fixture
    def db_temp(self, tmp_path):
        """Cria um banco de dados temporário"""
        db_path = tmp_path / "test_integration.db"
        db = DatabasePMPV(str(db_path))
        yield db
        db.fechar()
    
    @pytest.fixture
    def dados_completos(self):
        """Dados completos para teste de integração"""
        return {
            'dados_por_mes': {
                'Janeiro': [
                    {
                        'empresa': 'PETROBRAS',
                        'molecula': 1.5000,
                        'transporte': 0.3000,
                        'logistica': 0.2000,
                        'volume': 1000
                    },
                    {
                        'empresa': 'GALP',
                        'molecula': 1.6000,
                        'transporte': 0.3500,
                        'logistica': 0.2500,
                        'volume': 800
                    }
                ],
                'Fevereiro': [
                    {
                        'empresa': 'PETROBRAS',
                        'molecula': 1.5200,
                        'transporte': 0.3100,
                        'logistica': 0.2100,
                        'volume': 1050
                    }
                ],
                'Março': [
                    {
                        'empresa': 'GALP',
                        'molecula': 1.5800,
                        'transporte': 0.3400,
                        'logistica': 0.2400,
                        'volume': 900
                    }
                ]
            },
            'resultado': {
                'volume_total': 90000,
                'custo_total': 180000,
                'pmpv': 2.0000,
                'conta_grafica': -0.0210,
                'preco_final': 1.9790
            }
        }
    
    def test_fluxo_completo_salvar_e_exportar(self, db_temp, dados_completos, tmp_path):
        """Testa o fluxo completo: salvar no BD e exportar para Excel"""
        # 1. Criar sessão
        sessao_id = db_temp.criar_sessao("Teste Integração", "Teste completo")
        assert sessao_id > 0
        
        # 2. Salvar dados dos meses
        for i, (mes, dados) in enumerate(dados_completos['dados_por_mes'].items(), 1):
            sucesso = db_temp.salvar_dados_mes(sessao_id, i, dados)
            assert sucesso is True
        
        # 3. Salvar resultado
        res = dados_completos['resultado']
        sucesso = db_temp.salvar_resultado(
            sessao_id,
            res['volume_total'],
            res['custo_total'],
            res['pmpv'],
            res['conta_grafica'],
            res['preco_final']
        )
        assert sucesso is True
        
        # 4. Exportar para Excel
        arquivo_excel = tmp_path / "teste_integracao.xlsx"
        nome_arquivo = ExcelHandlerPMPV.exportar_trimestre(
            dados_completos['dados_por_mes'],
            dados_completos['resultado'],
            str(arquivo_excel)
        )
        
        assert Path(nome_arquivo).exists()
        
        # 5. Verificar conteúdo do Excel
        wb = openpyxl.load_workbook(arquivo_excel)
        
        # Verifica aba de resumo
        ws_resumo = wb["Resumo Executivo"]
        assert ws_resumo["B3"].value == 90000  # Volume Total
        assert ws_resumo["B4"].value == 180000  # Custo Total
        
        # Verifica aba Janeiro
        ws_jan = wb["Janeiro"]
        assert ws_jan.cell(2, 1).value == "PETROBRAS"
        assert ws_jan.cell(3, 1).value == "GALP"
        
        wb.close()
    
    def test_recuperar_e_reexportar(self, db_temp, dados_completos, tmp_path):
        """Testa recuperar dados do BD e reexportar"""
        # Salva dados
        sessao_id = db_temp.criar_sessao("Teste Recuperação")
        
        for i, (mes, dados) in enumerate(dados_completos['dados_por_mes'].items(), 1):
            db_temp.salvar_dados_mes(sessao_id, i, dados)
        
        # Recupera dados
        dados_recuperados = {}
        meses = ['Janeiro', 'Fevereiro', 'Março']
        
        for i, mes in enumerate(meses, 1):
            dados_mes = db_temp.carregar_dados_mes(sessao_id, i)
            dados_recuperados[mes] = dados_mes
        
        # Verifica que recuperou corretamente
        assert len(dados_recuperados['Janeiro']) == 2
        assert len(dados_recuperados['Fevereiro']) == 1
        assert len(dados_recuperados['Março']) == 1
        
        # Exporta dados recuperados
        arquivo_excel = tmp_path / "teste_recuperacao.xlsx"
        nome_arquivo = ExcelHandlerPMPV.exportar_trimestre(
            dados_recuperados,
            dados_completos['resultado'],
            str(arquivo_excel)
        )
        
        assert Path(nome_arquivo).exists()


class TestCalculosPMPV:
    """Testes de cálculos PMPV"""
    
    def test_calculo_pmpv_simples(self):
        """Testa cálculo PMPV com dados simples"""
        # Dados: 1 empresa, 1 mês
        molecula = 1.5
        transporte = 0.3
        logistica = 0.2
        volume_diario = 1000
        dias = 30
        
        preco_unitario = molecula + transporte + logistica
        volume_total = volume_diario * dias
        custo_total = preco_unitario * volume_total
        pmpv = custo_total / volume_total
        
        assert preco_unitario == 2.0
        assert volume_total == 30000
        assert custo_total == 60000
        assert pmpv == 2.0
    
    def test_calculo_pmpv_multiplas_empresas(self):
        """Testa cálculo PMPV com múltiplas empresas"""
        empresas = [
            {'preco': 2.0, 'volume_diario': 1000, 'dias': 30},  # 60.000
            {'preco': 2.2, 'volume_diario': 800, 'dias': 30},   # 52.800
        ]
        
        custo_total = sum(e['preco'] * e['volume_diario'] * e['dias'] for e in empresas)
        volume_total = sum(e['volume_diario'] * e['dias'] for e in empresas)
        pmpv = custo_total / volume_total
        
        assert custo_total == 112800
        assert volume_total == 54000
        assert round(pmpv, 4) == 2.0889
    
    def test_calculo_preco_final_com_conta_grafica(self):
        """Testa cálculo do preço final incluindo conta gráfica"""
        pmpv = 2.0
        conta_grafica = -0.021
        preco_final = pmpv + conta_grafica
        
        assert preco_final == 1.979
    
    def test_calculo_trimestre_completo(self):
        """Testa cálculo de um trimestre completo"""
        # 3 meses, cada um com diferentes volumes
        mes1 = {'preco': 2.0, 'volume_diario': 1000, 'dias': 30}
        mes2 = {'preco': 2.1, 'volume_diario': 1100, 'dias': 28}
        mes3 = {'preco': 1.9, 'volume_diario': 900, 'dias': 31}
        
        custo_total = sum([
            mes1['preco'] * mes1['volume_diario'] * mes1['dias'],
            mes2['preco'] * mes2['volume_diario'] * mes2['dias'],
            mes3['preco'] * mes3['volume_diario'] * mes3['dias']
        ])
        
        volume_total = sum([
            mes1['volume_diario'] * mes1['dias'],
            mes2['volume_diario'] * mes2['dias'],
            mes3['volume_diario'] * mes3['dias']
        ])
        
        pmpv = custo_total / volume_total
        
        assert custo_total == 177690.0  # 60000 + 64680 + 52920
        assert volume_total == 88700   # 30000 + 30800 + 27900
        assert round(pmpv, 4) == 2.0033
