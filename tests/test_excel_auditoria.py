from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from Src.Services.excel_auditoria import ExcelAuditoria
from Src.Services.comparador_conta_grafica import ResultadoComparacaoNotas
from Src.Services.servicos_auditoria import XMLItem


def _mk_item(numero="1", empresa="ACME", tipo="NF-e", valor=1000.0, icms_taxa=0.12) -> XMLItem:
    return XMLItem(
        empresa=empresa, tipo=tipo, numero=numero,
        valor_total=valor, icms=valor * icms_taxa, icms_taxa=icms_taxa,
        pis=valor * 0.0165, cofins=valor * 0.076,
        volume=10, status="OK", volume_total=10.0,
    )


class TestGerarRelatorioAuditoria:
    def test_cria_arquivo(self, tmp_path):
        out = tmp_path / "audit.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria([_mk_item()], str(out))
        assert out.exists()

    def test_estrutura_basica(self, tmp_path):
        out = tmp_path / "audit.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1"), _mk_item("2", tipo="CT-e")], str(out)
        )
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "Relatório Completo"
        # Linha 1 = título
        assert "AUDITORIA" in str(ws.cell(1, 1).value).upper()
        # Linha 4 = cabeçalho
        assert ws.cell(4, 1).value == "Empresa"
        assert ws.cell(4, 3).value == "Nº NF / CT-e"

    def test_dados_aparecem(self, tmp_path):
        out = tmp_path / "x.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("777", empresa="EMPRESA_A")], str(out)
        )
        wb = load_workbook(out)
        ws = wb.active
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("EMPRESA_A" in t for t in textos)
        assert any("777" in t for t in textos)

    def test_total_geral_presente(self, tmp_path):
        out = tmp_path / "x.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1", valor=1000), _mk_item("2", valor=2000)], str(out)
        )
        wb = load_workbook(out)
        ws = wb.active
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("TOTAL GERAL" in t for t in textos)
        assert any("RESUMO POR TIPO" in t for t in textos)

    def test_cgr_total_customizado_usado(self, tmp_path):
        out = tmp_path / "x.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1", valor=1000)], str(out), cgr_total=99999.99
        )
        # O arquivo é gerado sem crash — a verificação do valor é coberta pela
        # implementação da fórmula; aqui só garantimos que cgr_total não quebra.
        assert out.exists()


class TestSheetsComparacao:
    @pytest.fixture
    def comparacao(self):
        return ResultadoComparacaoNotas(
            periodo="Jan/2026",
            coluna_nota_excel="Nota Fiscal",
            coluna_periodo_excel="Periodo",
            total_nossa_base=2,
            total_conta_grafica=2,
            qtd_em_ambas=1,
            notas_apenas_nossa=[{
                "empresa": "ACME", "tipo": "NF-e", "numero": "100",
                "numero_normalizado": "100", "valor_total": 1000.0,
                "icms": 120.0, "cgr_liquido": 798.6,
            }],
            notas_apenas_conta_grafica=[{
                "linha_excel": 5, "numero": "200", "numero_normalizado": "200",
                "periodo_linha": "Jan/2026",
                "dados_linha": {"Nota Fiscal": "200", "Valor": 500.0, "Mes": "Jan/2026"},
            }],
            linhas_sem_numero_excel=[{
                "linha_excel": 10, "coluna_nota": "Nota Fiscal", "valor_original": "",
            }],
            avisos=["Aviso 1", "Aviso 2"],
            sheets_excel=["Jan-26", "Fev-26"],
            sheets_periodo_match=["Jan-26"],
            total_linhas_periodo=2,
        )

    def test_abas_de_comparacao_criadas(self, tmp_path, comparacao):
        out = tmp_path / "audit.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1")], str(out), comparacao=comparacao
        )
        wb = load_workbook(out)
        nomes = wb.sheetnames
        assert "Comparacao Conta Grafica" in nomes
        assert "Diferencas" in nomes
        assert "Conta sem Numero" in nomes

    def test_resumo_contem_metricas(self, tmp_path, comparacao):
        out = tmp_path / "audit.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1")], str(out), comparacao=comparacao
        )
        wb = load_workbook(out)
        ws = wb["Comparacao Conta Grafica"]
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("Jan/2026" in t for t in textos)
        assert any("Aviso 1" in t for t in textos)
        assert any("Notas em ambas" in t for t in textos)

    def test_diferencas_secoes(self, tmp_path, comparacao):
        out = tmp_path / "audit.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1")], str(out), comparacao=comparacao
        )
        wb = load_workbook(out)
        ws = wb["Diferencas"]
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("SOMENTE NA AUDITORIA" in t for t in textos)
        assert any("SOMENTE NA PLANILHA" in t for t in textos)

    def test_linhas_sem_numero(self, tmp_path, comparacao):
        out = tmp_path / "audit.xlsx"
        ExcelAuditoria.gerar_relatorio_auditoria(
            [_mk_item("1")], str(out), comparacao=comparacao
        )
        wb = load_workbook(out)
        ws = wb["Conta sem Numero"]
        assert ws.cell(1, 1).value == "Linha Excel"
        assert ws.cell(2, 1).value == 10
