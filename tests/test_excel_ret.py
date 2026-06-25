from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from Src.Services.excel_ret import ExcelRET, _mes_abrev_para_num, _trimestre_do_mes, _chave_trimestre


class TestHelpersInternos:
    def test_mes_abrev_jan(self):
        assert _mes_abrev_para_num("Jan") == 1

    def test_mes_abrev_dez(self):
        assert _mes_abrev_para_num("Dez") == 12

    def test_mes_abrev_invalido(self):
        assert _mes_abrev_para_num("XYZ") == 0

    def test_trimestre_jan_no_nov_jan(self):
        assert _trimestre_do_mes(1) == "Nov - Jan"

    def test_trimestre_abril_no_fev_abr(self):
        assert _trimestre_do_mes(4) == "Fev - Abr"

    def test_trimestre_mes_invalido(self):
        assert _trimestre_do_mes(0) == "Sem Trimestre"

    def test_chave_trimestre_nov_jan_zero(self):
        assert _chave_trimestre("Nov - Jan") == 0

    def test_chave_trimestre_desconhecido(self):
        assert _chave_trimestre("xyz") == 99


# ════════════════════════════════════════════════════════════════════════════
# Relatório completo
# ════════════════════════════════════════════════════════════════════════════

def _doc(tipo="EAT", empresa="ACME", valor=1000.0, qt=10.0, mes_ref="Jan/2026"):
    return {
        "tipo_encargo": tipo,
        "empresa": empresa,
        "nota_tipo": "Débito",
        "numero_nd": "001",
        "data_vencimento": "01/01/2026",
        "valor_total": valor,
        "quantidade": qt,
        "valor_unitario": valor / qt if qt else 0.0,
        "arquivo": f"arq_{tipo}.pdf",
        "mes_ref": mes_ref,
    }


@pytest.fixture
def calc_ret():
    return {
        "eat_bruto": 1000.0,
        "pis_cofins_rate": 0.0925,
        "ret": 907.5,
    }


class TestGerarRelatorioCompleto:
    def test_cria_arquivo(self, tmp_path, calc_ret):
        out = tmp_path / "ret.xlsx"
        ExcelRET.gerar_relatorio_completo([_doc()], calc_ret, out)
        assert out.exists()

    def test_abas_esperadas(self, tmp_path, calc_ret):
        out = tmp_path / "ret.xlsx"
        ExcelRET.gerar_relatorio_completo([_doc()], calc_ret, out)
        wb = load_workbook(out)
        nomes = wb.sheetnames
        assert "Dados Completos" in nomes
        assert "Resumo por Tipo" in nomes
        assert "Resumo Geral" in nomes
        assert "Por Mês" in nomes

    def test_aba_dados_tem_header(self, tmp_path, calc_ret):
        out = tmp_path / "ret.xlsx"
        ExcelRET.gerar_relatorio_completo([_doc()], calc_ret, out)
        wb = load_workbook(out)
        ws = wb["Dados Completos"]
        assert ws.cell(1, 1).value == "Tipo de Encargo"
        assert ws.cell(2, 1).value == "EAT"

    def test_resumo_geral_inclui_metricas(self, tmp_path, calc_ret):
        out = tmp_path / "ret.xlsx"
        ExcelRET.gerar_relatorio_completo([_doc(), _doc(valor=500)], calc_ret, out)
        wb = load_workbook(out)
        ws = wb["Resumo Geral"]
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("RESUMO GERAL" in t for t in textos)
        assert any("EAT bruto" in t for t in textos)

    def test_por_mes_agrupa_por_trimestre(self, tmp_path, calc_ret):
        out = tmp_path / "ret.xlsx"
        ExcelRET.gerar_relatorio_completo(
            [
                _doc(mes_ref="Jan/2026", valor=100),
                _doc(mes_ref="Fev/2026", valor=200),
                _doc(mes_ref="Mar/2026", valor=300),
            ],
            calc_ret, out,
        )
        wb = load_workbook(out)
        ws = wb["Por Mês"]
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        # Deve haver pelo menos 1 cabeçalho de trimestre
        assert any("Trimestre:" in t for t in textos)
        # E o TOTAL GERAL
        assert any("TOTAL GERAL" in t for t in textos)

    def test_docs_sem_mes_vao_para_secao_especial(self, tmp_path, calc_ret):
        out = tmp_path / "ret.xlsx"
        ExcelRET.gerar_relatorio_completo(
            [_doc(mes_ref="")],
            calc_ret, out,
        )
        wb = load_workbook(out)
        ws = wb["Por Mês"]
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert any("Sem Mês" in t for t in textos)
