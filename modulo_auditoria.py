import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import os
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configuração Visual
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# ==========================================
# CLASSES DE DADOS
# ==========================================

class XMLItem:
    """Representa um item auditado de XML fiscal"""
    def __init__(self, empresa: str, tipo: str, numero: str, valor_total: float, 
                 icms: float, pis: float, cofins: float, volume: int, status: str):
        self.empresa = empresa
        self.tipo = tipo  # NF-e ou CT-e
        self.numero = numero
        self.valor_total = valor_total
        self.icms = icms
        self.pis = pis
        self.cofins = cofins
        self.volume = volume
        self.status = status

# ==========================================
# FUNÇÕES DE PARSE XML
# ==========================================

def parse_nfe(xml_path: Path) -> Dict:
    """Extrai dados de uma NF-e"""
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        # Namespace NFe
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
        
        # Dados básicos
        numero = root.find('.//nfe:ide/nfe:nNF', ns)
        valor_total = root.find('.//nfe:total/nfe:ICMSTot/nfe:vNF', ns)
        icms = root.find('.//nfe:total/nfe:ICMSTot/nfe:vICMS', ns)
        pis = root.find('.//nfe:total/nfe:ICMSTot/nfe:vPIS', ns)
        cofins = root.find('.//nfe:total/nfe:ICMSTot/nfe:vCOFINS', ns)
        
        # Volume (tenta pegar da tag qVol)
        volume = root.find('.//nfe:vol/nfe:qVol', ns)
        
        return {
            'tipo': 'NF-e',
            'numero': numero.text if numero is not None else 'N/A',
            'valor_total': float(valor_total.text) if valor_total is not None else 0.0,
            'icms': float(icms.text) if icms is not None else 0.0,
            'pis': float(pis.text) if pis is not None else 0.0,
            'cofins': float(cofins.text) if cofins is not None else 0.0,
            'volume': int(float(volume.text)) if volume is not None else 0
        }
    except Exception as e:
        return {'erro': str(e)}

def parse_cte(xml_path: Path) -> Dict:
    """Extrai dados de um CT-e"""
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        # Namespace CTe
        ns = {'cte': 'http://www.portalfiscal.inf.br/cte'}
        
        numero = root.find('.//cte:ide/cte:nCT', ns)
        valor_total = root.find('.//cte:vPrest/cte:vTPrest', ns)
        icms = root.find('.//cte:ICMS//cte:vICMS', ns)
        
        # CT-e pode não ter PIS/COFINS no XML
        pis = root.find('.//cte:vPIS', ns)
        cofins = root.find('.//cte:vCOFINS', ns)
        
        # Volume (qCarga)
        volume = root.find('.//cte:infQ/cte:qCarga', ns)
        
        return {
            'tipo': 'CT-e',
            'numero': numero.text if numero is not None else 'N/A',
            'valor_total': float(valor_total.text) if valor_total is not None else 0.0,
            'icms': float(icms.text) if icms is not None else 0.0,
            'pis': float(pis.text) if pis is not None else 0.0,
            'cofins': float(cofins.text) if cofins is not None else 0.0,
            'volume': int(float(volume.text)) if volume is not None else 0
        }
    except Exception as e:
        return {'erro': str(e)}

def detectar_tipo_xml(xml_path: Path) -> str:
    """Detecta se é NF-e ou CT-e pelo conteúdo"""
    try:
        with open(xml_path, 'r', encoding='utf-8') as f:
            conteudo = f.read(500)  # Lê só o início
            if 'nfeProc' in conteudo or 'NFe' in conteudo:
                return 'nfe'
            elif 'cteProc' in conteudo or 'CTe' in conteudo:
                return 'cte'
    except:
        pass
    return 'desconhecido'

# ==========================================
# INTERFACE GRÁFICA
# ==========================================

class AppAuditoriaXML(ctk.CTkToplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.title("Auditoria XML - NF-e e CT-e")
        self.geometry("1300x850")
        
        # Variáveis de controle
        self.pasta_selecionada = None
        self.empresas_disponiveis = []
        self.empresas_selecionadas = []
        self.excel_path = None
        self.df_excel = None
        self.resultados = []
        
        self._setup_ui()
    
    def _setup_ui(self):
        # HEADER
        header = ctk.CTkFrame(self, height=60, corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(header, text="🔍 Auditoria XML Fiscal", 
                     font=("Roboto", 24, "bold")).pack(side="left", padx=20, pady=10)
        
        # CONTAINER PRINCIPAL
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # ========== PASSO 1: PASTA PAI ==========
        frame_pasta = ctk.CTkFrame(container)
        frame_pasta.pack(fill="x", pady=10)
        
        ctk.CTkLabel(frame_pasta, text="📁 Passo 1: Selecione a pasta PAI com subpastas de empresas",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        btn_frame = ctk.CTkFrame(frame_pasta, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(btn_frame, text="📂 Selecionar Pasta", 
                      command=self.selecionar_pasta,
                      fg_color="#3498db", hover_color="#2980b9").pack(side="left", padx=5)
        
        self.lbl_pasta = ctk.CTkLabel(btn_frame, text="Nenhuma pasta selecionada", 
                                      text_color="gray")
        self.lbl_pasta.pack(side="left", padx=10)
        
        # ========== PASSO 2: EMPRESAS ==========
        frame_empresas = ctk.CTkFrame(container)
        frame_empresas.pack(fill="x", pady=10)
        
        ctk.CTkLabel(frame_empresas, text="🏢 Passo 2: Selecione as empresas para auditar",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        # Frame com scroll para empresas
        self.scroll_empresas = ctk.CTkScrollableFrame(frame_empresas, height=120)
        self.scroll_empresas.pack(fill="x", padx=10, pady=5)
        
        self.checkboxes_empresas = []
        
        # ========== PASSO 3: EXCEL ==========
        frame_excel = ctk.CTkFrame(container)
        frame_excel.pack(fill="x", pady=10)
        
        ctk.CTkLabel(frame_excel, text="📊 Passo 3: Selecione o arquivo Excel de referência",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        btn_excel_frame = ctk.CTkFrame(frame_excel, fg_color="transparent")
        btn_excel_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(btn_excel_frame, text="📄 Selecionar Excel", 
                      command=self.selecionar_excel,
                      fg_color="#27ae60", hover_color="#2ecc71").pack(side="left", padx=5)
        
        self.lbl_excel = ctk.CTkLabel(btn_excel_frame, text="Nenhum arquivo selecionado",
                                      text_color="gray")
        self.lbl_excel.pack(side="left", padx=10)
        
        # ========== ÁREA DE STATUS ==========
        frame_status = ctk.CTkFrame(container, fg_color="#1a1a1a")
        frame_status.pack(fill="x", pady=10)
        
        self.lbl_status = ctk.CTkLabel(frame_status, text="Aguardando seleções...",
                                       font=("Roboto", 14), text_color="#f39c12")
        self.lbl_status.pack(pady=15)
        
        # ========== BOTÃO AUDITAR ==========
        self.btn_auditar = ctk.CTkButton(container, text="⚡ INICIAR AUDITORIA",
                                         command=self.iniciar_auditoria,
                                         font=("Roboto", 16, "bold"),
                                         height=50,
                                         fg_color="#e74c3c", hover_color="#c0392b",
                                         state="disabled")
        self.btn_auditar.pack(pady=20, fill="x")
        
        # ========== ÁREA DE RESULTADOS ==========
        frame_resultados = ctk.CTkFrame(container)
        frame_resultados.pack(fill="both", expand=True, pady=10)
        
        ctk.CTkLabel(frame_resultados, text="📋 Resultados da Auditoria",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        self.text_resultados = ctk.CTkTextbox(frame_resultados, height=200,
                                              font=("Consolas", 11))
        self.text_resultados.pack(fill="both", expand=True, padx=10, pady=5)
    
    # ==========================================
    # FUNÇÕES DE SELEÇÃO
    # ==========================================
    
    def selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta PAI com empresas")
        if pasta:
            self.pasta_selecionada = Path(pasta)
            self.lbl_pasta.configure(text=f"✅ {pasta}", text_color="#27ae60")
            
            # Detectar subpastas (empresas)
            self.empresas_disponiveis = [d.name for d in self.pasta_selecionada.iterdir() 
                                         if d.is_dir()]
            
            if not self.empresas_disponiveis:
                messagebox.showwarning("Aviso", "Nenhuma subpasta de empresa encontrada!")
                return
            
            self._criar_checkboxes_empresas()
            self._verificar_habilitacao()
    
    def _criar_checkboxes_empresas(self):
        # Limpar checkboxes antigos
        for cb in self.checkboxes_empresas:
            cb.destroy()
        self.checkboxes_empresas.clear()
        
        # Criar novos
        for empresa in self.empresas_disponiveis:
            var = tk.BooleanVar(value=True)
            cb = ctk.CTkCheckBox(self.scroll_empresas, text=empresa, variable=var,
                                 command=self._verificar_habilitacao)
            cb.pack(anchor="w", padx=10, pady=3)
            self.checkboxes_empresas.append((empresa, var, cb))
    
    def selecionar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o Excel de referência",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.excel_path = arquivo
            self.lbl_excel.configure(text=f"✅ {Path(arquivo).name}", text_color="#27ae60")
            
            # Carregar Excel
            try:
                self.df_excel = pd.read_excel(arquivo)
                self.lbl_status.configure(text=f"Excel carregado: {len(self.df_excel)} linhas",
                                         text_color="#27ae60")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao carregar Excel:\n{e}")
                return
            
            self._verificar_habilitacao()
    
    def _verificar_habilitacao(self):
        # Verifica se pode habilitar botão auditar
        empresas_sel = [emp for emp, var, _ in self.checkboxes_empresas if var.get()]
        
        if self.pasta_selecionada and empresas_sel and self.excel_path:
            self.btn_auditar.configure(state="normal")
            self.lbl_status.configure(text=f"Pronto! {len(empresas_sel)} empresas selecionadas",
                                     text_color="#27ae60")
        else:
            self.btn_auditar.configure(state="disabled")
    
    # ==========================================
    # LÓGICA DE AUDITORIA
    # ==========================================
    
    def iniciar_auditoria(self):
        self.btn_auditar.configure(state="disabled")
        self.text_resultados.delete("1.0", "end")
        self.text_resultados.insert("1.0", "🔄 Iniciando auditoria...\n")
        self.resultados.clear()
        
        # Empresas selecionadas
        empresas = [emp for emp, var, _ in self.checkboxes_empresas if var.get()]
        
        total_xmls = 0
        
        for empresa in empresas:
            self.text_resultados.insert("end", f"\n📂 Auditando: {empresa}\n")
            self.text_resultados.see("end")
            self.update()
            
            pasta_empresa = self.pasta_selecionada / empresa
            xmls = list(pasta_empresa.rglob("*.xml")) + list(pasta_empresa.rglob("*.XML"))
            
            self.text_resultados.insert("end", f"   Encontrados: {len(xmls)} XMLs\n")
            
            for xml_file in xmls:
                total_xmls += 1
                resultado = self._auditar_xml(xml_file, empresa)
                if resultado:
                    self.resultados.append(resultado)
        
        # Resumo
        self.text_resultados.insert("end", f"\n{'='*50}\n")
        self.text_resultados.insert("end", f"✅ Auditoria concluída!\n")
        self.text_resultados.insert("end", f"   Total de XMLs: {total_xmls}\n")
        self.text_resultados.insert("end", f"   Processados: {len(self.resultados)}\n")
        
        erros = sum(1 for r in self.resultados if r.status != "OK")
        self.text_resultados.insert("end", f"   Erros/Divergências: {erros}\n")
        
        self.btn_auditar.configure(state="normal")
        
        # Perguntar se quer gerar relatório
        if messagebox.askyesno("Concluído", "Deseja gerar o relatório em Excel?"):
            self._gerar_relatorio()
    
    def _auditar_xml(self, xml_path: Path, empresa: str) -> XMLItem:
        """Audita um XML individual"""
        tipo = detectar_tipo_xml(xml_path)
        
        if tipo == 'nfe':
            dados = parse_nfe(xml_path)
        elif tipo == 'cte':
            dados = parse_cte(xml_path)
        else:
            return XMLItem(empresa, "ERRO", xml_path.name, 0, 0, 0, 0, 0, "ERRO_PARSE")
        
        if 'erro' in dados:
            return XMLItem(empresa, "ERRO", xml_path.name, 0, 0, 0, 0, 0, "ERRO_PARSE")
        
        # Comparar com Excel (simplificado - assumindo coluna 'Numero' no Excel)
        # Na prática, você precisa fazer o match correto com suas colunas
        status = "OK"
        
        # Aqui você faria a comparação real com self.df_excel
        # Exemplo: buscar linha no Excel com mesmo número e comparar valores
        
        return XMLItem(
            empresa=empresa,
            tipo=dados['tipo'],
            numero=dados['numero'],
            valor_total=dados['valor_total'],
            icms=dados['icms'],
            pis=dados['pis'],
            cofins=dados['cofins'],
            volume=dados['volume'],
            status=status
        )
    
    def _gerar_relatorio(self):
        """Gera relatório Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria"
        
        # Cabeçalho
        headers = ["Empresa", "Tipo", "Número", "Valor Total", "ICMS", "PIS", 
                   "COFINS", "Volume", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for row, item in enumerate(self.resultados, 2):
            ws.cell(row, 1, item.empresa)
            ws.cell(row, 2, item.tipo)
            ws.cell(row, 3, item.numero)
            ws.cell(row, 4, item.valor_total)
            ws.cell(row, 5, item.icms)
            ws.cell(row, 6, item.pis)
            ws.cell(row, 7, item.cofins)
            ws.cell(row, 8, item.volume)
            ws.cell(row, 9, item.status)
            
            # Colorir status
            status_cell = ws.cell(row, 9)
            if item.status == "OK":
                status_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        # Salvar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"Auditoria_XML_{timestamp}.xlsx"
        wb.save(nome_arquivo)
        
        # ===== SALVAR CGR NO BANCO =====
        cgr_total = sum(item.valor_total for item in self.resultados)
        
        periodo = simpledialog.askstring("Período CGR", 
                                        "Digite o período (ex: Q1 2026):",
                                        initialvalue="Q1 2026")
        
        if periodo:
            from database import DatabasePMPV
            db = DatabasePMPV()
            
            if not db.buscar_consolidacao(periodo):
                db.criar_periodo_consolidacao(periodo, "Auditoria XML")
            
            db.atualizar_cgr(periodo, cgr_total)
            db.fechar()
            
            messagebox.showinfo("CGR Salvo", 
                               f"CGR: R$ {cgr_total:,.2f}\nPeríodo: {periodo}\n\n"
                               f"Relatório: {nome_arquivo}")
        else:
            messagebox.showinfo("Sucesso", f"Relatório salvo:\n{nome_arquivo}")

if __name__ == "__main__":
    app = AppAuditoriaXML()
    app.mainloop()
