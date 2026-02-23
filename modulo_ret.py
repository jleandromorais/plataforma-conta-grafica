import os
import sys
import sqlite3
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
import pdfplumber
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuração Visual
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# Taxa de câmbio EUR → BRL (ajuste conforme a cotação desejada)
TAXA_EUR_BRL = 6.0

# Diretório base da aplicação (funciona tanto em .py quanto em .exe)
_APP_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) \
           else os.path.dirname(os.path.abspath(__file__))

class SistemaRET(ctk.CTkToplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.title("Sistema RET - Processamento de PDFs")
        self.geometry("1400x900")
        
        # Dados
        self.pasta_selecionada = None
        self.dados_processados = []
        self.resultados = None
        
        self._setup_ui()
    
    def _setup_ui(self):
        # HEADER
        header = ctk.CTkFrame(self, height=80, corner_radius=0, fg_color="#1a1a2e")
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header, 
            text="Sistema RET Master", 
            font=("Roboto", 32, "bold"),
            text_color="#00d9ff"
        ).pack(side="left", padx=30, pady=20)
        
        ctk.CTkLabel(
            header, 
            text="Processamento Automatizado de Encargos", 
            font=("Roboto", 14),
            text_color="#a0a0a0"
        ).pack(side="left", padx=10)
        
        # CONTAINER PRINCIPAL
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=20, pady=20)
        
        # PAINEL ESQUERDO - Seleção
        left = ctk.CTkFrame(main, width=400, corner_radius=15)
        left.pack(side="left", fill="both", padx=(0, 10), pady=0)
        left.pack_propagate(False)
        
        ctk.CTkLabel(
            left, 
            text="Seleção de Pasta", 
            font=("Roboto", 20, "bold")
        ).pack(pady=(20, 10), padx=20, anchor="w")
        
        self.lbl_pasta = ctk.CTkLabel(
            left, 
            text="Nenhuma pasta selecionada",
            font=("Roboto", 12),
            wraplength=350,
            text_color="#808080"
        )
        self.lbl_pasta.pack(pady=10, padx=20)
        
        ctk.CTkButton(
            left,
            text="Selecionar Pasta",
            command=self.selecionar_pasta,
            height=40,
            font=("Roboto", 14, "bold"),
            fg_color="#2196F3",
            hover_color="#1976D2"
        ).pack(pady=10, padx=20, fill="x")
        
        # BOTÃO PROCESSAR
        ctk.CTkButton(
            left,
            text="PROCESSAR PDFs",
            command=self.processar,
            height=50,
            font=("Roboto", 16, "bold"),
            fg_color="#4CAF50",
            hover_color="#45a049"
        ).pack(pady=30, padx=20, fill="x")
        
        # PAINEL DIREITO - Resultados
        right = ctk.CTkFrame(main, corner_radius=15)
        right.pack(side="right", fill="both", expand=True)
        
        ctk.CTkLabel(
            right, 
            text="Resultados do Processamento", 
            font=("Roboto", 20, "bold")
        ).pack(pady=(20, 10), padx=20, anchor="w")
        
        # TABELA DE RESULTADOS
        self.tabview = ctk.CTkTabview(right)
        self.tabview.pack(fill="both", expand=True, padx=20, pady=10)
        
        self.tabview.add("Resumo")
        self.tabview.add("Dados Detalhados")
        self.tabview.add("Logs")
        self.tabview.add("Sem Valores")
        
        # ABA RESUMO
        self.frame_resumo = ctk.CTkScrollableFrame(self.tabview.tab("Resumo"))
        self.frame_resumo.pack(fill="both", expand=True)
        
        self.lbl_stats = ctk.CTkLabel(
            self.frame_resumo,
            text="Aguardando processamento...",
            font=("Roboto", 14),
            justify="left"
        )
        self.lbl_stats.pack(pady=20, padx=20, anchor="w")
        
        # ABA DADOS DETALHADOS
        self.frame_dados = ctk.CTkScrollableFrame(self.tabview.tab("Dados Detalhados"))
        self.frame_dados.pack(fill="both", expand=True)
        
        # ABA LOGS
        self.txt_logs = ctk.CTkTextbox(
            self.tabview.tab("Logs"),
            font=("Consolas", 11)
        )
        self.txt_logs.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ABA SEM VALORES (PDFs processados mas sem valores extraídos)
        self.txt_sem_valores = ctk.CTkTextbox(
            self.tabview.tab("Sem Valores"),
            font=("Consolas", 11)
        )
        self.txt_sem_valores.pack(fill="both", expand=True, padx=10, pady=10)
        self.txt_sem_valores.insert("end", "Nenhum processamento realizado.\nSelecione a pasta e clique em PROCESSAR PDFs.")
        
        # RODAPÉ
        footer = ctk.CTkFrame(self, height=100, corner_radius=15, fg_color="#1a1a2e")
        footer.pack(fill="x", padx=20, pady=(0, 20))
        footer.pack_propagate(False)
        
        # RESULTADO TOTAL
        result_frame = ctk.CTkFrame(footer, fg_color="transparent")
        result_frame.pack(side="left", padx=30, pady=20)
        
        ctk.CTkLabel(
            result_frame,
            text="TOTAL GERAL:",
            font=("Roboto", 14)
        ).pack(anchor="w")
        
        self.lbl_total = ctk.CTkLabel(
            result_frame,
            text="R$ 0,00",
            font=("Roboto", 28, "bold"),
            text_color="#00d9ff"
        )
        self.lbl_total.pack(anchor="w")
        
        # BOTÕES DE AÇÃO
        btn_frame = ctk.CTkFrame(footer, fg_color="transparent")
        btn_frame.pack(side="right", padx=30, pady=20)
        
        ctk.CTkButton(
            btn_frame,
            text="Salvar no Banco",
            command=self.salvar_db,
            width=140,
            height=35,
            fg_color="#9C27B0",
            hover_color="#7B1FA2"
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="Exportar Excel",
            command=self.exportar_excel,
            width=140,
            height=35,
            fg_color="#FF9800",
            hover_color="#F57C00"
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="💾 Salvar RET",
            command=self._salvar_ret_scg,
            width=140,
            height=35,
            fg_color="#27ae60",
            hover_color="#229954"
        ).pack(side="left", padx=5)
    
    def log(self, mensagem):
        """Adiciona mensagem ao log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.txt_logs.insert("end", f"[{timestamp}] {mensagem}\n")
        self.txt_logs.see("end")
        self.update()
    
    def selecionar_pasta(self):
        """Seleciona pasta para processamento"""
        pasta = filedialog.askdirectory(title="Selecione a Pasta Principal (RET)")
        
        if pasta:
            self.pasta_selecionada = pasta
            self.lbl_pasta.configure(
                text=f"Pasta: {pasta}",
                text_color="#4CAF50"
            )
            self.log(f"Pasta selecionada: {pasta}")
    
    def extrair_dados_pdf(self, caminho_pdf):
        """Extrai informações estruturadas do PDF"""
        dados = {
            'arquivo': os.path.basename(caminho_pdf),
            'caminho': caminho_pdf,
            'tipo_encargo': self._identificar_tipo(caminho_pdf),
            'empresa': self._extrair_empresa(caminho_pdf),
            'nota_tipo': self._extrair_tipo_nota(caminho_pdf),
            'numero_nd': '',
            'data_vencimento': '',
            'valor_total': 0.0,       # sempre em BRL
            'quantidade': 0.0,
            'valor_unitario': 0.0,
            'moeda_detectada': 'BRL', # 'BRL' ou 'EUR'
            'valores_encontrados': []  # todos já convertidos para BRL
        }
        
        try:
            with pdfplumber.open(caminho_pdf) as pdf:
                texto_completo = ''
                
                for pagina in pdf.pages:
                    texto = pagina.extract_text()
                    if texto:
                        texto_completo += texto + '\n'
                
                # Extrair número ND
                nd_match = re.search(r'ND\s*[:\-]?\s*(\d+)', texto_completo, re.IGNORECASE)
                if nd_match:
                    dados['numero_nd'] = nd_match.group(1)
                
                # Extrair data
                data_match = re.search(r'(\d{2}[/-]\d{2}[/-]\d{4})', texto_completo)
                if data_match:
                    dados['data_vencimento'] = data_match.group(1)
                
                # Extrair valores separando moeda para evitar conversão dupla
                # Padrões com símbolo explícito têm prioridade; o genérico só é
                # usado como fallback quando nenhum símbolo é encontrado.
                padrao_brl = r'R\$\s*(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)'
                padrao_eur = r'€\s*(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)'
                padrao_gen = r'(?<![€$\d,])(\d{1,3}(?:\.\d{3})+,\d{2})(?!\d)'

                def _parse(s):
                    try:
                        v = float(s.replace('.', '').replace(',', '.'))
                        return v if v > 0 else None
                    except Exception:
                        return None

                valores_brl = [v for m in re.findall(padrao_brl, texto_completo)
                               if (v := _parse(m)) is not None]
                valores_eur = [v for m in re.findall(padrao_eur, texto_completo)
                               if (v := _parse(m)) is not None]

                # Fallback genérico apenas quando nenhum símbolo foi detectado
                if not valores_brl and not valores_eur:
                    valores_brl = [v for m in re.findall(padrao_gen, texto_completo)
                                   if (v := _parse(m)) is not None]

                # Converter EUR→BRL na extração; armazenar tudo em BRL
                todos_brl = valores_brl + [v * TAXA_EUR_BRL for v in valores_eur]
                dados['valores_encontrados'] = todos_brl
                dados['moeda_detectada']     = 'EUR' if valores_eur and not valores_brl else 'BRL'

                if todos_brl:
                    dados['valor_total'] = max(todos_brl)

                    qt_match = re.search(r'(?:QT|Quantidade)[:\s]*(\d+(?:[.,]\d+)?)',
                                         texto_completo, re.IGNORECASE)
                    if qt_match:
                        dados['quantidade'] = float(qt_match.group(1).replace(',', '.'))

                    if dados['quantidade'] > 0:
                        dados['valor_unitario'] = dados['valor_total'] / dados['quantidade']
                
        except Exception as e:
            self.log(f"Erro ao processar {caminho_pdf}: {e}")
        
        return dados
    
    def _identificar_tipo(self, caminho):
        """Identifica tipo de encargo verificando os nomes das PASTAS do caminho
        (exclui o nome do arquivo para evitar falsos positivos)."""
        partes = [p.upper() for p in Path(caminho).parts[:-1]]  # ignora o arquivo
        if any('EAT' in p for p in partes):
            return 'EAT'
        if any('PENALIDADE' in p for p in partes):
            return 'Penalidades'
        # Verifica TOP como palavra inteira ou início de componente (evita DESKTOP, LAPTOP)
        if any(re.fullmatch(r'TOP[\s_\-]?.*', p) or p == 'TOP' for p in partes):
            return 'TOP'
        return 'Outros'
    
    def _extrair_empresa(self, caminho):
        """Extrai nome da empresa do nome do arquivo"""
        nome = os.path.basename(caminho).upper()
        empresas_conhecidas = [
            'COPERGAS', 'AMBEV', 'CBA', 'CERVEJARIA', 'DEXCO', 'GERDAU',
            'INDORAMA', 'INGREDION', 'KLABIN', 'MONDELEZ', 'NISSIN', 'VETRUS',
            'M DIAS BRANCO', 'PETROBRAS', 'GALP'
        ]
        
        for empresa in empresas_conhecidas:
            if empresa in nome:
                return empresa
        
        return 'N/A'
    
    def _extrair_tipo_nota(self, caminho):
        """Identifica se é Nota Débito ou Crédito.
        Tokeniza o nome do arquivo por separadores (espaço, _, -, .) para
        verificar 'ND' e 'NC' como tokens isolados, evitando falsos positivos
        em palavras como FUNDO, AGENDA, SEGUNDA (que contêm 'ND' internamente)."""
        nome = os.path.basename(caminho).upper()
        tokens = set(re.split(r'[\s_\-\.]+', nome))

        if 'ND' in tokens or 'DEBITO' in nome or 'DÉBITO' in nome:
            return 'Débito'
        if 'NC' in tokens or 'CREDITO' in nome or 'CRÉDITO' in nome:
            return 'Crédito'
        return 'N/A'
    
    def processar(self):
        """Processa todos os PDFs da pasta selecionada"""
        if not self.pasta_selecionada:
            messagebox.showwarning("Aviso", "Selecione uma pasta primeiro!")
            return
        
        self.log("="*60)
        self.log("INICIANDO PROCESSAMENTO")
        self.log("="*60)
        
        self.dados_processados = []
        arquivos_processados = 0
        
        for raiz, _, ficheiros in os.walk(self.pasta_selecionada):
            for ficheiro in ficheiros:
                if ficheiro.lower().endswith('.pdf'):
                    caminho_completo = os.path.join(raiz, ficheiro)
                    
                    self.log(f"[PDF] Processando: {ficheiro}")
                    
                    dados_pdf = self.extrair_dados_pdf(caminho_completo)
                    
                    if dados_pdf['valores_encontrados']:
                        self.dados_processados.append(dados_pdf)
                        self.log(f"   [OK] {len(dados_pdf['valores_encontrados'])} valores")
                    else:
                        self.dados_processados.append(dados_pdf)
                        self.log(f"   [AVISO] Sem valores")
                    
                    arquivos_processados += 1
        
        # Processar resultados
        self._mostrar_resultados(arquivos_processados)
    
    def _mostrar_resultados(self, total_arquivos):
        """Exibe resultados do processamento"""
        # Sempre atualizar a aba Sem Valores (mesmo quando nenhum foi processado)
        self._mostrar_sem_valores()
        
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Nenhum PDF foi processado! Verifique a pasta e os tipos de encargo selecionados.")
            return
        
        # Calcular estatísticas (valor_total já está em BRL desde a extração)
        total_geral = sum(d['valor_total'] for d in self.dados_processados)
        com_valores = len([d for d in self.dados_processados if d['valor_total'] > 0])

        self.lbl_total.configure(
            text=f"R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )

        resumo_tipos = {}
        for d in self.dados_processados:
            tipo = d['tipo_encargo']
            if tipo not in resumo_tipos:
                resumo_tipos[tipo] = {'count': 0, 'total': 0}
            resumo_tipos[tipo]['count'] += 1
            resumo_tipos[tipo]['total'] += d['valor_total']

        for widget in self.frame_resumo.winfo_children():
            widget.destroy()

        total_fmt = f"R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        stats_text = f"""
ESTATÍSTICAS DO PROCESSAMENTO

Total de PDFs: {total_arquivos}
PDFs com valores: {com_valores}
Valor Total (R$): {total_fmt}

RESUMO POR TIPO:
"""
        for tipo, stats in resumo_tipos.items():
            total_tipo_fmt = f"R$ {stats['total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            stats_text += f"\n{tipo}:\n"
            stats_text += f"  - Arquivos: {stats['count']}\n"
            stats_text += f"  - Total: {total_tipo_fmt}\n"
        
        ctk.CTkLabel(
            self.frame_resumo,
            text=stats_text,
            font=("Consolas", 13),
            justify="left"
        ).pack(pady=20, padx=20, anchor="w")
        
        # Atualizar aba de dados detalhados
        self._mostrar_dados_detalhados()
        
        self.log("="*60)
        self.log(f"PROCESSAMENTO CONCLUÍDO - {total_arquivos} arquivos")
        self.log("="*60)

        total_msg = f"R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        messagebox.showinfo("Sucesso", f"Processados {total_arquivos} PDFs!\nTotal: {total_msg}")
    
    def _mostrar_dados_detalhados(self):
        """Mostra tabela com dados detalhados"""
        for widget in self.frame_dados.winfo_children():
            widget.destroy()
        
        # Cabeçalho
        header = ctk.CTkFrame(self.frame_dados, fg_color="#2c3e50")
        header.pack(fill="x", pady=(0, 5))
        
        colunas = [
            ("Tipo", 80), ("Empresa", 150), ("Nota", 80),
            ("Nº", 100), ("Vencimento", 100), ("Valor Total", 120),
            ("QT", 80), ("Valor Unit.", 100)
        ]
        
        for txt, w in colunas:
            ctk.CTkLabel(
                header, 
                text=txt, 
                width=w,
                font=("Roboto", 11, "bold")
            ).pack(side="left", padx=2)
        
        total_regs = len(self.dados_processados)
        if total_regs > 500:
            ctk.CTkLabel(self.frame_dados, text=f"⚠ Exibindo 500 de {total_regs} registros.",
                         text_color="#f39c12", font=("Roboto", 11)).pack(anchor="w", padx=10)

        # Dados já em BRL (conversão feita na extração)
        for d in self.dados_processados[:500]:
            row = ctk.CTkFrame(self.frame_dados, fg_color="#34495e")
            row.pack(fill="x", pady=1)
            valores = [
                (d['tipo_encargo'],          80),
                (d['empresa'],              150),
                (d['nota_tipo'],             80),
                (d['numero_nd'],            100),
                (d['data_vencimento'],      100),
                (f"{d['valor_total']:.2f}", 120),
                (f"{d['quantidade']:.2f}",   80),
                (f"{d['valor_unitario']:.2f}", 100),
            ]
            
            for val, w in valores:
                ctk.CTkLabel(
                    row,
                    text=str(val),
                    width=w,
                    font=("Roboto", 10)
                ).pack(side="left", padx=2)
    
    def _mostrar_sem_valores(self):
        """Preenche a aba Sem Valores com os PDFs processados nos quais não foi extraído nenhum valor"""
        self.txt_sem_valores.delete("1.0", "end")
        if not self.dados_processados:
            self.txt_sem_valores.insert("end", "Nenhum processamento realizado.\nSelecione a pasta e clique em PROCESSAR PDFs.")
            return
        sem_valores = [d for d in self.dados_processados if not d.get('valores_encontrados') or d.get('valor_total', 0) == 0]
        if not sem_valores:
            self.txt_sem_valores.insert("end", "Nenhum arquivo sem valores.\n\nTodos os PDFs processados tiveram pelo menos um valor extraído.")
            return
        self.txt_sem_valores.insert("end", f"ARQUIVOS SEM VALORES ({len(sem_valores)})\n")
        self.txt_sem_valores.insert("end", "PDFs lidos nos quais não foi possível extrair valores (valor total = 0):\n")
        self.txt_sem_valores.insert("end", "=" * 60 + "\n\n")
        for i, d in enumerate(sem_valores, 1):
            arquivo = d.get("arquivo", "")
            caminho = d.get("caminho", "")
            tipo = d.get("tipo_encargo", "")
            self.txt_sem_valores.insert("end", f"{i:4}. [{tipo}] {arquivo}\n")
            self.txt_sem_valores.insert("end", f"      Nenhum valor extraído\n")
            self.txt_sem_valores.insert("end", f"      {caminho}\n\n")
        self.txt_sem_valores.see("1.0")
    
    def salvar_db(self):
        """Salva dados no banco de dados SQLite"""
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return
        
        try:
            db_path = os.path.join(_APP_DIR, 'RET_dados.db')
            conexao = sqlite3.connect(db_path)
            cursor = conexao.cursor()
            
            # Criar tabela
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS dados_ret (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tipo_encargo TEXT,
                    empresa TEXT,
                    nota_tipo TEXT,
                    numero_nd TEXT,
                    data_vencimento TEXT,
                    valor_total REAL,
                    quantidade REAL,
                    valor_unitario REAL,
                    arquivo TEXT,
                    caminho TEXT,
                    data_processamento TEXT
                )
            ''')
            
            # Inserir dados
            for d in self.dados_processados:
                cursor.execute('''
                    INSERT INTO dados_ret (
                        tipo_encargo, empresa, nota_tipo, numero_nd,
                        data_vencimento, valor_total, quantidade, valor_unitario,
                        arquivo, caminho, data_processamento
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    d['tipo_encargo'], d['empresa'], d['nota_tipo'], d['numero_nd'],
                    d['data_vencimento'], d['valor_total'], d['quantidade'], d['valor_unitario'],
                    d['arquivo'], d['caminho'], datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            
            conexao.commit()
            conexao.close()
            
            self.log(f"[OK] Dados salvos em: {db_path}")
            messagebox.showinfo("Sucesso", f"Dados salvos no banco!\n{db_path}")
            
        except Exception as e:
            self.log(f"[ERRO] Falha ao salvar: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar: {e}")
    
    def exportar_excel(self):
        """Exporta dados para Excel formatado"""
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return
        
        try:
            # Nome único com timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_base = f'RET_Relatorio_{timestamp}.xlsx'
            excel_path = os.path.join(self.pasta_selecionada, excel_base)
            
            # Verificar se arquivo está em uso e adicionar número se necessário
            contador = 1
            while True:
                try:
                    with open(excel_path, 'w') as f:
                        pass
                    os.remove(excel_path)
                    break
                except (PermissionError, IOError):
                    excel_path = os.path.join(self.pasta_selecionada, f'RET_Relatorio_{timestamp}_{contador}.xlsx')
                    contador += 1
                    if contador > 100:
                        excel_path = os.path.join(self.pasta_selecionada, f'RET_Relatorio_{datetime.now().strftime("%Y%m%d_%H%M%S%f")}.xlsx')
                        break
            
            # Criar DataFrame
            df = pd.DataFrame([{
                'Tipo de Encargo': d['tipo_encargo'],
                'Empresa': d['empresa'],
                'Nota Débito/Crédito': d['nota_tipo'],
                'Nº': d['numero_nd'],
                'Data Vencimento': d['data_vencimento'],
                'Valor Total': d['valor_total'],
                'QT': d['quantidade'],
                'Valor Unitário': d['valor_unitario'],
                'Arquivo': d['arquivo']
            } for d in self.dados_processados])
            
            # Criar workbook com formatação
            wb = Workbook()
            ws_dados = wb.active
            ws_dados.title = "Dados Completos"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Adicionar dados
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if r_idx == 1:  # Header
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        if c_idx in [6, 7, 8]:  # Colunas numéricas
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.00'
            
            # Ajustar larguras
            ws_dados.column_dimensions['A'].width = 20
            ws_dados.column_dimensions['B'].width = 25
            ws_dados.column_dimensions['C'].width = 20
            ws_dados.column_dimensions['D'].width = 15
            ws_dados.column_dimensions['E'].width = 18
            ws_dados.column_dimensions['F'].width = 15
            ws_dados.column_dimensions['G'].width = 12
            ws_dados.column_dimensions['H'].width = 15
            ws_dados.column_dimensions['I'].width = 40
            
            # ABA RESUMO POR TIPO
            ws_resumo = wb.create_sheet("Resumo por Tipo")
            
            resumo = df.groupby('Tipo de Encargo').agg({
                'Valor Total': 'sum',
                'QT': 'sum',
                'Arquivo': 'count'
            }).rename(columns={'Arquivo': 'Quantidade de Arquivos'}).reset_index()
            
            for r_idx, row in enumerate(dataframe_to_rows(resumo, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        if c_idx > 1:
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.00'
            
            ws_resumo.column_dimensions['A'].width = 25
            ws_resumo.column_dimensions['B'].width = 18
            ws_resumo.column_dimensions['C'].width = 15
            ws_resumo.column_dimensions['D'].width = 25
            
            # ABA RESUMO GERAL
            ws_geral = wb.create_sheet("Resumo Geral")
            
            total_geral = df['Valor Total'].sum()   # já em BRL
            total_qt = df['QT'].sum()
            total_arquivos = len(df)

            dados_geral = [
                ['RESUMO GERAL DO PROCESSAMENTO', ''],
                ['', ''],
                ['Métrica', 'Valor'],
                ['Total de PDFs Processados', total_arquivos],
                ['Quantidade Total (QT)', total_qt],
                ['Valor Total (R$)', total_geral],
                ['', ''],
                ['Data do Processamento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            for r_idx, row in enumerate(dados_geral, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_geral.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, size=16, color="1F4788")
                    elif r_idx == 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        if c_idx == 2 and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                # Mesclar célula do título só depois de escrever toda a linha (evita MergedCell read-only)
                if r_idx == 1:
                    ws_geral.merge_cells('A1:B1')
            
            ws_geral.column_dimensions['A'].width = 30
            ws_geral.column_dimensions['B'].width = 25
            
            # Salvar e fechar
            wb.save(excel_path)
            wb.close()
            
            self.log(f"[OK] Excel criado: {excel_path}")
            messagebox.showinfo("Sucesso", f"Excel exportado com sucesso!\n{excel_path}")
            
        except Exception as e:
            self.log(f"[ERRO] Falha ao exportar: {e}")
            messagebox.showerror("Erro", f"Erro ao exportar: {e}")
    
    def _salvar_ret_scg(self):
        """Salva o total RET na consolidação"""
        if not hasattr(self, 'dados_processados') or not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return
        
        from tkinter import simpledialog
        
        total_geral = sum(d['valor_total'] for d in self.dados_processados)  # já em BRL

        periodo = simpledialog.askstring("Período RET",
                                         "Digite o período (ex: Q1 2026):",
                                         initialvalue="Q1 2026")
        if periodo:
            from database import DatabasePMPV
            db = DatabasePMPV()
            if not db.buscar_consolidacao(periodo):
                db.criar_periodo_consolidacao(periodo, "RET")
            db.atualizar_ret(periodo, total_geral)
            db.fechar()

            total_fmt = f"R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            messagebox.showinfo("RET Salvo", f"RET: {total_fmt}\nPeríodo: {periodo}")

if __name__ == "__main__":
    root = ctk.CTk()
    root.withdraw()
    app = SistemaRET(root)
    app.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()
