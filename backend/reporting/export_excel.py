import os
import sys
import logging
import datetime
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Configuração de Logs
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

def load_env_credentials() -> str:
    """Lê as credenciais do ficheiro .env e retorna a string de conexão."""
    # Procura o .env na raiz do projeto (ajuste o nível de diretórios conforme necessário)
    env_path = os.path.join(os.path.dirname(__file__), '..', '..', '.env')
    load_dotenv(dotenv_path=env_path)
    
    db_user = os.getenv("PG_USER", "postgres")
    db_pass = os.getenv("PG_PASSWORD", "admin")
    db_host = os.getenv("PG_HOST", "localhost")
    db_port = os.getenv("PG_PORT", "5432")
    db_name = os.getenv("PG_DB", "plataforma")
    
    return f"postgresql+psycopg2://{db_user}:{db_pass}@{db_host}:{db_port}/{db_name}"

def fetch_data(engine, query: str) -> pd.DataFrame:
    """Executa a query e retorna um DataFrame. Lida com tabelas vazias ou erros."""
    try:
        df = pd.read_sql(query, engine)
        return df
    except SQLAlchemyError as e:
        logger.error(f"Erro ao executar a query: {e}")
        return pd.DataFrame() # Retorna DataFrame vazio para não quebrar o processo
    except Exception as e:
        logger.error(f"Erro inesperado na extração de dados: {e}")
        return pd.DataFrame()

def format_excel(file_path: str):
    """Aplica formatação avançada ao ficheiro Excel gerado."""
    try:
        wb = load_workbook(file_path)
        
        # Estilos definidos
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Cores de Status para Data Quality
        colors_status = {
            'OK': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # Verde Claro
            'FAIL': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), # Vermelho Claro
            'WARN': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarelo Claro
        }
        
        timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        footer_text = f"Relatório gerado em: {timestamp_str}"
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 1. Formatar Header (Linha 1)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 2. Aplicar Bordas e Auto-fit de Colunas
            for col_idx, col in enumerate(ws.columns, 1):
                max_length = 0
                col_letter = col[0].column_letter
                
                for cell in col:
                    if cell.value:
                        cell.border = thin_border
                        max_length = max(max_length, len(str(cell.value)))
                
                # Ajustar largura da coluna com um pequeno buffer
                ws.column_dimensions[col_letter].width = max_length + 2

            # 3. Formatação Condicional para a sheet "Data Quality"
            if sheet_name == "Data Quality":
                # Encontrar qual coluna tem o nome "status" ou "Status"
                status_col_idx = None
                for cell in ws[1]:
                    if cell.value and str(cell.value).upper() == "STATUS":
                        status_col_idx = cell.column
                        break
                
                if status_col_idx:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=status_col_idx)
                        status_val = str(cell.value).upper() if cell.value else ""
                        if status_val in colors_status:
                            # Colorir toda a linha (opcional) ou apenas a célula. Vamos colorir a célula:
                            cell.fill = colors_status[status_val]
            
            # 4. Adicionar Rodapé (Footer)
            footer_row = ws.max_row + 2
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=ws.max_column)
            footer_cell = ws.cell(row=footer_row, column=1)
            footer_cell.value = footer_text
            footer_cell.font = Font(italic=True, size=10)
            footer_cell.alignment = Alignment(horizontal="right")

        wb.save(file_path)
        logger.info(f"Formatação concluída com sucesso no ficheiro: {file_path}")
        
    except Exception as e:
        logger.error(f"Erro ao formatar Excel: {e}")

def main():
    logger.info("A iniciar processo de exportação para Excel...")
    
    # Garantir a criação da pasta reports na raiz do projeto
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    reports_dir = os.path.join(base_dir, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # Criar nome do ficheiro (YYYYMMDD)
    hoje_str = datetime.datetime.now().strftime("%Y%m%d")
    file_path = os.path.join(reports_dir, f"relatorio_geral_{hoje_str}.xlsx")
    
    # Obter credenciais e engine
    conn_string = load_env_credentials()
    try:
        engine = create_engine(conn_string)
    except Exception as e:
        logger.error(f"Falha ao criar engine de BD: {e}")
        sys.exit(1)
        
    # Dicionário de Queries (Ajustar de acordo com o esquema real)
    queries = {
        "KPIs": "SELECT periodo, pmpv, cgr_liquido_total, volume_cgf, valor_ret_total FROM marts.visao_geral ORDER BY periodo DESC LIMIT 12;",
        "Auditoria": "SELECT empresa, periodo, total_documentos, valor_bruto_total, cgr_liquido_total FROM marts.resumo_auditoria ORDER BY periodo DESC;",
        "RET": "SELECT empresa, periodo, total_encargos, valor_total_encargos FROM marts.resumo_ret ORDER BY periodo DESC;",
        "Data Quality": "SELECT dag_id, check_name, status, n_failed, run_ts FROM marts.data_quality_results WHERE run_ts >= CURRENT_DATE - INTERVAL '7 days' ORDER BY run_ts DESC;"
    }
    
    # Extrair e guardar dados
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, query in queries.items():
            logger.info(f"A extrair dados para a sheet: {sheet_name}")
            df = fetch_data(engine, query)
            
            # Se vier vazio, garantir que cria a sheet na mesma com colunas de aviso
            if df.empty:
                logger.warning(f"Tabela vazia ou erro para {sheet_name}. A criar sheet vazia.")
                df = pd.DataFrame({"Aviso": ["Sem dados disponíveis no momento."]})
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # Fechar engine e formatar
    engine.dispose()
    
    logger.info("A aplicar estilos no Excel...")
    format_excel(file_path)
    logger.info(f"Processo concluído! Ficheiro disponível em: {file_path}")

if __name__ == "__main__":
    main()