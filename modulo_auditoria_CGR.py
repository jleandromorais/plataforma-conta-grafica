import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── OCR (Tesseract) ──────────────────────────────────────────────────────────
# Detecção automática do Tesseract — mesma lógica usada em modulo_concilia_RP.py
try:
    import pdfplumber
    import pytesseract
    _TESSERACT_CANDIDATOS = [
        r'C:\Users\jose.demorais\AppData\Local\Programs\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Programs',
                     'Tesseract-OCR', 'tesseract.exe'),
    ]
    _tess = next((p for p in _TESSERACT_CANDIDATOS if os.path.exists(p)), None)
    if _tess:
        pytesseract.pytesseract.tesseract_cmd = _tess
    OCR_ATIVADO = _tess is not None
    PDF_ATIVADO = True
except ImportError:
    OCR_ATIVADO = False
    PDF_ATIVADO = False

# Configuração Visual
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# ── Fórmula regulatória CGR ──────────────────────────────────────────────────
# Validada contra a planilha "Conta Gráfica e Apuração de Custos":
#
#   G (coluna "Compra R$ s/tributos") = F − ICMS − PIS(1,65%) − COFINS(7,60%)
#   G = (F − ICMS) × (1 − 0,0165 − 0,0760)
#   G = (F − ICMS) × (1 − PIS_COFINS_CGR_RATE)
#
#   CGR = Σ G = (Σ vNF − Σ vICMS) × (1 − PIS_COFINS_CGR_RATE)
#
# Verificação Dez/25: (112.441.134,15 − 12.440.114,91) × 0,9075 = 90.750.924,96 ✓
PIS_COFINS_CGR_RATE = 0.0925   # PIS 1,65% + COFINS 7,60% = 9,25%


def calcular_cgr_liquido(valor_bruto: float, icms: float) -> float:
    """Retorna o valor líquido de tributos de um documento para o CGR.

    Formula: (vNF − vICMS) × (1 − PIS_COFINS_CGR_RATE)
    """
    return (valor_bruto - icms) * (1.0 - PIS_COFINS_CGR_RATE)

# ==========================================
# CLASSES DE DADOS
# ==========================================

class XMLItem:
    """Representa um item auditado de XML fiscal"""
    def __init__(self, empresa: str, tipo: str, numero: str, valor_total: float, 
                 icms: float, pis: float, cofins: float, volume: int, status: str,volume_total:float):
        self.empresa = empresa
        self.tipo = tipo  # NF-e ou CT-e
        self.numero = numero
        self.valor_total = valor_total
        self.icms = icms
        self.pis = pis
        self.cofins = cofins
        self.volume = volume
        self.status = status
        self.volume_total = volume_total

# ==========================================
# FUNÇÕES DE PARSE XML
# ==========================================

def parse_nfe(xml_path: Path) -> Dict:
    """Extrai dados de uma NF-e.

    Valor : total/ICMSTot/vNF  (padrão SEFAZ, igual em todas as empresas)
    Volume: soma de qCom nos itens onde uCom = M3 — campo mais confiável para
            gás natural. vol/qVol representa volumes de embalagem (caixas,
            paletes) e está incorreto para gás.
    Fallback 1 → vol/qVol  (caso não haja itens M3)
    Fallback 2 → vol/pesoL (último recurso)
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

        numero     = root.find('.//nfe:ide/nfe:nNF', ns)
        valor_tag  = root.find('.//nfe:total/nfe:ICMSTot/nfe:vNF', ns)
        # vNFTot inclui IBS/CBS (tributos 2026); se presente, usar como total
        valor_ext  = root.find('.//nfe:total/nfe:vNFTot', ns)
        icms       = root.find('.//nfe:total/nfe:ICMSTot/nfe:vICMS', ns)
        pis        = root.find('.//nfe:total/nfe:ICMSTot/nfe:vPIS', ns)
        cofins     = root.find('.//nfe:total/nfe:ICMSTot/nfe:vCOFINS', ns)

        # Volume M3: soma qCom dos itens cujo uCom é metro cúbico
        UNIDADES_M3 = {'M3', 'M³', 'M 3', 'M3.'}
        vol_m3 = 0.0
        for det in root.findall('.//nfe:det', ns):
            u_com = det.find('nfe:prod/nfe:uCom', ns)
            q_com = det.find('nfe:prod/nfe:qCom', ns)
            if u_com is not None and q_com is not None:
                if u_com.text.strip().upper() in UNIDADES_M3:
                    try:
                        vol_m3 += float(q_com.text)
                    except Exception:
                        pass

        # Fallback 1: vol/qVol
        if vol_m3 == 0.0:
            for vol in root.findall('.//nfe:vol', ns):
                q_vol = vol.find('nfe:qVol', ns)
                if q_vol is not None and q_vol.text:
                    try:
                        vol_m3 += float(q_vol.text)
                    except Exception:
                        pass

        # Fallback 2: pesoL do <vol>
        if vol_m3 == 0.0:
            peso = root.find('.//nfe:vol/nfe:pesoL', ns)
            if peso is not None and peso.text:
                try:
                    vol_m3 = float(peso.text)
                except Exception:
                    pass

        valor = (float(valor_ext.text) if valor_ext is not None else
                 float(valor_tag.text) if valor_tag is not None else 0.0)

        return {
            'tipo': 'NF-e',
            'numero': numero.text if numero is not None else 'N/A',
            'valor_total': valor,
            'icms':   float(icms.text)   if icms   is not None else 0.0,
            'pis':    float(pis.text)    if pis    is not None else 0.0,
            'cofins': float(cofins.text) if cofins is not None else 0.0,
            'volume_total': vol_m3,
            'volume': int(vol_m3),  # retrocompatibilidade
        }
    except Exception as e:
        return {'erro': str(e)}

def parse_cte(xml_path: Path) -> Dict:
    """Extrai dados de um CT-e.

    Valor : vPrest/vTPrest  (padrão SEFAZ para CT-e)
    Volume: infQ/qCarga onde cUnid='00' (M3) tem prioridade.
            Tabela cUnid CT-e: 00=M3, 01=KG, 02=TON, 03=Un, 04=L, 05=MMBTU
            Se não houver M3, usa o primeiro qCarga > 0 disponível.
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        ns = {'cte': 'http://www.portalfiscal.inf.br/cte'}

        numero      = root.find('.//cte:ide/cte:nCT', ns)
        valor_total = root.find('.//cte:vPrest/cte:vTPrest', ns)
        icms        = root.find('.//cte:ICMS//cte:vICMS', ns)
        pis         = root.find('.//cte:vPIS', ns)
        cofins      = root.find('.//cte:vCOFINS', ns)

        # Volume: prioridade para cUnid='00' (M3)
        vol_m3 = 0.0
        unid_encontrada = ''

        for infQ in root.findall('.//cte:infQ', ns):
            c_unid  = infQ.find('cte:cUnid', ns)
            q_carga = infQ.find('cte:qCarga', ns)
            if c_unid is not None and q_carga is not None and c_unid.text == '00':
                try:
                    v = float(q_carga.text)
                    if v > 0:
                        vol_m3 = v
                        unid_encontrada = 'M3'
                        break
                except Exception:
                    pass

        # Fallback: qualquer infQ com qCarga > 0
        if vol_m3 == 0.0:
            for infQ in root.findall('.//cte:infQ', ns):
                q_carga = infQ.find('cte:qCarga', ns)
                c_unid  = infQ.find('cte:cUnid', ns)
                tp_med  = infQ.find('cte:tpMed', ns)
                if q_carga is not None:
                    try:
                        v = float(q_carga.text)
                        if v > 0:
                            vol_m3 = v
                            unid_encontrada = (tp_med.text if tp_med is not None else
                                               c_unid.text if c_unid is not None else '?')
                            break
                    except Exception:
                        pass

        return {
            'tipo': 'CT-e',
            'numero': numero.text if numero is not None else 'N/A',
            'valor_total': float(valor_total.text) if valor_total is not None else 0.0,
            'icms':   float(icms.text)   if icms   is not None else 0.0,
            'pis':    float(pis.text)    if pis    is not None else 0.0,
            'cofins': float(cofins.text) if cofins is not None else 0.0,
            'volume_total': vol_m3,
            'unidade_volume': unid_encontrada,
            'volume': int(vol_m3),  # retrocompatibilidade
        }
    except Exception as e:
        return {'erro': str(e)}

def detectar_tipo_xml(xml_path: Path) -> str:
    """Detecta se é NF-e ou CT-e pelo conteúdo"""
    try:
        with open(xml_path, 'r', encoding='utf-8') as f:
            conteudo = f.read(500)  # Lê só o início
            if 'nfeProc' in conteudo or 'NFe' in conteudo:
                return 'nfe'
            elif 'cteProc' in conteudo or 'CTe' in conteudo:
                return 'cte'
    except:
        pass
    return 'desconhecido'


def parse_pdf_ocr(pdf_path: Path) -> Dict:
    """Extrai valor total de um PDF fiscal (NF-e/CT-e) quando não há XML.

    Estratégia:
      1. pdfplumber.extract_text() — rápido, funciona para PDFs com texto real.
      2. Fallback OCR via pytesseract — para PDFs escaneados/vetoriais.

    Padrões de valor buscados (prioridade decrescente):
      a) "VALOR TOTAL DA NOTA"  ou  "TOTAL DA NOTA"  + número
      b) "TOTAL"                                     + número
      c) Maior valor  R$ XX.XXX.XXX,XX  no documento
      d) Maior valor  XX.XXX.XXX,XX     (sem prefixo)

    Retorna dict compatível com parse_nfe/parse_cte:
      tipo, numero, valor_total, icms, pis, cofins, volume_total
    """
    if not PDF_ATIVADO:
        return {'erro': 'pdfplumber não instalado'}

    def _extrair_texto(path: Path) -> str:
        texto = ''
        try:
            with pdfplumber.open(str(path)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        texto += t + '\n'
            if not texto.strip() and OCR_ATIVADO:
                with pdfplumber.open(str(path)) as pdf:
                    for page in pdf.pages:
                        img = page.to_image(resolution=200).original
                        texto += pytesseract.image_to_string(img, lang='eng') + '\n'
        except Exception:
            pass
        return texto

    def _parse_brl(s: str) -> float:
        """Converte string BR (ex: '2.200.128,79') em float."""
        s = s.strip().replace(' ', '')
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        return float(s)

    texto = _extrair_texto(pdf_path)
    txt_up = texto.upper()

    valor = 0.0

    # a) VALOR TOTAL DA NOTA / TOTAL DA NOTA
    for padrao in [
        r'VALOR\s+TOTAL\s+DA\s+NOTA\s*[:\-]?\s*([\d.,]+)',
        r'TOTAL\s+DA\s+NOTA\s*[:\-]?\s*([\d.,]+)',
        r'VALOR\s+TOTAL\s*[:\-]?\s*([\d.,]+)',
        r'TOTAL\s+GERAL\s*[:\-]?\s*([\d.,]+)',
        r'TOTAL\s*[:\-]?\s*([\d.,]+)',
    ]:
        m = re.search(padrao, txt_up)
        if m:
            try:
                v = _parse_brl(m.group(1))
                if v > 1.0:
                    valor = v
                    break
            except ValueError:
                pass

    # b) Maior valor R$ XX.XXX,XX se acima não encontrou
    if valor == 0.0:
        todos_brl = []
        for m in re.finditer(r'R\$\s*([\d.,]+)', txt_up):
            try:
                todos_brl.append(_parse_brl(m.group(1)))
            except ValueError:
                pass
        if todos_brl:
            valor = max(todos_brl)

    # c) Maior número no formato XX.XXX.XXX,XX (último recurso)
    if valor == 0.0:
        todos_num = []
        for m in re.finditer(r'\b(\d{1,3}(?:\.\d{3})+,\d{2})\b', txt_up):
            try:
                todos_num.append(_parse_brl(m.group(1)))
            except ValueError:
                pass
        if todos_num:
            valor = max(todos_num)

    # Número do documento
    num = 'N/A'
    for padrao in [r'N[Oº°\.\s]*\.*\s*(\d{6,})', r'NF\s*[:\-]?\s*(\d+)',
                   r'CT-?E\s*[:\-]?\s*(\d+)', r'DANFE.*?(\d{6,})']:
        m = re.search(padrao, txt_up)
        if m:
            num = m.group(1)
            break

    # Tipo pelo nome do arquivo / conteúdo
    nome_up = pdf_path.name.upper()
    if 'CT-E' in nome_up or 'CTE' in nome_up or 'CT_E' in nome_up:
        tipo = 'CT-e'
    else:
        tipo = 'NF-e'

    return {
        'tipo': tipo,
        'numero': num,
        'valor_total': valor,
        'icms': 0.0,
        'pis': 0.0,
        'cofins': 0.0,
        'volume_total': 0.0,
        'volume': 0,
        'fonte': 'OCR',
    }

# ==========================================
# INTERFACE GRÁFICA
# ==========================================

class AppAuditoriaXML(ctk.CTkToplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.title("Auditoria XML - NF-e e CT-e")
        self.geometry("1300x850")
        
        # Variáveis de controle
        self.pasta_selecionada = None
        self.empresas_disponiveis = []
        self.empresas_selecionadas = []
        self.excel_path = None
        self.df_excel = None
        self.resultados = []

        # Somatórios — disponíveis após a auditoria para cálculos posteriores
        self.valor_total_geral  = 0.0   # soma bruta (c/tributos) de TODOS os documentos
        self.volume_total_geral = 0.0   # soma volume de TODOS os documentos
        self.valor_total_nfe    = 0.0   # soma valor somente das NF-e
        self.volume_total_nfe   = 0.0   # soma volume somente das NF-e
        self.valor_total_cte    = 0.0   # soma valor somente dos CT-e
        self.volume_total_cte   = 0.0   # soma volume somente dos CT-e
        self.cgr_liquido        = 0.0   # (Σ vNF − Σ vICMS) × (1 − 9,25%)

        # Modo de leitura: "XML" ou "PDF"
        self.modo_fonte = tk.StringVar(value="XML")

        self._setup_ui()
    
    def _setup_ui(self):
        # HEADER
        header = ctk.CTkFrame(self, height=60, corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(header, text="🔍 Auditoria XML Fiscal", 
                     font=("Roboto", 24, "bold")).pack(side="left", padx=20, pady=10)
        
        # CONTAINER PRINCIPAL
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # ========== PASSO 1: PASTA PAI ==========
        frame_pasta = ctk.CTkFrame(container)
        frame_pasta.pack(fill="x", pady=10)
        
        ctk.CTkLabel(frame_pasta, text="📁 Passo 1: Selecione a pasta PAI com subpastas de empresas",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        btn_frame = ctk.CTkFrame(frame_pasta, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(btn_frame, text="📂 Selecionar Pasta", 
                      command=self.selecionar_pasta,
                      fg_color="#3498db", hover_color="#2980b9").pack(side="left", padx=5)
        
        self.lbl_pasta = ctk.CTkLabel(btn_frame, text="Nenhuma pasta selecionada", 
                                      text_color="gray")
        self.lbl_pasta.pack(side="left", padx=10)
        
        # ========== PASSO 2: EMPRESAS ==========
        frame_empresas = ctk.CTkFrame(container)
        frame_empresas.pack(fill="x", pady=10)
        
        ctk.CTkLabel(frame_empresas, text="🏢 Passo 2: Selecione as empresas para auditar",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        # Frame com scroll para empresas
        self.scroll_empresas = ctk.CTkScrollableFrame(frame_empresas, height=120)
        self.scroll_empresas.pack(fill="x", padx=10, pady=5)
        
        self.checkboxes_empresas = []
        
        # ========== PASSO 3: EXCEL ==========
        frame_excel = ctk.CTkFrame(container)
        frame_excel.pack(fill="x", pady=10)
        
        ctk.CTkLabel(frame_excel, text="📊 Passo 3: Selecione o arquivo Excel de referência",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        btn_excel_frame = ctk.CTkFrame(frame_excel, fg_color="transparent")
        btn_excel_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkButton(btn_excel_frame, text="📄 Selecionar Excel", 
                      command=self.selecionar_excel,
                      fg_color="#27ae60", hover_color="#2ecc71").pack(side="left", padx=5)
        
        self.lbl_excel = ctk.CTkLabel(btn_excel_frame, text="Nenhum arquivo selecionado",
                                      text_color="gray")
        self.lbl_excel.pack(side="left", padx=10)
        
        # ========== PAINEL: MODO DE LEITURA + STATUS TESSERACT ==========
        frame_modo = ctk.CTkFrame(container)
        frame_modo.pack(fill="x", pady=(5, 0))

        ctk.CTkLabel(frame_modo, text="📂 Fonte dos dados:",
                     font=("Roboto", 13, "bold")).pack(side="left", padx=(12, 6), pady=8)

        rb_xml = ctk.CTkRadioButton(
            frame_modo, text="XML", variable=self.modo_fonte, value="XML",
            font=("Roboto", 13), command=self._atualizar_badge_modo,
        )
        rb_xml.pack(side="left", padx=6, pady=8)

        rb_pdf = ctk.CTkRadioButton(
            frame_modo, text="PDF (OCR)", variable=self.modo_fonte, value="PDF",
            font=("Roboto", 13), command=self._atualizar_badge_modo,
            state="normal" if PDF_ATIVADO else "disabled",
        )
        rb_pdf.pack(side="left", padx=6, pady=8)

        # Badge Tesseract
        _tess_txt   = "🟢 Tesseract ativo"  if OCR_ATIVADO else "🔴 Tesseract inativo"
        _tess_color = "#27ae60"             if OCR_ATIVADO else "#e74c3c"
        self.lbl_tesseract = ctk.CTkLabel(
            frame_modo, text=_tess_txt,
            font=("Roboto", 12, "bold"), text_color=_tess_color,
            fg_color="#2c2c2c", corner_radius=8,
        )
        self.lbl_tesseract.pack(side="right", padx=12, pady=8)

        # Badge modo atual
        self.lbl_modo_badge = ctk.CTkLabel(
            frame_modo, text="Modo: XML",
            font=("Roboto", 12, "bold"), text_color="#3498db",
            fg_color="#2c2c2c", corner_radius=8,
        )
        self.lbl_modo_badge.pack(side="right", padx=6, pady=8)

        # ========== ÁREA DE STATUS ==========
        frame_status = ctk.CTkFrame(container, fg_color="#1a1a1a")
        frame_status.pack(fill="x", pady=10)
        
        self.lbl_status = ctk.CTkLabel(frame_status, text="Aguardando seleções...",
                                       font=("Roboto", 14), text_color="#f39c12")
        self.lbl_status.pack(pady=15)
        
        # ========== BOTÕES DE AÇÃO (lado a lado) ==========
        frame_btns = ctk.CTkFrame(container, fg_color="transparent")
        frame_btns.pack(fill="x", pady=20)

        # Botão 1: Auditoria completa (precisa pasta + empresas + Excel)
        self.btn_auditar = ctk.CTkButton(frame_btns, text="⚡ AUDITORIA COMPLETA",
                                         command=self.iniciar_auditoria,
                                         font=("Roboto", 15, "bold"),
                                         height=50,
                                         fg_color="#e74c3c", hover_color="#c0392b",
                                         state="disabled")
        self.btn_auditar.pack(side="left", expand=True, fill="x", padx=(0, 8))

        # Botão 2: Só somatório (precisa só da pasta)
        self.btn_somatorio = ctk.CTkButton(frame_btns, text="📊 SÓ SOMATÓRIO\n(Valor e Volume das NFs/CTes)",
                                           command=self.calcular_somatorio,
                                           font=("Roboto", 13, "bold"),
                                           height=50,
                                           fg_color="#2980b9", hover_color="#3498db",
                                           state="disabled")
        self.btn_somatorio.pack(side="left", expand=True, fill="x", padx=(8, 0))

        # Botão 3: Salvar resultado no SCG (habilitado após auditoria ou somatório)
        self.btn_salvar_scg = ctk.CTkButton(
            container, text="💾 SALVAR RESULTADO NO SCG",
            command=self._salvar_cgr_scg,
            font=("Roboto", 13, "bold"),
            height=40,
            fg_color="#27ae60", hover_color="#1e8449",
            state="disabled",
        )
        self.btn_salvar_scg.pack(fill="x", pady=(0, 5))

        # ========== ÁREA DE RESULTADOS ==========
        frame_resultados = ctk.CTkFrame(container)
        frame_resultados.pack(fill="both", expand=True, pady=10)
        
        ctk.CTkLabel(frame_resultados, text="📋 Resultados da Auditoria",
                     font=("Roboto", 14, "bold")).pack(anchor="w", padx=10, pady=5)
        
        self.text_resultados = ctk.CTkTextbox(frame_resultados, height=200,
                                              font=("Consolas", 11))
        self.text_resultados.pack(fill="both", expand=True, padx=10, pady=5)
    
    # ==========================================
    # HELPERS DE UI
    # ==========================================

    def _atualizar_badge_modo(self):
        modo = self.modo_fonte.get()
        if modo == "XML":
            self.lbl_modo_badge.configure(text="Modo: XML", text_color="#3498db")
        else:
            cor = "#27ae60" if OCR_ATIVADO else "#e74c3c"
            aviso = "" if OCR_ATIVADO else " ⚠️ sem OCR"
            self.lbl_modo_badge.configure(text=f"Modo: PDF{aviso}", text_color=cor)

    # ==========================================
    # FUNÇÕES DE SELEÇÃO
    # ==========================================
    
    def selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta PAI com empresas")
        if pasta:
            self.pasta_selecionada = Path(pasta)
            self.lbl_pasta.configure(text=f"✅ {pasta}", text_color="#27ae60")
            
            # Detectar subpastas (empresas)
            self.empresas_disponiveis = [d.name for d in self.pasta_selecionada.iterdir() 
                                         if d.is_dir()]
            
            if not self.empresas_disponiveis:
                messagebox.showwarning("Aviso", "Nenhuma subpasta de empresa encontrada!")
                return
            
            self._criar_checkboxes_empresas()
            self._verificar_habilitacao()
    
    def _criar_checkboxes_empresas(self):
        # Limpar checkboxes antigos
        for cb in self.checkboxes_empresas:
            cb.destroy()
        self.checkboxes_empresas.clear()
        
        # Criar novos
        for empresa in self.empresas_disponiveis:
            var = tk.BooleanVar(value=True)
            cb = ctk.CTkCheckBox(self.scroll_empresas, text=empresa, variable=var,
                                 command=self._verificar_habilitacao)
            cb.pack(anchor="w", padx=10, pady=3)
            self.checkboxes_empresas.append((empresa, var, cb))
    
    def selecionar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o Excel de referência",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.excel_path = arquivo
            self.lbl_excel.configure(text=f"✅ {Path(arquivo).name}", text_color="#27ae60")
            
            # Carregar Excel
            try:
                self.df_excel = pd.read_excel(arquivo)
                self.lbl_status.configure(text=f"Excel carregado: {len(self.df_excel)} linhas",
                                         text_color="#27ae60")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao carregar Excel:\n{e}")
                return
            
            self._verificar_habilitacao()
    
    def _verificar_habilitacao(self):
        empresas_sel = [emp for emp, var, _ in self.checkboxes_empresas if var.get()]

        # Somatório: precisa só da pasta
        if self.pasta_selecionada:
            self.btn_somatorio.configure(state="normal")
        else:
            self.btn_somatorio.configure(state="disabled")

        # Auditoria completa: precisa de pasta + empresas + excel
        if self.pasta_selecionada and empresas_sel and self.excel_path:
            self.btn_auditar.configure(state="normal")
            self.lbl_status.configure(text=f"Pronto! {len(empresas_sel)} empresas selecionadas",
                                     text_color="#27ae60")
        else:
            self.btn_auditar.configure(state="disabled")
    
    # ==========================================
    # LÓGICA DE AUDITORIA
    # ==========================================
    
    def iniciar_auditoria(self):
        self.btn_auditar.configure(state="disabled")
        self.text_resultados.delete("1.0", "end")
        self.text_resultados.insert("1.0", "🔄 Iniciando auditoria...\n")
        self.resultados.clear()
        
        # Empresas selecionadas
        empresas = [emp for emp, var, _ in self.checkboxes_empresas if var.get()]
        
        total_xmls = 0
        
        total_ocr = 0

        for empresa in empresas:
            self.text_resultados.insert("end", f"\n📂 Auditando: {empresa}\n")
            self.text_resultados.see("end")
            self.update()

            pasta_empresa = self.pasta_selecionada / empresa
            # resolve() deduplica no Windows (case-insensitive): *.xml e *.XML seriam o mesmo arquivo
            xmls = list({p.resolve() for p in pasta_empresa.rglob("*.xml")})
            pdfs = list({p.resolve() for p in pasta_empresa.rglob("*.pdf")})

            usar_pdf = self.modo_fonte.get() == "PDF"
            self.text_resultados.insert(
                "end",
                f"   XMLs: {len(xmls)}  |  PDFs: {len(pdfs)}  "
                f"[{'PDF/OCR' if usar_pdf else 'XML'}]\n"
            )

            if not usar_pdf:
                # ── Modo XML ──────────────────────────────────────────────
                for xml_file in xmls:
                    total_xmls += 1
                    resultado = self._auditar_xml(xml_file, empresa)
                    if resultado:
                        self.resultados.append(resultado)
            else:
                # ── Modo PDF — processa somente PDFs, ignora XMLs ─────────
                ocr_status = "✅ OCR ativo" if OCR_ATIVADO else "⚠️ OCR inativo (só texto nativo)"
                self.text_resultados.insert(
                    "end",
                    f"   📄 {len(pdfs)} PDF(s) → PDF/OCR  [{ocr_status}]\n"
                )
                for pdf_file in pdfs:
                    total_ocr += 1
                    dados = parse_pdf_ocr(pdf_file)
                    if 'erro' in dados:
                        self.text_resultados.insert("end", f"     ❌ {pdf_file.name}: {dados['erro']}\n")
                        continue
                    vf = dados.get('valor_total', 0.0)
                    self.text_resultados.insert(
                        "end",
                        f"     📄 {pdf_file.name}  →  R$ {vf:,.2f}  [OCR]\n"
                    )
                    self.resultados.append(XMLItem(
                        empresa=empresa,
                        tipo=dados.get('tipo', 'NF-e'),
                        numero=dados.get('numero', 'N/A'),
                        valor_total=vf,
                        icms=dados.get('icms', 0.0),
                        pis=dados.get('pis', 0.0),
                        cofins=dados.get('cofins', 0.0),
                        volume=dados.get('volume', 0),
                        status='OCR',
                        volume_total=dados.get('volume_total', 0.0),
                    ))

        # Resumo
        self.text_resultados.insert("end", f"\n{'='*50}\n")
        self.text_resultados.insert("end", f"✅ Auditoria concluída!\n")
        self.text_resultados.insert("end", f"   XMLs processados : {total_xmls}\n")
        self.text_resultados.insert("end", f"   PDFs via OCR     : {total_ocr}\n")
        self.text_resultados.insert("end", f"   Total documentos : {len(self.resultados)}\n")

        erros = sum(1 for r in self.resultados if r.status not in ("OK", "OCR"))
        self.text_resultados.insert("end", f"   Erros/Divergências: {erros}\n")

        
        # ===== SOMATÓRIOS — prontos para cálculos posteriores =====
        nfes = [r for r in self.resultados if r.tipo == 'NF-e']
        ctes = [r for r in self.resultados if r.tipo == 'CT-e']

        self.valor_total_nfe    = sum(r.valor_total           for r in nfes)
        self.volume_total_nfe   = sum(getattr(r, 'volume', 0) for r in nfes)
        self.valor_total_cte    = sum(r.valor_total           for r in ctes)
        self.volume_total_cte   = sum(getattr(r, 'volume', 0) for r in ctes)
        self.valor_total_geral  = self.valor_total_nfe + self.valor_total_cte
        self.volume_total_geral = self.volume_total_nfe + self.volume_total_cte

        # CGR líquido = (Σ vNF+vTPrest − Σ vICMS) × (1 − 9,25%)
        icms_total = sum(r.icms for r in self.resultados)
        self.cgr_liquido = calcular_cgr_liquido(self.valor_total_geral, icms_total)

        aliq_pct = PIS_COFINS_CGR_RATE * 100
        self.text_resultados.insert("end", f"\n{'='*52}\n")
        self.text_resultados.insert("end", f"📄 NF-e  → R$ {self.valor_total_nfe:>18,.2f}  | Vol: {self.volume_total_nfe:,.0f}\n")
        self.text_resultados.insert("end", f"🚚 CT-e  → R$ {self.valor_total_cte:>18,.2f}  | Vol: {self.volume_total_cte:,.0f}\n")
        self.text_resultados.insert("end", f"{'─'*52}\n")
        self.text_resultados.insert("end", f"📊 Σ bruto (c/tributos)     R$ {self.valor_total_geral:>18,.2f}\n")
        self.text_resultados.insert("end", f"   − ICMS                   R$ {icms_total:>18,.2f}\n")
        self.text_resultados.insert("end", f"   × (1 − {aliq_pct:.2f}% PIS/COFINS)        × {1 - PIS_COFINS_CGR_RATE:.4f}\n")
        self.text_resultados.insert("end", f"{'─'*52}\n")
        self.text_resultados.insert("end", f"✅ CGR s/tributos           R$ {self.cgr_liquido:>18,.2f}\n")
        
        self.btn_auditar.configure(state="normal")
        self.btn_salvar_scg.configure(state="normal")

        # Perguntar se quer gerar relatório
        if messagebox.askyesno("Concluído", "Deseja gerar o relatório em Excel?"):
            self._gerar_relatorio()
    
    # ------------------------------------------------------------------
    # SOMATÓRIO RÁPIDO — não precisa de Excel nem de empresas selecionadas
    # ------------------------------------------------------------------
    def calcular_somatorio(self):
        """Lê todos os XMLs (e PDFs órfãos via OCR) da pasta e exibe o somatório."""
        if not self.pasta_selecionada:
            messagebox.showwarning("Atenção", "Selecione uma pasta com XMLs primeiro.")
            return

        pasta = Path(self.pasta_selecionada)
        # resolve() deduplica no Windows (case-insensitive): *.xml e *.XML seriam o mesmo arquivo
        xmls = list({p.resolve() for p in pasta.rglob("*.xml")})
        pdfs = list({p.resolve() for p in pasta.rglob("*.pdf")})

        usar_pdf = self.modo_fonte.get() == "PDF"

        # Modo PDF → processa apenas PDFs; Modo XML → processa apenas XMLs
        arquivos_xml  = xmls  if not usar_pdf else []
        arquivos_pdf  = pdfs  if usar_pdf     else []

        if not arquivos_xml and not arquivos_pdf:
            messagebox.showinfo("Sem arquivos", "Nenhum XML ou PDF encontrado na pasta.")
            return

        modo_label = "PDF/OCR" if usar_pdf else "XML"
        n_docs = len(arquivos_xml) + len(arquivos_pdf)
        ocr_status_txt = ("✅ OCR ativo" if OCR_ATIVADO else "⚠️ sem OCR") if usar_pdf else ""
        self.lbl_status.configure(
            text=f"Calculando somatório de {n_docs} arquivo(s) [{modo_label}] {ocr_status_txt}…",
            text_color="#f39c12"
        )
        self.update()

        val_nfe = vol_nfe = 0.0
        val_cte = vol_cte = 0.0
        val_ocr = icm_total = 0.0
        erros = 0
        ocr_detalhes = []

        # 1️⃣ XMLs (somente no modo XML)
        for xml_path in arquivos_xml:
            try:
                tipo = detectar_tipo_xml(xml_path)
                if tipo == 'nfe':
                    dados = parse_nfe(xml_path)
                    val_nfe   += float(dados.get('valor_total', 0) or 0)
                    vol_nfe   += float(dados.get('volume_total', 0) or 0)
                    icm_total += float(dados.get('icms', 0) or 0)
                elif tipo == 'cte':
                    dados = parse_cte(xml_path)
                    val_cte   += float(dados.get('valor_total', 0) or 0)
                    vol_cte   += float(dados.get('volume_total', 0) or 0)
                    icm_total += float(dados.get('icms', 0) or 0)
            except Exception:
                erros += 1

        # 2️⃣ PDFs via OCR (somente no modo PDF)
        for pdf_path in arquivos_pdf:
            try:
                dados = parse_pdf_ocr(pdf_path)
                v  = float(dados.get('valor_total', 0) or 0)
                vi = float(dados.get('icms', 0) or 0)
                val_ocr   += v
                icm_total += vi
                if dados.get('tipo') == 'CT-e':
                    val_cte += v
                else:
                    val_nfe += v
                ocr_detalhes.append((pdf_path.name, v))
            except Exception:
                erros += 1

        # Salva nos atributos para reutilização futura
        self.valor_total_nfe    = val_nfe
        self.volume_total_nfe   = vol_nfe
        self.valor_total_cte    = val_cte
        self.volume_total_cte   = vol_cte
        self.valor_total_geral  = val_nfe + val_cte
        self.volume_total_geral = vol_nfe + vol_cte

        # CGR líquido = (Σ bruto − Σ ICMS) × (1 − 9,25%)
        self.cgr_liquido = calcular_cgr_liquido(self.valor_total_geral, icm_total)

        self.lbl_status.configure(text="Somatório calculado!", text_color="#27ae60")
        self.btn_salvar_scg.configure(state="normal")

        aviso_erros = f"\n\n⚠️ {erros} arquivo(s) com erro." if erros else ""
        ocr_info = ""
        if ocr_detalhes:
            ocr_info = f"\n\n  🔍  PDFs via OCR ({len(ocr_detalhes)} doc(s)):\n"
            for nome, v in ocr_detalhes:
                ocr_info += f"       {nome[:40]:<40}  R$ {v:>15,.2f}\n"
            ocr_info += f"       {'─'*57}\n"
            ocr_info += f"       {'SUBTOTAL OCR':<40}  R$ {val_ocr:>15,.2f}\n"

        aliq_pct = PIS_COFINS_CGR_RATE * 100
        msg = (
            f"📊  SOMATÓRIO — {len(xmls)} XML(s) + {len(pdfs_sem_xml)} PDF(s) OCR{aviso_erros}\n"
            f"{'─' * 50}\n\n"
            f"  📄  NF-e\n"
            f"       Valor Total : R$ {val_nfe:>18,.2f}\n"
            f"       Volume Total:     {vol_nfe:>18,.2f}\n\n"
            f"  🚚  CT-e\n"
            f"       Valor Total : R$ {val_cte:>18,.2f}\n"
            f"       Volume Total:     {vol_cte:>18,.2f}\n\n"
            f"{'─' * 50}\n"
            f"  📊  Σ bruto (c/tributos) : R$ {self.valor_total_geral:>18,.2f}\n"
            f"  📉  − ICMS               : R$ {icm_total:>18,.2f}\n"
            f"  ×  (1 − {aliq_pct:.2f}% PIS/COFINS) :     × {1 - PIS_COFINS_CGR_RATE:.4f}\n"
            f"{'─' * 50}\n"
            f"  ✅  CGR s/tributos       : R$ {self.cgr_liquido:>18,.2f}\n"
            f"       Volume Total         :     {self.volume_total_geral:>18,.2f}"
            + ocr_info
        )
        messagebox.showinfo("Somatório Final — CGR", msg)

    def _auditar_xml(self, xml_path: Path, empresa: str) -> XMLItem:
        """Audita um XML individual"""
        tipo = detectar_tipo_xml(xml_path)
        
        if tipo == 'nfe':
            dados = parse_nfe(xml_path)
        elif tipo == 'cte':
            dados = parse_cte(xml_path)
        else:
            return XMLItem(empresa, "ERRO", xml_path.name, 0.0, 0.0, 0.0, 0.0, 0, "ERRO_PARSE", 0.0)
        
        if 'erro' in dados:
            return XMLItem(empresa, "ERRO", xml_path.name, 0.0, 0.0, 0.0, 0.0, 0, "ERRO_PARSE", 0.0)
        
        # Comparar com Excel (simplificado - assumindo coluna 'Numero' no Excel)
        # Na prática, você precisa fazer o match correto com suas colunas
        status = "OK"
        
        # Aqui você faria a comparação real com self.df_excel
        # Exemplo: buscar linha no Excel com mesmo número e comparar valores
        
        vol_total = dados.get('volume_total', float(dados.get('volume', 0)))
        return XMLItem(
            empresa=empresa,
            tipo=dados['tipo'],
            numero=dados['numero'],
            valor_total=dados['valor_total'],
            icms=dados['icms'],
            pis=dados['pis'],
            cofins=dados['cofins'],
            volume=int(vol_total),
            status=status,
            volume_total=vol_total
        )
    
    # ------------------------------------------------------------------
    def _salvar_cgr_scg(self):
        """Salva o CGR líquido s/tributos no banco de consolidação SCG.

        CGR = (Σ vNF − Σ vICMS) × (1 − PIS_COFINS_CGR_RATE)
        """
        from tkinter import simpledialog
        from database import DatabasePMPV

        cgr = getattr(self, 'cgr_liquido', 0.0)
        if cgr == 0.0:
            messagebox.showwarning("Aviso", "Execute a auditoria ou o somatório antes de salvar.")
            return

        periodo = simpledialog.askstring(
            "Salvar CGR no SCG",
            "Digite o período (ex: Dez/2025):",
            initialvalue="Dez/2025",
        )
        if not periodo:
            return

        db = DatabasePMPV()
        db.atualizar_cgr(periodo, cgr)
        rpv = db.calcular_e_salvar_rpv(periodo)
        db.fechar()

        aliq_pct = PIS_COFINS_CGR_RATE * 100
        cgr_fmt  = f"R$ {cgr:,.2f}"
        bruto_fmt = f"R$ {self.valor_total_geral:,.2f}"
        rpv_fmt  = f"R$ {rpv:,.2f}"
        messagebox.showinfo(
            "CGR Salvo ✅",
            f"Período        : {periodo}\n"
            f"Σ bruto (c/trib): {bruto_fmt}\n"
            f"× (1 − {aliq_pct:.2f}%): aplicado PIS/COFINS\n"
            f"CGR s/tributos : {cgr_fmt}\n"
            f"RPV = CGR − CGF: {rpv_fmt}\n\n"
            f"Acesse o módulo SCG para ver o resultado final.",
        )

    # ==========================================
    # GETTERS — use estes em outros cálculos
    # ==========================================

    def obter_totais(self) -> Tuple[float, float]:
        """Retorna (valor_total_geral, volume_total_geral)."""
        return (self.valor_total_geral, self.volume_total_geral)

    def obter_valor_total(self) -> float:
        """Somatório do VALOR de todos os documentos (NF-e + CT-e)."""
        return self.valor_total_geral

    def obter_volume_total(self) -> float:
        """Somatório do VOLUME de todos os documentos (NF-e + CT-e)."""
        return self.volume_total_geral

    def obter_totais_nfe(self) -> Tuple[float, float]:
        """Retorna (valor_total_nfe, volume_total_nfe) — só NF-e."""
        return (self.valor_total_nfe, self.volume_total_nfe)

    def obter_totais_cte(self) -> Tuple[float, float]:
        """Retorna (valor_total_cte, volume_total_cte) — só CT-e."""
        return (self.valor_total_cte, self.volume_total_cte)

    def _gerar_relatorio(self):
        """Gera relatório Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria"
        
        # Cabeçalho
        headers = ["Empresa", "Tipo", "Número", "Valor Total", "ICMS", "PIS", 
                   "COFINS", "Volume", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for row, item in enumerate(self.resultados, 2):
            ws.cell(row, 1, item.empresa)
            ws.cell(row, 2, item.tipo)
            ws.cell(row, 3, item.numero)
            ws.cell(row, 4, item.valor_total)
            ws.cell(row, 5, item.icms)
            ws.cell(row, 6, item.pis)
            ws.cell(row, 7, item.cofins)
            ws.cell(row, 8, item.volume)
            ws.cell(row, 9, item.status)
            
            # Colorir status
            status_cell = ws.cell(row, 9)
            if item.status == "OK":
                status_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        # Salvar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"Auditoria_XML_{timestamp}.xlsx"
        wb.save(nome_arquivo)
        
        # ===== SALVAR CGR NO BANCO =====
        cgr_total = sum(item.valor_total for item in self.resultados)
        
        periodo = simpledialog.askstring("Período CGR", 
                                        "Digite o período (ex: Q1 2026):",
                                        initialvalue="Q1 2026")
        
        if periodo:
            from database import DatabasePMPV
            db = DatabasePMPV()
            
            if not db.buscar_consolidacao(periodo):
                db.criar_periodo_consolidacao(periodo, "Auditoria XML")
            
            db.atualizar_cgr(periodo, cgr_total)
            db.fechar()
            
            messagebox.showinfo("CGR Salvo", 
                               f"CGR: R$ {cgr_total:,.2f}\nPeríodo: {periodo}\n\n"
                               f"Relatório: {nome_arquivo}")
        else:
            messagebox.showinfo("Sucesso", f"Relatório salvo:\n{nome_arquivo}")

if __name__ == "__main__":
    app = AppAuditoriaXML()
    app.mainloop()
