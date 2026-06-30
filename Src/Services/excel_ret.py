import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

from Src.common.periodos import TRIMESTRES_FISCAIS, MESES_ABREVS as _MESES_BR

# Converte TRIMESTRES_FISCAIS (índices 0-based) para números de mês 1-based
_TRIMESTRES: list[tuple[str, list[int]]] = [
    (nome, [(idx % 12) + 1 for idx in indices])
    for nome, indices in TRIMESTRES_FISCAIS.items()
]


def _mes_abrev_para_num(abrev: str) -> int:
    """'Jan' → 1, 'Fev' → 2, …  Retorna 0 se não reconhecido."""
    try:
        return _MESES_BR.index(abrev) + 1
    except ValueError:
        return 0


def _trimestre_do_mes(mes_num: int) -> str:
    """Retorna o nome do trimestre ('Fev - Abr') dado o número do mês (1-12)."""
    for nome, meses in _TRIMESTRES:
        if mes_num in meses:
            return nome
    return "Sem Trimestre"


def _chave_trimestre(tri_nome: str) -> int:
    """Ordenação dos trimestres: Nov-Jan=0, Fev-Abr=1, Mai-Jul=2, Ago-Out=3."""
    for i, (nome, _) in enumerate(_TRIMESTRES):
        if nome == tri_nome:
            return i
    return 99


class ExcelRET:
    @staticmethod
    def gerar_relatorio_completo(dados_processados, calc_ret, excel_path):
        df = pd.DataFrame([{
            'Tipo de Encargo': d['tipo_encargo'],
            'Empresa': d['empresa'],
            'Nota Débito/Crédito': d['nota_tipo'],
            'Nº': d['numero_nd'],
            'Data Vencimento': d['data_vencimento'],
            'Valor Total': d['valor_total'],
            'QT': d['quantidade'],
            'Valor Unitário': d['valor_unitario'],
            'Arquivo': d['arquivo']
        } for d in dados_processados])
        
        wb = Workbook()
        ws_dados = wb.active
        ws_dados.title = "Dados Completos"
        
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if r_idx == 1:  
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    if c_idx in [6, 7, 8]:  
                        if isinstance(value, (int, float)): cell.number_format = '#,##0.00'
        
        ws_dados.column_dimensions['A'].width = 20
        ws_dados.column_dimensions['B'].width = 25
        ws_dados.column_dimensions['C'].width = 20
        ws_dados.column_dimensions['D'].width = 15
        ws_dados.column_dimensions['E'].width = 18
        ws_dados.column_dimensions['F'].width = 15
        ws_dados.column_dimensions['G'].width = 12
        ws_dados.column_dimensions['H'].width = 15
        ws_dados.column_dimensions['I'].width = 40
        
        # ABA RESUMO POR TIPO
        ws_resumo = wb.create_sheet("Resumo por Tipo")
        resumo = df.groupby('Tipo de Encargo').agg({
            'Valor Total': 'sum',
            'QT': 'sum',
            'Arquivo': 'count'
        }).rename(columns={'Arquivo': 'Quantidade de Arquivos'}).reset_index()
        
        for r_idx, row in enumerate(dataframe_to_rows(resumo, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    if c_idx > 1:
                        if isinstance(value, (int, float)): cell.number_format = '#,##0.00'
        
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 18
        ws_resumo.column_dimensions['C'].width = 15
        ws_resumo.column_dimensions['D'].width = 25
        
        # ABA RESUMO GERAL
        ws_geral = wb.create_sheet("Resumo Geral")

        eat_bruto_xls = calc_ret['eat_bruto']
        pis_rate      = calc_ret['pis_cofins_rate']
        ec_ret_xls    = calc_ret['ret']
        total_qt      = df['QT'].sum()
        total_arqs    = len(df)

        ret_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        ret_font = Font(bold=True, color="FFFFFF", size=12)

        dados_geral = [
            ['RESUMO GERAL DO PROCESSAMENTO', ''],
            ['', ''], 
            ['Métrica', 'Valor'], 
            ['Total de PDFs Processados', total_arqs], 
            ['Quantidade Total (QT)', total_qt],
            ['', ''],
            ['Σ EAT bruto (R$)', eat_bruto_xls],
            [f'× (1 − {pis_rate*100:.2f}% PIS/COFINS)', 1.0 - pis_rate],
            ['EC = RET  (R$)', ec_ret_xls], 
            ['', ''],
            ['Data do Processamento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')] 
        ]

        for r_idx, row in enumerate(dados_geral, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_geral.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if r_idx == 1:
                    cell.font = Font(bold=True, size=16, color="1F4788")
                elif r_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                elif r_idx == 9:
                    cell.fill = ret_fill
                    cell.font = ret_font
                    if c_idx == 2: cell.number_format = '#,##0.000000'
                else:
                    if c_idx == 2 and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.000000'
            if r_idx == 1:
                ws_geral.merge_cells('A1:B1')

        ws_geral.column_dimensions['A'].width = 35
        ws_geral.column_dimensions['B'].width = 25

        # ABA POR MÊS
        ExcelRET._criar_aba_por_mes(wb, dados_processados)

        wb.save(excel_path)
        wb.close()

    @staticmethod
    def _criar_aba_por_mes(wb: Workbook, dados_processados: list) -> None:
        """Cria aba 'Por Mês' agrupada por trimestre (Nov-Jan, Fev-Abr, …) usando
        o mês detectado no nome das pastas (campo 'mes_ref')."""
        ws = wb.create_sheet("Por Mês")

        # ── Estilos ──────────────────────────────────────────────────────────
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        tri_fill     = PatternFill(start_color="0D2444", end_color="0D2444", fill_type="solid")
        tri_font     = Font(bold=True, color="FFFFFF", size=12)
        header_fill  = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        sub_mes_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        sub_mes_font = Font(bold=True, size=10, color="0D2444")
        sub_tri_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        sub_tri_font = Font(bold=True, size=11, color="000000")
        total_fill   = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        total_font   = Font(bold=True, color="FFFFFF", size=13)
        center   = Alignment(horizontal="center", vertical="center")
        left_al  = Alignment(horizontal="left",   vertical="center")

        def _cell(row, col, value, fill=None, font=None, fmt=None, align=None):
            c = ws.cell(row=row, column=col, value=value)
            c.border = border
            c.alignment = align or center
            if fill: c.fill = fill
            if font: c.font = font
            if fmt:  c.number_format = fmt
            return c

        def _merge_row(row, label, fill, font, val=None, qt=None, height=20):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            _cell(row, 1, label, fill=fill, font=font, align=left_al)
            ws.cell(row=row, column=2).fill = fill  # célula mesclada
            if val is not None:
                _cell(row, 3, val, fill=fill, font=font, fmt="#,##0.00")
            else:
                _cell(row, 3, "", fill=fill, font=font)
            if qt is not None:
                _cell(row, 4, qt, fill=fill, font=font, fmt="#,##0")
            else:
                _cell(row, 4, "", fill=fill, font=font)
            ws.row_dimensions[row].height = height

        # ── Cabeçalho de colunas ─────────────────────────────────────────────
        for col, h in enumerate(["Mês/Ano", "Tipo de Encargo", "Valor Total (R$)", "Quantidade"], 1):
            _cell(1, col, h, fill=header_fill, font=header_font)
        ws.row_dimensions[1].height = 22

        # ── Agrupamento: tri_ano → mes_ano → tipo → {valor, qt} ──────────────
        # tri_ano: ex. "Fev - Abr / 2026"
        # mes_ano: ex. "Fev/2026"
        grupos: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(
            lambda: {"valor": 0.0, "qt": 0.0}
        )))
        sem_mes: dict = defaultdict(lambda: {"valor": 0.0, "qt": 0.0})

        for d in dados_processados:
            mes_ref = (d.get("mes_ref") or "").strip()
            tipo = d.get("tipo_encargo") or "Outros"
            val  = float(d.get("valor_total", 0) or 0)
            qt   = float(d.get("quantidade",  0) or 0)

            if not mes_ref:
                sem_mes[tipo]["valor"] += val
                sem_mes[tipo]["qt"]    += qt
                continue

            # Extrai número do mês de "Jan/2026"
            partes = mes_ref.split("/")
            mes_num = _mes_abrev_para_num(partes[0]) if len(partes) == 2 else 0
            ano_str = partes[1] if len(partes) == 2 else "0000"

            tri_nome = _trimestre_do_mes(mes_num) if mes_num else "Sem Trimestre"
            tri_ano  = f"{tri_nome} / {ano_str}"
            grupos[tri_ano][mes_ref][tipo]["valor"] += val
            grupos[tri_ano][mes_ref][tipo]["qt"]    += qt

        # ── Ordenação dos trimestres por ano depois por ordem trimestral ───────
        def _chave_tri_ano(ta: str) -> tuple:
            # "Fev - Abr / 2026" → (2026, 1)
            parts = ta.rsplit("/", 1)
            try:
                ano = int(parts[-1].strip())
            except ValueError:
                ano = 9999
            tri = parts[0].strip() if len(parts) == 2 else ta
            return (ano, _chave_trimestre(tri))

        row = 2
        total_valor = 0.0
        total_qt    = 0.0

        for tri_ano in sorted(grupos.keys(), key=_chave_tri_ano):
            tri_nome_display = tri_ano  # ex: "Fev - Abr / 2026"

            # Cabeçalho do trimestre (linha azul-escura larga)
            _merge_row(row, f"  Trimestre: {tri_nome_display}", tri_fill, tri_font, height=24)
            row += 1

            tri_val = 0.0
            tri_qt  = 0.0

            # Meses dentro do trimestre em ordem cronológica
            def _chave_mes(m: str) -> tuple:
                p = m.split("/")
                try:
                    return (int(p[1]), _mes_abrev_para_num(p[0]))
                except Exception:
                    return (9999, 99)

            for mes_ano in sorted(grupos[tri_ano].keys(), key=_chave_mes):
                tipos = sorted(grupos[tri_ano][mes_ano].keys())
                sub_v = 0.0
                sub_q = 0.0

                for tipo in tipos:
                    v = grupos[tri_ano][mes_ano][tipo]["valor"]
                    q = grupos[tri_ano][mes_ano][tipo]["qt"]
                    _cell(row, 1, mes_ano)
                    _cell(row, 2, tipo, align=left_al)
                    _cell(row, 3, v, fmt="#,##0.00")
                    _cell(row, 4, q, fmt="#,##0")
                    sub_v += v
                    sub_q += q
                    row += 1

                # Subtotal do mês (azul claro)
                _merge_row(row, f"  Subtotal {mes_ano}", sub_mes_fill, sub_mes_font,
                           val=sub_v, qt=sub_q, height=16)
                tri_val += sub_v
                tri_qt  += sub_q
                row += 1

            # Subtotal do trimestre (cinza)
            _merge_row(row, f"  Subtotal Trimestre {tri_nome_display}",
                       sub_tri_fill, sub_tri_font, val=tri_val, qt=tri_qt, height=18)
            total_valor += tri_val
            total_qt    += tri_qt
            row += 1

        # Documentos sem mês detectado
        if sem_mes:
            _merge_row(row, "  Sem Mês Identificado", tri_fill, tri_font, height=22)
            row += 1
            sub_v = sub_q = 0.0
            for tipo in sorted(sem_mes.keys()):
                v = sem_mes[tipo]["valor"]
                q = sem_mes[tipo]["qt"]
                _cell(row, 1, "Sem Data")
                _cell(row, 2, tipo, align=left_al)
                _cell(row, 3, v, fmt="#,##0.00")
                _cell(row, 4, q, fmt="#,##0")
                sub_v += v
                sub_q += q
                row += 1
            _merge_row(row, "  Subtotal Sem Mês", sub_tri_fill, sub_tri_font,
                       val=sub_v, qt=sub_q, height=16)
            total_valor += sub_v
            total_qt    += sub_q
            row += 1

        # ── Total Geral ───────────────────────────────────────────────────────
        _merge_row(row, "  TOTAL GERAL", total_fill, total_font,
                   val=total_valor, qt=total_qt, height=24)

        # ── Larguras ─────────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15