from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta de cores ──────────────────────────────────────────────────────────
_TITULO_BG      = "0D2137"
_HEADER_BG      = "1A3A5C"
_SUBTITULO_BG   = "E8F4FD"
_NFE_PAR        = "EBF7EE"
_NFE_IMPAR      = "D5F5E3"
_CTE_PAR        = "FEF9E7"
_CTE_IMPAR      = "FDF2D0"
_OUTRO_PAR      = "EBF3FB"
_OUTRO_IMPAR    = "F7F9FA"
_TOTAL_BG       = "D4E6F1"
_RESUMO_HEADER  = "2E86AB"

_FMT_MOEDA  = 'R$ #,##0.00'
_FMT_PCT    = '0.0%'
_FMT_INT    = '#,##0'

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    s = Side(style="medium", color="1A3A5C")
    return Border(left=s, right=s, top=s, bottom=s)


class ExcelAuditoria:

    @staticmethod
    def gerar_relatorio_auditoria(resultados, nome_arquivo, cgr_total: float = 0.0, comparacao=None):
        """Gera relatório Excel completo e formatado com todos os dados auditados."""
        from Src.Services.servicos_auditoria import PIS_COFINS_RATE

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Completo"

        COLS = [
            "Empresa",
            "Tipo",
            "Nº NF / CT-e",
            "Valor Total (R$)",
            "ICMS (R$)",
            "ICMS (%)",
            "PIS (R$)",
            "COFINS (R$)",
            "Volume (m³)",
            "CGR Líquido (R$)",
            "Fonte",
        ]
        N = len(COLS)
        last_col = get_column_letter(N)

        # ── Linha 1: Título ──────────────────────────────────────────────────
        ws.row_dimensions[1].height = 36
        c = ws.cell(1, 1, "RELATÓRIO DE AUDITORIA FISCAL")
        c.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color=_TITULO_BG, end_color=_TITULO_BG, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A1:{last_col}1")

        # ── Linha 2: Subtítulo ───────────────────────────────────────────────
        ws.row_dimensions[2].height = 20
        sub_txt = (
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y  %H:%M')}     |"
            f"     Total de registros: {len(resultados)}"
        )
        c2 = ws.cell(2, 1, sub_txt)
        c2.font = Font(name="Calibri", size=10, italic=True, color="555555")
        c2.fill = PatternFill(start_color=_SUBTITULO_BG, end_color=_SUBTITULO_BG, fill_type="solid")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A2:{last_col}2")

        # ── Linha 3: vazia ───────────────────────────────────────────────────
        ws.row_dimensions[3].height = 6

        # ── Linha 4: Cabeçalhos ──────────────────────────────────────────────
        ws.row_dimensions[4].height = 28
        for col, header in enumerate(COLS, 1):
            c = ws.cell(4, col, header)
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _thin_border()

        # ── Linhas de dados (ordenadas por empresa → tipo → número) ──────────
        dados = sorted(resultados, key=lambda r: (r.empresa.lower(), r.tipo, r.numero))
        PRIMEIRA = 5

        for idx, item in enumerate(dados, PRIMEIRA):
            par = (idx - PRIMEIRA) % 2 == 0
            if item.tipo == "NF-e":
                bg = _NFE_PAR if par else _NFE_IMPAR
            elif item.tipo == "CT-e":
                bg = _CTE_PAR if par else _CTE_IMPAR
            else:
                bg = _OUTRO_PAR if par else _OUTRO_IMPAR

            fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cgr_item = (item.valor_total - item.icms) * (1.0 - PIS_COFINS_RATE)

            linha = [
                item.empresa,
                item.tipo,
                item.numero,
                item.valor_total,
                item.icms,
                item.icms_taxa,
                item.pis,
                item.cofins,
                item.volume_total,
                cgr_item,
                item.status,
            ]
            for col, val in enumerate(linha, 1):
                c = ws.cell(idx, col, val)
                c.fill = fill
                c.border = _thin_border()
                c.font = Font(name="Calibri", size=10)
                if col == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif col in (2, 3, 11):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col in (4, 5, 7, 8, 10):
                    c.number_format = _FMT_MOEDA
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif col == 6:
                    c.number_format = _FMT_PCT
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 9:
                    c.number_format = _FMT_INT
                    c.alignment = Alignment(horizontal="right", vertical="center")

        # ── Linha TOTAL ───────────────────────────────────────────────────────
        total_row = PRIMEIRA + len(dados)
        ws.row_dimensions[total_row].height = 26
        fill_tot = PatternFill(start_color=_TOTAL_BG, end_color=_TOTAL_BG, fill_type="solid")
        borda_tot = _thick_border()

        cell_lbl = ws.cell(total_row, 1, "TOTAL GERAL")
        cell_lbl.font = Font(name="Calibri", size=12, bold=True, color="0D2137")
        cell_lbl.fill = fill_tot
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
        cell_lbl.border = borda_tot
        ws.merge_cells(f"A{total_row}:C{total_row}")

        totais = {
            4: sum(r.valor_total  for r in resultados),
            5: sum(r.icms         for r in resultados),
            7: sum(r.pis          for r in resultados),
            8: sum(r.cofins       for r in resultados),
            9: sum(r.volume_total for r in resultados),
            10: cgr_total if cgr_total else sum(
                (r.valor_total - r.icms) * (1.0 - PIS_COFINS_RATE) for r in resultados
            ),
        }
        for col in range(1, N + 1):
            c = ws.cell(total_row, col)
            if col <= 3:
                pass  # já mergeado
            else:
                val = totais.get(col)
                if val is not None:
                    c.value = val
                    c.number_format = _FMT_INT if col == 9 else _FMT_MOEDA
                    c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = fill_tot
                c.font = Font(name="Calibri", size=11, bold=True, color="0D2137")
                c.border = borda_tot

        # ── Seção RESUMO POR TIPO ─────────────────────────────────────────────
        res_start = total_row + 2
        ws.row_dimensions[res_start].height = 24

        lbl_res = ws.cell(res_start, 1, "RESUMO POR TIPO DE DOCUMENTO")
        lbl_res.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        lbl_res.fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        lbl_res.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A{res_start}:F{res_start}")

        res_cols_hdr = ["Tipo", "Qtd.", "Valor Total (R$)", "ICMS Total (R$)", "CGR Líquido (R$)", "% do CGR"]
        for col, h in enumerate(res_cols_hdr, 1):
            c = ws.cell(res_start + 1, col, h)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=_RESUMO_HEADER, end_color=_RESUMO_HEADER, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()

        tipos: dict = {}
        for r in resultados:
            t = r.tipo
            if t not in tipos:
                tipos[t] = {"count": 0, "valor": 0.0, "icms": 0.0, "cgr": 0.0}
            tipos[t]["count"] += 1
            tipos[t]["valor"] += r.valor_total
            tipos[t]["icms"]  += r.icms
            tipos[t]["cgr"]   += (r.valor_total - r.icms) * (1.0 - PIS_COFINS_RATE)

        total_cgr_all = sum(v["cgr"] for v in tipos.values()) or 1.0
        for i, (tipo, vals) in enumerate(sorted(tipos.items()), 1):
            ri = res_start + 1 + i
            bg = _OUTRO_PAR if i % 2 == 0 else "FFFFFF"
            fill_r = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            row_v = [tipo, vals["count"], vals["valor"], vals["icms"], vals["cgr"], vals["cgr"] / total_cgr_all]
            for col, val in enumerate(row_v, 1):
                c = ws.cell(ri, col, val)
                c.fill = fill_r
                c.border = _thin_border()
                c.font = Font(name="Calibri", size=10)
                if col in (3, 4, 5):
                    c.number_format = _FMT_MOEDA
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif col == 6:
                    c.number_format = _FMT_PCT
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # ── Larguras de colunas ───────────────────────────────────────────────
        for col, w in enumerate([34, 8, 16, 18, 16, 10, 14, 14, 14, 20, 10], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.freeze_panes = "A5"

        if comparacao:
            ExcelAuditoria._adicionar_sheets_comparacao(wb, comparacao)

        wb.save(nome_arquivo)

    @staticmethod
    def _adicionar_sheets_comparacao(wb, comparacao):
        def _label_coluna_planilha(col_name: str) -> str:
            nome = str(col_name or "").strip()
            mapa_unnamed = {
                "Unnamed: 0": "Linha Base",
                "Unnamed: 1": "Mes",
                "Unnamed: 2": "Supridor",
                "Unnamed: 3": "Nota Fiscal",
                "Unnamed: 4": "Vol m3",
                "Unnamed: 5": "Compra c/tributo",
                "Unnamed: 6": "Compra s/tributos",
                "Unnamed: 7": "Custo",
                "Unnamed: 8": "ICMS",
                "Unnamed: 9": "PIS",
                "Unnamed: 10": "COFINS",
                "Unnamed: 11": "CSLL",
                "Unnamed: 12": "ICMS (R$)",
                "Unnamed: 13": "PIS (R$)",
                "Unnamed: 14": "COFINS (R$)",
                "Unnamed: 15": "Observacao",
            }
            return mapa_unnamed.get(nome, nome)

        def _fmt_por_coluna(nome_coluna: str) -> str | None:
            n = str(nome_coluna or "").strip().lower()
            if any(k in n for k in ["compra", "custo", "icms", "pis", "cofins", "csll", "valor"]):
                return _FMT_MOEDA
            if "vol" in n:
                return "#,##0.00"
            return None

        def _largura_por_coluna(nome_coluna: str) -> int:
            n = str(nome_coluna or "").strip().lower()
            if "nota" in n:
                return 18
            if "supridor" in n or "empresa" in n:
                return 24
            if "mes" in n or "periodo" in n:
                return 14
            if "linha" in n:
                return 12
            if any(k in n for k in ["compra", "custo", "icms", "pis", "cofins", "csll", "valor"]):
                return 16
            if "vol" in n:
                return 12
            return 15

        # ── Aba 1: Resumo da comparação ──────────────────────────────────────
        ws_resumo = wb.create_sheet("Comparacao Conta Grafica")
        ws_resumo.append(["Comparacao de notas fiscais por periodo"])
        ws_resumo.append([])
        ws_resumo.append(["Periodo", comparacao.periodo])
        ws_resumo.append(["Coluna de nota na planilha", comparacao.coluna_nota_excel])
        ws_resumo.append(["Coluna de periodo", comparacao.coluna_periodo_excel or "Nao identificada"])
        ws_resumo.append(["Total notas na nossa base (auditoria)", comparacao.total_nossa_base])
        ws_resumo.append(["Total notas na planilha (periodo)", comparacao.total_conta_grafica])
        ws_resumo.append(["Notas em ambas (confirmadas)", comparacao.qtd_em_ambas])
        ws_resumo.append(["So na auditoria (divergencia)", len(comparacao.notas_apenas_nossa)])
        ws_resumo.append(["So na planilha (divergencia)", len(comparacao.notas_apenas_conta_grafica)])
        ws_resumo.append(["Total divergencias", len(comparacao.notas_apenas_nossa) + len(comparacao.notas_apenas_conta_grafica)])
        ws_resumo.append(["Linhas sem numero na planilha", len(comparacao.linhas_sem_numero_excel)])
        ws_resumo.append(["Linhas lidas para o periodo", getattr(comparacao, "total_linhas_periodo", "N/A")])
        ws_resumo.append([])
        ws_resumo.append(["DIAGNOSTICO DE ABAS"])
        abas_todas = getattr(comparacao, "sheets_excel", [])
        abas_match = getattr(comparacao, "sheets_periodo_match", [])
        ws_resumo.append(["Abas encontradas na planilha", ", ".join(abas_todas) if abas_todas else "N/A"])
        ws_resumo.append(["Abas que corresponderam ao periodo", ", ".join(abas_match) if abas_match else "NENHUMA – arquivo pode nao ter dados deste periodo"])
        ws_resumo.append([])

        linha_inicio_avisos = ws_resumo.max_row + 1
        ws_resumo.cell(linha_inicio_avisos, 1, "Avisos / Alertas")
        avisos = comparacao.avisos or ["Sem avisos"]
        for i, aviso in enumerate(avisos, start=linha_inicio_avisos + 1):
            ws_resumo.cell(i, 1, f"  {aviso}")

        for col, largura in {"A": 44, "B": 48}.items():
            ws_resumo.column_dimensions[col].width = largura

        # ── Aba 2: Diferenças – seção 1 (Auditoria) + seção 2 (Planilha) ────────
        _COR_NOSSA = "FEF9E7"
        _COR_CONTA = "D6EAF8"
        _COR_SEC1  = "784212"   # caramelo escuro
        _COR_SEC2  = "1A5276"   # azul escuro

        ws_diff = wb.create_sheet("Diferencas")
        row_idx = 1

        def _titulo_secao(ws, row, texto, cor_bg, n_cols):
            ws.row_dimensions[row].height = 26
            fill_sec = PatternFill(start_color=cor_bg, end_color=cor_bg, fill_type="solid")
            border_sec = _thick_border()
            c = ws.cell(row, 1, texto)
            c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            c.fill = fill_sec
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = border_sec
            if n_cols > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

        def _cabecalho(ws, row, colunas, cor_bg):
            ws.row_dimensions[row].height = 22
            for col, h in enumerate(colunas, 1):
                c = ws.cell(row, col, h)
                c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color=cor_bg, end_color=cor_bg, fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = _thin_border()

        # ── Título global ─────────────────────────────────────────────────────
        ws_diff.row_dimensions[row_idx].height = 30
        c_tit = ws_diff.cell(row_idx, 1, f"DIFERENCAS DE NOTAS FISCAIS  –  Periodo: {comparacao.periodo}")
        c_tit.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        c_tit.fill = PatternFill(start_color=_TITULO_BG, end_color=_TITULO_BG, fill_type="solid")
        c_tit.alignment = Alignment(horizontal="center", vertical="center")
        ws_diff.merge_cells(f"A{row_idx}:I{row_idx}")
        row_idx += 1

        # ══════════════════════════════════════════════════════════════════════
        # SEÇÃO 1 – Notas somente na Auditoria
        # ══════════════════════════════════════════════════════════════════════
        N_NOSSA = len(comparacao.notas_apenas_nossa)
        _titulo_secao(ws_diff, row_idx,
                      f"  SOMENTE NA AUDITORIA ({N_NOSSA} nota(s))  –  "
                      "notas auditadas que NAO foram encontradas na planilha",
                      _COR_SEC1, 7)
        row_idx += 1

        COLS_A = ["Empresa", "Tipo", "N Nota", "N Normalizado", "Valor Total (R$)", "ICMS (R$)", "CGR Liquido (R$)"]
        _cabecalho(ws_diff, row_idx, COLS_A, _COR_SEC1)
        row_idx += 1

        fill_nossa = PatternFill(start_color=_COR_NOSSA, end_color=_COR_NOSSA, fill_type="solid")
        soma_valor = soma_icms = soma_cgr = 0.0

        for item in comparacao.notas_apenas_nossa:
            vt = float(item.get("valor_total") or 0.0)
            ic = float(item.get("icms") or 0.0)
            cg = float(item.get("cgr_liquido") or 0.0)
            soma_valor += vt; soma_icms += ic; soma_cgr += cg
            vals = [item.get("empresa", ""), item.get("tipo", ""), item.get("numero", ""),
                    item.get("numero_normalizado", ""), vt, ic, cg]
            for col, val in enumerate(vals, 1):
                c = ws_diff.cell(row_idx, col, val)
                c.fill = fill_nossa; c.border = _thin_border()
                c.font = Font(name="Calibri", size=10)
                if col in (5, 6, 7):
                    c.number_format = _FMT_MOEDA
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            row_idx += 1

        if comparacao.notas_apenas_nossa:
            fill_tot = PatternFill(start_color=_TOTAL_BG, end_color=_TOTAL_BG, fill_type="solid")
            borda_tot = _thick_border()
            c_tot = ws_diff.cell(row_idx, 1, "TOTAL")
            c_tot.font = Font(name="Calibri", size=10, bold=True, color="0D2137")
            c_tot.fill = fill_tot; c_tot.border = borda_tot
            c_tot.alignment = Alignment(horizontal="center", vertical="center")
            ws_diff.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            for col, val in ((5, soma_valor), (6, soma_icms), (7, soma_cgr)):
                c = ws_diff.cell(row_idx, col, val)
                c.font = Font(name="Calibri", size=10, bold=True, color="0D2137")
                c.fill = fill_tot; c.number_format = _FMT_MOEDA
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.border = borda_tot
            row_idx += 1

        row_idx += 1  # linha separadora

        # ══════════════════════════════════════════════════════════════════════
        # SEÇÃO 2 – Notas somente na Planilha
        # ══════════════════════════════════════════════════════════════════════
        N_CONTA = len(comparacao.notas_apenas_conta_grafica)

        # Descobrir colunas disponíveis nos dados da planilha
        colunas_planilha_raw: list[str] = []
        seen_cols: set[str] = set()
        for item in comparacao.notas_apenas_conta_grafica:
            for k in (item.get("dados_linha") or {}).keys():
                if k not in seen_cols:
                    seen_cols.add(k)
                    colunas_planilha_raw.append(k)

        # Mantém ordem natural dos Unnamed (0..N), que costuma refletir o layout original da planilha.
        def _ord_coluna(col: str) -> tuple[int, str]:
            nome = str(col)
            if nome.startswith("Unnamed: "):
                try:
                    return (0, f"{int(nome.split(':', 1)[1].strip()):03d}")
                except Exception:
                    return (0, nome)
            return (1, nome)

        colunas_planilha_raw = sorted(colunas_planilha_raw, key=_ord_coluna)
        colunas_planilha = [_label_coluna_planilha(c) for c in colunas_planilha_raw]

        # Se não tiver dados_linha, fallback para colunas fixas
        if not colunas_planilha:
            colunas_planilha = ["Numero", "N Normalizado", "Periodo", "Linha Excel"]
            colunas_planilha_raw = []
            usar_dados_linha = False
        else:
            usar_dados_linha = True

        n_cols_sec2 = len(colunas_planilha) + 1  # +1 para "Linha Excel"
        _titulo_secao(ws_diff, row_idx,
                      f"  SOMENTE NA PLANILHA ({N_CONTA} nota(s))  –  "
                      "notas da planilha que NAO foram encontradas na auditoria",
                      _COR_SEC2, max(7, n_cols_sec2))
        row_idx += 1

        if usar_dados_linha:
            COLS_B = list(colunas_planilha) + ["Linha Excel"]
        else:
            COLS_B = colunas_planilha
        _cabecalho(ws_diff, row_idx, COLS_B, _COR_SEC2)
        row_idx += 1

        fill_conta = PatternFill(start_color=_COR_CONTA, end_color=_COR_CONTA, fill_type="solid")

        for item in comparacao.notas_apenas_conta_grafica:
            if usar_dados_linha:
                dados = item.get("dados_linha") or {}
                linha_vals = [dados.get(k, "") for k in colunas_planilha_raw]
                linha_vals.append(item.get("linha_excel", ""))
            else:
                linha_vals = [
                    item.get("numero", ""),
                    item.get("numero_normalizado", ""),
                    item.get("periodo_linha", ""),
                    item.get("linha_excel", ""),
                ]
            for col, val in enumerate(linha_vals, 1):
                v = "" if val is None else val
                c = ws_diff.cell(row_idx, col, v)
                c.fill = fill_conta; c.border = _thin_border()
                c.font = Font(name="Calibri", size=10)
                nome_col = COLS_B[col - 1] if col - 1 < len(COLS_B) else ""
                fmt = _fmt_por_coluna(nome_col)
                if fmt and isinstance(val, (int, float)) and val != "":
                    c.number_format = fmt
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(val, (int, float)) and val != "":
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            row_idx += 1

        # Larguras
        max_cols = max(len(COLS_A), len(COLS_B) if usar_dados_linha else len(colunas_planilha))
        for col in range(1, max_cols + 1):
            nome_col = ""
            if col - 1 < len(COLS_B):
                nome_col = COLS_B[col - 1]
            w = _largura_por_coluna(nome_col)
            ws_diff.column_dimensions[get_column_letter(col)].width = w
        ws_diff.freeze_panes = "A2"

        # ── Aba 3: Linhas sem número ──────────────────────────────────────────
        ws_invalidas = wb.create_sheet("Conta sem Numero")
        ws_invalidas.append(["Linha Excel", "Coluna Nota", "Valor Original"])
        for item in comparacao.linhas_sem_numero_excel:
            ws_invalidas.append([
                item.get("linha_excel", ""),
                item.get("coluna_nota", ""),
                item.get("valor_original", ""),
            ])

        for col in range(1, ws_invalidas.max_column + 1):
            ws_invalidas.column_dimensions[get_column_letter(col)].width = 22
        for c in ws_invalidas[1]:
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()
        for row in ws_invalidas.iter_rows(min_row=2, max_row=ws_invalidas.max_row, min_col=1, max_col=ws_invalidas.max_column):
            for c in row:
                c.border = _thin_border()
        ws_invalidas.freeze_panes = "A2"