from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

class ExcelConcilia:
    @staticmethod
    def gerar_relatorio(caminho: Path, itens, layout_bonito: bool = False):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatorio"

        # Estilos base
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(bold=True, color="FFFFFF")

        thin = Side(style="thin", color="D5D8DC")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        if layout_bonito:
            ws.merge_cells("A1:F1")
            c_title = ws["A1"]
            c_title.value = "CONCILIACAO FINANCEIRA - RELATORIO"
            c_title.font = Font(bold=True, size=14, color="FFFFFF")
            c_title.fill = PatternFill("solid", fgColor="1F3A5F")
            c_title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

        ws.append(["Arquivo", "Categoria", "Valor", "Status", "Método", "Caminho"])

        header_row = 2 if layout_bonito else 1
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            if layout_bonito:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_all

        total_rec = 0.0
        total_desp = 0.0

        for i in itens:
            ws.append([i.file_name, i.category, i.amount, i.status, i.method, i.file_path])
            cell_val = ws.cell(row=ws.max_row, column=3)
            cell_val.number_format = '"R$ "#,##0.00'

            if layout_bonito:
                row = ws.max_row
                for col in range(1, 7):
                    c = ws.cell(row=row, column=col)
                    c.border = border_all
                    if col in (1, 5, 6):
                        c.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        c.alignment = Alignment(horizontal="center", vertical="center")

                # Destaques visuais por categoria e status
                if i.category == "Receita":
                    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="E9F7EF")
                elif i.category == "Despesa":
                    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FDEDEC")

                if i.status == "OK":
                    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="D5F5E3")
                elif i.status == "REVISAR":
                    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="FCF3CF")
                else:
                    ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="F5B7B1")
            
            if i.status == "OK":
                if i.category == "Receita": total_rec += i.amount
                elif i.category == "Despesa": total_desp += i.amount

        # Rodapé com totais
        ws.append([])
        ws.append(["RESUMO FINAL", "", "", "", "", ""])
        ws.append(["(+) RECEITAS", "", total_rec, "", "", ""])
        ws.append(["(-) DESPESAS", "", total_desp, "", "", ""])
        ws.append(["(=) SALDO", "", total_rec - total_desp, "", "", ""])

        if layout_bonito:
            resumo_inicio = ws.max_row - 3
            ws.merge_cells(start_row=resumo_inicio, start_column=1, end_row=resumo_inicio, end_column=6)
            c_resumo = ws.cell(row=resumo_inicio, column=1)
            c_resumo.fill = PatternFill("solid", fgColor="34495E")
            c_resumo.font = Font(bold=True, color="FFFFFF")
            c_resumo.alignment = Alignment(horizontal="center", vertical="center")

            for r in range(resumo_inicio + 1, ws.max_row + 1):
                ws.cell(row=r, column=1).font = Font(bold=True)
                ws.cell(row=r, column=3).number_format = '"R$ "#,##0.00'
                ws.cell(row=r, column=1).border = border_all
                ws.cell(row=r, column=3).border = border_all

            ws.cell(row=resumo_inicio + 1, column=1).fill = PatternFill("solid", fgColor="E8F8F5")
            ws.cell(row=resumo_inicio + 2, column=1).fill = PatternFill("solid", fgColor="FDEDEC")
            ws.cell(row=resumo_inicio + 3, column=1).fill = PatternFill("solid", fgColor="EBF5FB")
        
        # Ajuste de largura (preserva leitura mesmo sem layout bonito)
        larguras = {
            "A": 46,
            "B": 14,
            "C": 16,
            "D": 14,
            "E": 34,
            "F": 72,
        }
        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura

        if layout_bonito:
            ws.freeze_panes = f"A{header_row + 1}"
            ws.auto_filter.ref = f"A{header_row}:F{header_row}"
            ws.sheet_view.showGridLines = False

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
                for c in row:
                    if c.alignment is None:
                        c.alignment = Alignment(vertical="center")
        
        wb.save(caminho)
        return total_rec, total_desp