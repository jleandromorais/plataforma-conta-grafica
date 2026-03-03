import os
import sys
import sqlite3
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox
import pdfplumber
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from Src.infra.ocr_pdf import OCR_ENABLED, read_pdf_text

# Taxa de câmbio EUR → BRL (ajuste conforme a cotação desejada)
TAXA_EUR_BRL = 6.0

# Alíquota PIS/COFINS usada no cálculo regulatório do EC:
#   EC = Σ(pasta EAT) × (1 − PIS_COFINS_RATE)
# Validado contra os PDFs de Dezembro/2025: resultado = R$ 154.768,562025 (exato).
PIS_COFINS_RATE = 0.0925

# Diretório base da aplicação (funciona tanto em .py quanto em .exe)
_APP_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) \
           else os.path.dirname(os.path.abspath(__file__))

class SistemaRET(ctk.CTkToplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.title("Sistema RET - Processamento de PDFs")
        self.geometry("1400x900")
        
        # Dados
        self.pasta_selecionada = None
        self.dados_processados = []
        self.resultados = None
        
        self._setup_ui()
    
    def _setup_ui(self):
        # HEADER
        header = ctk.CTkFrame(self, height=80, corner_radius=0, fg_color="#1a1a2e")
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header, 
            text="Sistema RET Master", 
            font=("Roboto", 32, "bold"),
            text_color="#00d9ff"
        ).pack(side="left", padx=30, pady=20)
        
        ctk.CTkLabel(
            header, 
            text="Processamento Automatizado de Encargos", 
            font=("Roboto", 14),
            text_color="#a0a0a0"
        ).pack(side="left", padx=10)
        
        # CONTAINER PRINCIPAL
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=20, pady=20)
        
        # PAINEL ESQUERDO - Seleção
        left = ctk.CTkFrame(main, width=400, corner_radius=15)
        left.pack(side="left", fill="both", padx=(0, 10), pady=0)
        left.pack_propagate(False)
        
        ctk.CTkLabel(
            left, 
            text="Seleção de Pasta", 
            font=("Roboto", 20, "bold")
        ).pack(pady=(20, 10), padx=20, anchor="w")
        
        self.lbl_pasta = ctk.CTkLabel(
            left, 
            text="Nenhuma pasta selecionada",
            font=("Roboto", 12),
            wraplength=350,
            text_color="#808080"
        )
        self.lbl_pasta.pack(pady=10, padx=20)
        
        ctk.CTkButton(
            left,
            text="Selecionar Pasta",
            command=self.selecionar_pasta,
            height=40,
            font=("Roboto", 14, "bold"),
            fg_color="#2196F3",
            hover_color="#1976D2"
        ).pack(pady=10, padx=20, fill="x")
        
        # BOTÃO PROCESSAR
        ctk.CTkButton(
            left,
            text="PROCESSAR PDFs",
            command=self.processar,
            height=50,
            font=("Roboto", 16, "bold"),
            fg_color="#4CAF50",
            hover_color="#45a049"
        ).pack(pady=30, padx=20, fill="x")
        
        # PAINEL DIREITO - Resultados
        right = ctk.CTkFrame(main, corner_radius=15)
        right.pack(side="right", fill="both", expand=True)
        
        ctk.CTkLabel(
            right, 
            text="Resultados do Processamento", 
            font=("Roboto", 20, "bold")
        ).pack(pady=(20, 10), padx=20, anchor="w")
        
        # TABELA DE RESULTADOS
        self.tabview = ctk.CTkTabview(right)
        self.tabview.pack(fill="both", expand=True, padx=20, pady=10)
        
        self.tabview.add("Resumo")
        self.tabview.add("Dados Detalhados")
        self.tabview.add("Logs")
        self.tabview.add("Sem Valores")
        
        # ABA RESUMO
        self.frame_resumo = ctk.CTkScrollableFrame(self.tabview.tab("Resumo"))
        self.frame_resumo.pack(fill="both", expand=True)
        
        self.lbl_stats = ctk.CTkLabel(
            self.frame_resumo,
            text="Aguardando processamento...",
            font=("Roboto", 14),
            justify="left"
        )
        self.lbl_stats.pack(pady=20, padx=20, anchor="w")
        
        # ABA DADOS DETALHADOS
        self.frame_dados = ctk.CTkScrollableFrame(self.tabview.tab("Dados Detalhados"))
        self.frame_dados.pack(fill="both", expand=True)
        
        # ABA LOGS
        self.txt_logs = ctk.CTkTextbox(
            self.tabview.tab("Logs"),
            font=("Consolas", 11)
        )
        self.txt_logs.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ABA SEM VALORES (PDFs processados mas sem valores extraídos)
        self.txt_sem_valores = ctk.CTkTextbox(
            self.tabview.tab("Sem Valores"),
            font=("Consolas", 11)
        )
        self.txt_sem_valores.pack(fill="both", expand=True, padx=10, pady=10)
        self.txt_sem_valores.insert("end", "Nenhum processamento realizado.\nSelecione a pasta e clique em PROCESSAR PDFs.")
        
        # RODAPÉ
        footer = ctk.CTkFrame(self, height=100, corner_radius=15, fg_color="#1a1a2e")
        footer.pack(fill="x", padx=20, pady=(0, 20))
        footer.pack_propagate(False)
        
        # RESULTADO TOTAL
        result_frame = ctk.CTkFrame(footer, fg_color="transparent")
        result_frame.pack(side="left", padx=30, pady=20)
        
        ctk.CTkLabel(
            result_frame,
            text="TOTAL GERAL:",
            font=("Roboto", 14)
        ).pack(anchor="w")
        
        self.lbl_total = ctk.CTkLabel(
            result_frame,
            text="R$ 0,00",
            font=("Roboto", 28, "bold"),
            text_color="#00d9ff"
        )
        self.lbl_total.pack(anchor="w")
        
        # BOTÕES DE AÇÃO
        btn_frame = ctk.CTkFrame(footer, fg_color="transparent")
        btn_frame.pack(side="right", padx=30, pady=20)
        
        ctk.CTkButton(
            btn_frame,
            text="Salvar no Banco",
            command=self.salvar_db,
            width=140,
            height=35,
            fg_color="#9C27B0",
            hover_color="#7B1FA2"
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="Exportar Excel",
            command=self.exportar_excel,
            width=140,
            height=35,
            fg_color="#FF9800",
            hover_color="#F57C00"
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="💾 Salvar RET",
            command=self._salvar_ret_scg,
            width=140,
            height=35,
            fg_color="#27ae60",
            hover_color="#229954"
        ).pack(side="left", padx=5)
    
    def log(self, mensagem):
        """Adiciona mensagem ao log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.txt_logs.insert("end", f"[{timestamp}] {mensagem}\n")
        self.txt_logs.see("end")
        self.update()
    
    def selecionar_pasta(self):
        """Seleciona pasta para processamento"""
        pasta = filedialog.askdirectory(title="Selecione a Pasta Principal (RET)")
        
        if pasta:
            self.pasta_selecionada = pasta
            self.lbl_pasta.configure(
                text=f"Pasta: {pasta}",
                text_color="#4CAF50"
            )
            self.log(f"Pasta selecionada: {pasta}")
    
    def extrair_dados_pdf(self, caminho_pdf):
        """Extrai informações estruturadas do PDF"""
        dados = {
            'arquivo': os.path.basename(caminho_pdf),
            'caminho': caminho_pdf,
            'tipo_encargo': self._identificar_tipo(caminho_pdf),
            'empresa': self._extrair_empresa(caminho_pdf),
            'nota_tipo': self._extrair_tipo_nota(caminho_pdf),
            'numero_nd': '',
            'data_vencimento': '',
            'valor_total': 0.0,       # sempre em BRL
            'quantidade': 0.0,
            'valor_unitario': 0.0,
            'moeda_detectada': 'BRL', # 'BRL' ou 'EUR'
            'valores_encontrados': [], # todos já convertidos para BRL
            'periodo_doc': '',         # MM/AAAA extraído do texto do PDF
            'contrib_ec': 'OUTROS',    # classificação para fórmula EC
        }
        
        try:
            texto_completo, metodo = read_pdf_text(Path(caminho_pdf), lang="eng")
            if metodo == "OCR":
                dados["ocr_usado"] = True

                # Extrair número ND
                nd_match = re.search(r'ND\s*[:\-]?\s*(\d+)', texto_completo, re.IGNORECASE)
                if nd_match:
                    dados['numero_nd'] = nd_match.group(1)

                # Extrair data
                data_match = re.search(r'(\d{2}[/-]\d{2}[/-]\d{4})', texto_completo)
                if data_match:
                    dados['data_vencimento'] = data_match.group(1)

                # ── Extração de valores ──────────────────────────────────────
                # padrao_brl: R$ com ou sem separador de milhar
                #   Testa \d{4,} ANTES de \d{1,3} para capturar "R$139789,99"
                #   sem separador de milhar, evitando truncagem em "139".
                #   ex.: R$ 142.029,44  /  R$ 2 524,85  /  R$139789,99
                padrao_brl = r'R\$\s*((?:\d{4,}|\d{1,3}(?:[.\s]\d{3})*)(?:,\d{2})?)'
                padrao_eur = r'€\s*(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)'
                # padrao_gen: número com pelo menos um grupo de milhar (ponto/espaço)
                padrao_gen = r'(?<![€$\d,])(\d{1,3}(?:[.\s]\d{3})+,\d{2})(?!\d)'
                # padrao_small: valores pequenos sem separador de milhar
                #   ex.: 37,88  (min R$10 para evitar unit-prices)
                padrao_small = r'(?<![€$\d,.])(\d{1,5},\d{2})(?!\d)'

                def _parse(s: str) -> 'float | None':
                    try:
                        v = float(s.replace(' ', '').replace('.', '').replace(',', '.'))
                        return v if v > 0 else None
                    except Exception:
                        return None

                valores_brl = [v for m in re.findall(padrao_brl, texto_completo)
                               if (v := _parse(m)) is not None]
                valores_eur = [v for m in re.findall(padrao_eur, texto_completo)
                               if (v := _parse(m)) is not None]

                # Fallback 1: números com separador de milhar sem símbolo
                if not valores_brl and not valores_eur:
                    valores_brl = [v for m in re.findall(padrao_gen, texto_completo)
                                   if (v := _parse(m)) is not None]

                # Fallback 2: valores pequenos (ex.: 37,88 da Eneva Garanhuns)
                if not valores_brl and not valores_eur:
                    valores_brl = [v for m in re.findall(padrao_small, texto_completo)
                                   if (v := _parse(m)) is not None and v >= 10]

                # Converter EUR→BRL na extração; armazenar tudo em BRL
                todos_brl = valores_brl + [v * TAXA_EUR_BRL for v in valores_eur]
                dados['valores_encontrados'] = todos_brl
                dados['moeda_detectada']     = 'EUR' if valores_eur and not valores_brl else 'BRL'

                if todos_brl:
                    # Prefere cálculo exato (vol × taxa) para evitar arredondamento
                    calc_exato = self._extrair_calc_exato(texto_completo)
                    if calc_exato > 0:
                        dados['valor_total'] = calc_exato
                    else:
                        dados['valor_total'] = max(todos_brl)

                    qt_match = re.search(r'(?:QT|Quantidade)[:\s]*(\d+(?:[.,]\d+)?)',
                                         texto_completo, re.IGNORECASE)
                    if qt_match:
                        dados['quantidade'] = float(qt_match.group(1).replace(',', '.'))

                    if dados['quantidade'] > 0:
                        dados['valor_unitario'] = dados['valor_total'] / dados['quantidade']

                # ── Período do documento ──────────────────────────────────
                # Extrai "Novembro/2025" → "11/2025" para filtrar TOPs por período
                _MESES = {
                    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARCO': 3, 'MARÇO': 3,
                    'ABRIL': 4, 'MAIO': 5, 'JUNHO': 6, 'JULHO': 7,
                    'AGOSTO': 8, 'SETEMBRO': 9, 'OUTUBRO': 10,
                    'NOVEMBRO': 11, 'DEZEMBRO': 12,
                }
                per_match = re.search(
                    r'(?:PERIODO|PER[IÍ]ODO)\s+\d+\s+[AÀ]\s+\d+\s+DE\s+([A-Z\u00C0-\u00FF]+)/(\d{4})',
                    texto_completo.upper()
                )
                if per_match:
                    mes_nome = per_match.group(1).strip()
                    mes_num = _MESES.get(mes_nome, 0)
                    if mes_num:
                        dados['periodo_doc'] = f"{mes_num:02d}/{per_match.group(2)}"

        except Exception as e:
            self.log(f"Erro ao processar {caminho_pdf}: {e}")

        # ── Sub-classificação informativa (não usada no cálculo EC) ──────
        tipo  = dados['tipo_encargo']
        arq_u = dados['arquivo'].upper()
        if tipo == 'Penalidades (Receita)' and 'NDPFP' in arq_u:
            dados['contrib_ec'] = 'NDPFP'
        elif tipo == 'EAT' and 'OAC' in arq_u:
            dados['contrib_ec'] = 'EAT_OAC'
        elif tipo == 'Penalidades (Despesa)' and ('OAC' in arq_u or 'VARIA' in arq_u):
            dados['contrib_ec'] = 'PFP_DESP_OAC'
        elif tipo == 'EAT':
            dados['contrib_ec'] = 'EAT_ND'
        elif tipo == 'Penalidades (Despesa)':
            dados['contrib_ec'] = 'PFP_DESP_ND'
        elif tipo == 'TOP':
            dados['contrib_ec'] = 'TOP'

        return dados
    
    def _identificar_tipo(self, caminho):
        """Identifica tipo de encargo verificando os nomes das PASTAS do caminho
        (exclui o nome do arquivo para evitar falsos positivos).

        Itera das pastas mais profundas para as mais rasas para que a pasta
        específica (ex: 'TOP Não recuperável') tenha prioridade sobre a pasta
        raiz genérica (ex: 'Penalidades').

        Categorias reconhecidas (estrutura real COPERGAS/ARPE):
          - EAT                   → pasta contém 'EAT'
          - EC                    → pasta é exatamente 'EC' (começa com 'EC ')
          - TOP                   → pasta começa com 'TOP' (evita DESKTOP/LAPTOP)
          - Penalidades (Despesa) → pasta contém 'PENALIDADE' e 'DESPESA'
          - Penalidades (Receita) → pasta contém 'PENALIDADE' e 'RECEITA'
          - Penalidades           → pasta contém 'PENALIDADE'
          - Notas Fiscais         → pasta contém 'NOTAS FISCAIS' ou 'NOTA FISCAL'
        """
        partes = [p.upper() for p in Path(caminho).parts[:-1]]
        for p in reversed(partes):  # mais profunda primeiro → maior especificidade
            if 'EAT' in p:
                return 'EAT'
            if p == 'EC' or p.startswith('EC ') or p.startswith('EC_'):
                return 'EC'
            if 'PENALIDADE' in p:
                if 'DESPESA' in p:
                    return 'Penalidades (Despesa)'
                if 'RECEITA' in p:
                    return 'Penalidades (Receita)'
                return 'Penalidades'
            if re.fullmatch(r'TOP[\s_\-]?.*', p) or p == 'TOP':
                return 'TOP'
        return 'Outros'

    def _extrair_empresa(self, caminho):
        """Extrai nome da empresa do nome do arquivo ou da pasta pai."""
        nome = os.path.basename(caminho).upper()
        # Também verifica a pasta pai direta (ex: Notas Fiscais\Petrobras\arquivo.pdf)
        pasta_pai = Path(caminho).parent.name.upper()

        empresas_conhecidas = [
            # Fornecedores de gás
            'PETROBRAS', 'GALP', 'BRAVA', 'ENEVA', 'MASTERGAS',
            'PETRORECONCAVO', 'PETRO RECONCAVO', 'TAG', 'GEB',
            'ORIZON', 'VECTOR',
            # Clientes industriais
            'AMBEV', 'ALPEK', 'CBA', 'CERVEJARIA', 'DEXCO', 'FIAT',
            'GERDAU', 'GYPSUM', 'INDORAMA', 'INGREDION', 'KLABIN',
            'M DIAS BRANCO', 'MONDELEZ', 'NISSIN', 'OWENS', 'ROCA',
            'TERPHANE', 'VETRUS',
            # Própria empresa
            'COPERGAS',
        ]

        for empresa in empresas_conhecidas:
            if empresa in nome or empresa in pasta_pai:
                return empresa

        return 'N/A'

    def _extrair_calc_exato(self, texto: str) -> float:
        """Tenta extrair o produto exato VOLUME × TAXA do texto do PDF.

        Padrão típico nos NDPFPs e TOPNRECs da COPERGAS:
          '49.779,00 m³ X R$2,8532 = R$142.029,44'
          '36.674,00 X R$2,8532 = R$104.638,26'
          '50400,1988 m³ X R$ 2,7736 = R$139789,99'

        Calcula VOLUME × TAXA internamente (sem arredondar para 2 casas),
        eliminando o erro de arredondamento do PDF (ex.: diferença de ~R$0,003).
        Retorna 0.0 se o padrão não for encontrado.
        """
        # Aceita:  NUM  [qualquer unidade opcional, ex.: m³, m?, mmbtu]  X  R$NUM
        # Nota: OCR frequentemente converte m³ em m?, m3, mÂ³ — por isso a
        # unidade é tratada como "0 ou mais caracteres não-numéricos antes do X".
        m = re.search(
            r'((?:\d{1,3}(?:[.\s]\d{3})*|\d{4,})(?:,\d+)?)'   # volume
            r'\s*[^\d\sxX×R]{0,6}\s*[xX×]\s*'                  # unidade (OCR-tolerante)
            r'R\$\s*((?:\d{4,}|\d{1,3}(?:[.\s]\d{3})*)(?:[,.]\d+)?)',  # taxa
            texto,
        )
        if not m:
            return 0.0
        try:
            vol  = float(m.group(1).replace(' ', '').replace('.', '').replace(',', '.'))
            taxa = float(m.group(2).replace(' ', '').replace('.', '').replace(',', '.'))
            if vol > 0 and taxa > 0:
                return vol * taxa
        except ValueError:
            pass
        return 0.0

    def _extrair_tipo_nota(self, caminho):
        """Identifica se é Nota Débito, Crédito ou Nota Fiscal.

        Regras (por ordem de prioridade):
          1. Palavras explícitas 'DEBITO'/'DÉBITO' ou 'CREDITO'/'CRÉDITO'.
          2. Prefixo 'NDPFP' (Nota de Débito – Penalidade Falha de Programação).
          3. Sigla 'ND' seguida ou não de números (ex: ND001, ND-123, ND_PENALIDADE).
          4. Sigla 'NC' seguida ou não de números.
          5. Arquivo de Nota Fiscal (NFE, NF como palavra, CT-e, DANFE) → 'NF'.
        """
        nome = os.path.basename(caminho).upper()

        if re.search(r'D[ÉE]BITO', nome):
            return 'Débito'
        if re.search(r'CR[ÉE]DITO', nome):
            return 'Crédito'
        if 'NDPFP' in nome:
            return 'Débito'
        if re.search(r'\bND[\d\-\_]*', nome):
            return 'Débito'
        if re.search(r'\bNC[\d\-\_]*', nome):
            return 'Crédito'
        if re.search(r'\b(NFE|NF|DANFE|CT-?E)\b', nome):
            return 'NF'
        return 'N/A'
    
    def _calcular_ret(self):
        """Calcula EC e RET a partir dos documentos processados.

        Fórmula regulatória validada (ARPE/Pernambuco, Dezembro 2025):

            eat_bruto = Σ(todos os documentos na pasta EAT)
            EC        = eat_bruto × (1 − PIS_COFINS_RATE)   ← 0,0925 = 9,25 %
            EC_docs   = Σ(documentos na pasta EC, se houver)
            RET       = EC + EC_docs

        Verificação:  170.543,87 × (1 − 0,0925) = 154.768,562025  (exato ao Excel AX16)

        Returns:
            dict com chaves: eat_bruto, eat_docs, ec_docs_total, ec_docs,
                             ec, ret, pis_cofins_rate, outros_docs
        """
        eat_docs    = [d for d in self.dados_processados if d['tipo_encargo'] == 'EAT']
        ec_docs     = [d for d in self.dados_processados if d['tipo_encargo'] == 'EC']
        outros_docs = [d for d in self.dados_processados
                       if d['tipo_encargo'] not in ('EAT', 'EC')]

        eat_bruto    = sum(d['valor_total'] for d in eat_docs)
        ec_docs_total = sum(d['valor_total'] for d in ec_docs)

        ec  = eat_bruto * (1.0 - PIS_COFINS_RATE) + ec_docs_total
        ret = ec

        return {
            'eat_bruto':       eat_bruto,
            'eat_docs':        eat_docs,
            'ec_docs_total':   ec_docs_total,
            'ec_docs':         ec_docs,
            'outros_docs':     outros_docs,
            'ec':              ec,
            'ret':             ret,
            'pis_cofins_rate': PIS_COFINS_RATE,
        }

    def processar(self):
        """Processa todos os PDFs da pasta selecionada"""
        if not self.pasta_selecionada:
            messagebox.showwarning("Aviso", "Selecione uma pasta primeiro!")
            return
        
        self.log("="*60)
        self.log("INICIANDO PROCESSAMENTO")
        self.log("="*60)
        
        self.dados_processados = []
        arquivos_processados = 0
        
        ignorados_nf = 0

        for raiz, _, ficheiros in os.walk(self.pasta_selecionada):
            # Ignorar toda a árvore dentro de pastas "Notas Fiscais"
            partes_raiz = [p.upper() for p in Path(raiz).parts]
            if any('NOTAS FISCAIS' in p or 'NOTA FISCAL' in p for p in partes_raiz):
                ignorados_nf += len([f for f in ficheiros if f.lower().endswith('.pdf')])
                continue

            for ficheiro in ficheiros:
                if ficheiro.lower().endswith('.pdf'):
                    caminho_completo = os.path.join(raiz, ficheiro)

                    # Ignorar Notas Fiscais (CT-e, NF) — apenas NDs e similares entram no RET
                    tipo_nota = self._extrair_tipo_nota(caminho_completo)
                    if tipo_nota == 'NF':
                        ignorados_nf += 1
                        self.log(f"   [NF] Ignorado (Nota Fiscal): {ficheiro}")
                        continue

                    self.log(f"[PDF] Processando: {ficheiro}")

                    dados_pdf = self.extrair_dados_pdf(caminho_completo)

                    if dados_pdf['valores_encontrados']:
                        self.dados_processados.append(dados_pdf)
                        self.log(f"   [OK] {len(dados_pdf['valores_encontrados'])} valores")
                    else:
                        self.dados_processados.append(dados_pdf)
                        self.log(f"   [AVISO] Sem valores")

                    arquivos_processados += 1

        if ignorados_nf:
            self.log(f"\n[INFO] {ignorados_nf} PDF(s) em 'Notas Fiscais' ignorados (fora do escopo do RET).")
        
        # Processar resultados
        self._mostrar_resultados(arquivos_processados)
    
    def _mostrar_resultados(self, total_arquivos):
        """Exibe resultados do processamento"""
        # Sempre atualizar a aba Sem Valores (mesmo quando nenhum foi processado)
        self._mostrar_sem_valores()
        
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Nenhum PDF foi processado! Verifique a pasta e os tipos de encargo selecionados.")
            return
        
        # Calcular RET primeiro — é o valor oficial que aparece em toda a UI
        calc = self._calcular_ret()
        com_valores = len([d for d in self.dados_processados if d['valor_total'] > 0])

        # Label principal mostra EC = RET (valor já multiplicado por (1 − PIS/COFINS))
        ret_fmt2 = f"R$ {calc['ret']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        self.lbl_total.configure(text=ret_fmt2)

        resumo_tipos = {}
        for d in self.dados_processados:
            tipo = d['tipo_encargo']
            if tipo not in resumo_tipos:
                resumo_tipos[tipo] = {'count': 0, 'total': 0}
            resumo_tipos[tipo]['count'] += 1
            resumo_tipos[tipo]['total'] += d['valor_total']

        for widget in self.frame_resumo.winfo_children():
            widget.destroy()

        eat_bruto_fmt = f"R$ {calc['eat_bruto']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        stats_text = f"""
ESTATÍSTICAS DO PROCESSAMENTO

Total de PDFs: {total_arquivos}
PDFs com valores: {com_valores}
Σ EAT (bruto): {eat_bruto_fmt}
EC = RET (líquido): {ret_fmt2}

RESUMO POR TIPO:
"""
        for tipo, stats in resumo_tipos.items():
            total_tipo_fmt = f"R$ {stats['total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            stats_text += f"\n{tipo}:\n"
            stats_text += f"  - Arquivos: {stats['count']}\n"
            stats_text += f"  - Total: {total_tipo_fmt}\n"

        def brl6(v):
            """Formata com 6 casas decimais, removendo zeros à direita (mín. 2)."""
            partes = f"{v:,.6f}".split('.')
            dec = partes[1].rstrip('0')
            dec = dec if len(dec) >= 2 else dec.ljust(2, '0')
            inteiro = partes[0].replace(',', 'X').replace('.', ',').replace('X', '.')
            return f"R$ {inteiro},{dec}"

        aliq_pct = calc['pis_cofins_rate'] * 100
        ret_text = f"""
{'='*56}
CÁLCULO EC / RET  [precisão: 6 casas decimais]
{'='*56}

  Σ pasta EAT (bruto)           {brl6(calc['eat_bruto']):>22s}
  × (1 − {aliq_pct:.2f}% PIS/COFINS)            × {1 - calc['pis_cofins_rate']:.4f}
  {'─'*56}
  EC  (pasta EAT líquida)       {brl6(calc['eat_bruto'] * (1 - calc['pis_cofins_rate'])):>22s}
"""
        if calc['ec_docs_total'] > 0:
            ret_text += f"  (+) Σ pasta EC (docs.)        {brl6(calc['ec_docs_total']):>22s}\n"
            ret_text += f"  {'─'*56}\n"

        ret_text += f"""  EC  =  RET                    {brl6(calc['ret']):>22s}
{'='*56}
"""
        if calc['outros_docs']:
            ret_text += f"\n  [INFO] Documentos fora das pastas EAT/EC ({len(calc['outros_docs'])}):\n"
            for d in calc['outros_docs'][:10]:
                ret_text += f"    • {d['arquivo']}  tipo={d['tipo_encargo']}\n"
            if len(calc['outros_docs']) > 10:
                ret_text += f"    ... e mais {len(calc['outros_docs']) - 10} arquivo(s)\n"

        # Detalhe por arquivo na pasta EAT
        if calc['eat_docs']:
            ret_text += f"\n  DETALHES PASTA EAT ({len(calc['eat_docs'])} docs):\n"
            for d in calc['eat_docs']:
                ret_text += f"    • {d['arquivo']:<45s}  {brl6(d['valor_total'])}\n"

        stats_text += ret_text

        ctk.CTkLabel(
            self.frame_resumo,
            text=stats_text,
            font=("Consolas", 13),
            justify="left"
        ).pack(pady=20, padx=20, anchor="w")
        
        # Atualizar aba de dados detalhados
        self._mostrar_dados_detalhados()
        
        self.log("="*60)
        self.log(f"PROCESSAMENTO CONCLUÍDO - {total_arquivos} arquivos")
        self.log("="*60)

        messagebox.showinfo(
            "Sucesso",
            f"Processados {total_arquivos} PDFs!\n"
            f"Σ EAT (bruto): {eat_bruto_fmt}\n"
            f"EC = RET:       {ret_fmt2}"
        )
    
    def _mostrar_dados_detalhados(self):
        """Mostra tabela com dados detalhados"""
        for widget in self.frame_dados.winfo_children():
            widget.destroy()
        
        # Cabeçalho
        header = ctk.CTkFrame(self.frame_dados, fg_color="#2c3e50")
        header.pack(fill="x", pady=(0, 5))
        
        colunas = [
            ("Tipo", 80), ("Empresa", 150), ("Nota", 80),
            ("Nº", 100), ("Vencimento", 100), ("Valor Total", 120),
            ("QT", 80), ("Valor Unit.", 100)
        ]
        
        for txt, w in colunas:
            ctk.CTkLabel(
                header, 
                text=txt, 
                width=w,
                font=("Roboto", 11, "bold")
            ).pack(side="left", padx=2)
        
        total_regs = len(self.dados_processados)
        if total_regs > 500:
            ctk.CTkLabel(self.frame_dados, text=f"⚠ Exibindo 500 de {total_regs} registros.",
                         text_color="#f39c12", font=("Roboto", 11)).pack(anchor="w", padx=10)

        # Dados já em BRL (conversão feita na extração)
        for d in self.dados_processados[:500]:
            row = ctk.CTkFrame(self.frame_dados, fg_color="#34495e")
            row.pack(fill="x", pady=1)
            valores = [
                (d['tipo_encargo'],          80),
                (d['empresa'],              150),
                (d['nota_tipo'],             80),
                (d['numero_nd'],            100),
                (d['data_vencimento'],      100),
                (f"{d['valor_total']:.2f}", 120),
                (f"{d['quantidade']:.2f}",   80),
                (f"{d['valor_unitario']:.2f}", 100),
            ]
            
            for val, w in valores:
                ctk.CTkLabel(
                    row,
                    text=str(val),
                    width=w,
                    font=("Roboto", 10)
                ).pack(side="left", padx=2)
    
    def _mostrar_sem_valores(self):
        """Preenche a aba Sem Valores com os PDFs processados nos quais não foi extraído nenhum valor"""
        self.txt_sem_valores.delete("1.0", "end")
        if not self.dados_processados:
            self.txt_sem_valores.insert("end", "Nenhum processamento realizado.\nSelecione a pasta e clique em PROCESSAR PDFs.")
            return
        sem_valores = [d for d in self.dados_processados if not d.get('valores_encontrados') or d.get('valor_total', 0) == 0]
        if not sem_valores:
            self.txt_sem_valores.insert("end", "Nenhum arquivo sem valores.\n\nTodos os PDFs processados tiveram pelo menos um valor extraído.")
            return
        self.txt_sem_valores.insert("end", f"ARQUIVOS SEM VALORES ({len(sem_valores)})\n")
        self.txt_sem_valores.insert("end", "PDFs lidos nos quais não foi possível extrair valores (valor total = 0):\n")
        self.txt_sem_valores.insert("end", "=" * 60 + "\n\n")
        for i, d in enumerate(sem_valores, 1):
            arquivo = d.get("arquivo", "")
            caminho = d.get("caminho", "")
            tipo = d.get("tipo_encargo", "")
            self.txt_sem_valores.insert("end", f"{i:4}. [{tipo}] {arquivo}\n")
            self.txt_sem_valores.insert("end", f"      Nenhum valor extraído\n")
            self.txt_sem_valores.insert("end", f"      {caminho}\n\n")
        self.txt_sem_valores.see("1.0")
    
    def salvar_db(self):
        """Salva dados no banco de dados SQLite.

        O arquivo principal (RET_dados.db) é sempre mantido em _APP_DIR para
        que o programa funcione corretamente.  Após salvar, o usuário pode
        opcionalmente gerar uma cópia de backup em qualquer pasta.
        """
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return

        try:
            db_path = os.path.join(_APP_DIR, 'RET_dados.db')
            conexao = sqlite3.connect(db_path)
            cursor = conexao.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS dados_ret (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tipo_encargo TEXT,
                    empresa TEXT,
                    nota_tipo TEXT,
                    numero_nd TEXT,
                    data_vencimento TEXT,
                    valor_total REAL,
                    quantidade REAL,
                    valor_unitario REAL,
                    arquivo TEXT,
                    caminho TEXT,
                    data_processamento TEXT
                )
            ''')

            for d in self.dados_processados:
                cursor.execute('''
                    INSERT INTO dados_ret (
                        tipo_encargo, empresa, nota_tipo, numero_nd,
                        data_vencimento, valor_total, quantidade, valor_unitario,
                        arquivo, caminho, data_processamento
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    d['tipo_encargo'], d['empresa'], d['nota_tipo'], d['numero_nd'],
                    d['data_vencimento'], d['valor_total'], d['quantidade'], d['valor_unitario'],
                    d['arquivo'], d['caminho'], datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))

            conexao.commit()
            conexao.close()

            self.log(f"[OK] Dados salvos em: {db_path}")

            # ── Oferecer cópia de backup em pasta escolhida pelo usuário ──
            resposta = messagebox.askyesno(
                "Backup do Banco de Dados",
                f"Dados salvos com sucesso!\nLocal do banco: {db_path}\n\n"
                "Deseja salvar uma cópia de backup em outra pasta?"
            )
            if resposta:
                from tkinter import filedialog
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = filedialog.asksaveasfilename(
                    title="Salvar cópia do banco de dados",
                    initialfile=f"RET_dados_backup_{timestamp}.db",
                    defaultextension=".db",
                    filetypes=[("Banco de dados SQLite", "*.db"), ("Todos os arquivos", "*.*")],
                    initialdir=self.pasta_selecionada or os.path.expanduser("~"),
                )
                if backup_path:
                    import shutil
                    shutil.copy2(db_path, backup_path)
                    self.log(f"[OK] Backup salvo em: {backup_path}")
                    messagebox.showinfo("Backup Salvo", f"Cópia salva em:\n{backup_path}")
            else:
                messagebox.showinfo("Sucesso", f"Dados salvos no banco!\n{db_path}")

        except Exception as e:
            self.log(f"[ERRO] Falha ao salvar: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar: {e}")
    
    def exportar_excel(self):
        """Exporta dados para Excel formatado.

        Abre um diálogo para o usuário escolher onde salvar o arquivo.
        A pasta sugerida é a pasta de PDFs selecionada (ou a pasta pessoal
        caso nenhuma pasta tenha sido selecionada ainda).
        """
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return

        try:
            from tkinter import filedialog
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = filedialog.asksaveasfilename(
                title="Salvar relatório Excel",
                initialfile=f"RET_Relatorio_{timestamp}.xlsx",
                defaultextension=".xlsx",
                filetypes=[("Planilha Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
                initialdir=self.pasta_selecionada or os.path.expanduser("~"),
            )
            if not excel_path:
                return   # usuário cancelou
            
            # Criar DataFrame
            df = pd.DataFrame([{
                'Tipo de Encargo': d['tipo_encargo'],
                'Empresa': d['empresa'],
                'Nota Débito/Crédito': d['nota_tipo'],
                'Nº': d['numero_nd'],
                'Data Vencimento': d['data_vencimento'],
                'Valor Total': d['valor_total'],
                'QT': d['quantidade'],
                'Valor Unitário': d['valor_unitario'],
                'Arquivo': d['arquivo']
            } for d in self.dados_processados])
            
            # Criar workbook com formatação
            wb = Workbook()
            ws_dados = wb.active
            ws_dados.title = "Dados Completos"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Adicionar dados
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if r_idx == 1:  # Header
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        if c_idx in [6, 7, 8]:  # Colunas numéricas
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.00'
            
            # Ajustar larguras
            ws_dados.column_dimensions['A'].width = 20
            ws_dados.column_dimensions['B'].width = 25
            ws_dados.column_dimensions['C'].width = 20
            ws_dados.column_dimensions['D'].width = 15
            ws_dados.column_dimensions['E'].width = 18
            ws_dados.column_dimensions['F'].width = 15
            ws_dados.column_dimensions['G'].width = 12
            ws_dados.column_dimensions['H'].width = 15
            ws_dados.column_dimensions['I'].width = 40
            
            # ABA RESUMO POR TIPO
            ws_resumo = wb.create_sheet("Resumo por Tipo")
            
            resumo = df.groupby('Tipo de Encargo').agg({
                'Valor Total': 'sum',
                'QT': 'sum',
                'Arquivo': 'count'
            }).rename(columns={'Arquivo': 'Quantidade de Arquivos'}).reset_index()
            
            for r_idx, row in enumerate(dataframe_to_rows(resumo, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_resumo.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        if c_idx > 1:
                            if isinstance(value, (int, float)):
                                cell.number_format = '#,##0.00'
            
            ws_resumo.column_dimensions['A'].width = 25
            ws_resumo.column_dimensions['B'].width = 18
            ws_resumo.column_dimensions['C'].width = 15
            ws_resumo.column_dimensions['D'].width = 25
            
            # ABA RESUMO GERAL
            ws_geral = wb.create_sheet("Resumo Geral")

            calc_ret = self._calcular_ret()
            eat_bruto_xls = calc_ret['eat_bruto']
            pis_rate      = calc_ret['pis_cofins_rate']
            ec_ret_xls    = calc_ret['ret']
            total_qt      = df['QT'].sum()
            total_arqs    = len(df)

            # Estilo para linha de resultado em destaque
            ret_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            ret_font = Font(bold=True, color="FFFFFF", size=12)

            dados_geral = [
                ['RESUMO GERAL DO PROCESSAMENTO', ''],        # 1 – título
                ['', ''],                                      # 2
                ['Métrica', 'Valor'],                          # 3 – cabeçalho
                ['Total de PDFs Processados', total_arqs],    # 4
                ['Quantidade Total (QT)', total_qt],           # 5
                ['', ''],                                      # 6
                ['Σ EAT bruto (R$)', eat_bruto_xls],          # 7
                [f'× (1 − {pis_rate*100:.2f}% PIS/COFINS)', 1.0 - pis_rate],  # 8
                ['EC = RET  (R$)', ec_ret_xls],                # 9 – RESULTADO FINAL
                ['', ''],                                      # 10
                ['Data do Processamento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]  # 11
            ]

            for r_idx, row in enumerate(dados_geral, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_geral.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    if r_idx == 1:
                        cell.font = Font(bold=True, size=16, color="1F4788")
                    elif r_idx == 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    elif r_idx == 9:   # linha EC = RET em destaque
                        cell.fill = ret_fill
                        cell.font = ret_font
                        if c_idx == 2:
                            cell.number_format = '#,##0.000000'
                    else:
                        if c_idx == 2 and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.000000'
                if r_idx == 1:
                    ws_geral.merge_cells('A1:B1')

            ws_geral.column_dimensions['A'].width = 35
            ws_geral.column_dimensions['B'].width = 25
            
            # Salvar e fechar
            wb.save(excel_path)
            wb.close()
            
            self.log(f"[OK] Excel criado: {excel_path}")
            messagebox.showinfo("Sucesso", f"Excel exportado com sucesso!\n{excel_path}")
            
        except Exception as e:
            self.log(f"[ERRO] Falha ao exportar: {e}")
            messagebox.showerror("Erro", f"Erro ao exportar: {e}")
    
    def _salvar_ret_scg(self):
        """Salva o total RET na consolidação"""
        if not hasattr(self, 'dados_processados') or not self.dados_processados:
            messagebox.showwarning("Aviso", "Processe os PDFs primeiro!")
            return
        
        from tkinter import simpledialog
        
        calc = self._calcular_ret()
        total_geral = calc['ret']   # RET = EC = Σ(EAT) × (1 − PIS_COFINS)

        periodo = simpledialog.askstring("Período RET",
                                         "Digite o período (ex: Q1 2026):",
                                         initialvalue="Q1 2026")
        if periodo:
            from database import DatabasePMPV
            db = DatabasePMPV()
            if not db.buscar_consolidacao(periodo):
                db.criar_periodo_consolidacao(periodo, "RET")
            db.atualizar_ret(periodo, total_geral)
            db.fechar()

            total_fmt = f"R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            messagebox.showinfo("RET Salvo", f"RET: {total_fmt}\nPeríodo: {periodo}")

if __name__ == "__main__":
    root = ctk.CTk()
    root.withdraw()
    app = SistemaRET(root)
    app.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()
