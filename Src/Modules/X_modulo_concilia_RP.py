import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import logging
import threading  # Para não travar a tela enquanto processa
from pathlib import Path
from dataclasses import dataclass
from typing import List, Tuple
from datetime import datetime
import re

import pdfplumber
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from Src.common.formatting import parse_brl, format_brl_plain
from Src.infra.ocr_pdf import OCR_ENABLED, read_pdf_text

@dataclass(frozen=True)
class PdfItem:
    file_name: str
    file_path: str
    category: str
    amount: float
    status: str
    method: str

def clean_ocr_text(text: str) -> str:
    if not text:
        return ""
    return (
        text.replace("|", "")
        .replace("!", "1")
        .replace("l", "1")
        .replace("$=", " ")
        .replace("=", " = ")
    )

def extrair_valor(text: str) -> Tuple[float, str]:
    text_clean = clean_ocr_text(text)
    text_upper = text_clean.upper()
    
    eh_documento_oficial = "NOTA" in text_upper or "PENALIDADE" in text_upper or "FISCAL" in text_upper
    todos_valores = re.findall(r"(\d{1,3}(?:\.\d{3})*,\d{2})", text_clean)
    
    lista_floats = []
    if todos_valores:
        for v in todos_valores:
            # Antes estava f = br_money_to_float(v)
            f = parse_brl(v)
            # Filtro de ano/datas
            if f in [2024.0, 2025.0, 2026.0, 2027.0]: continue
            
            if eh_documento_oficial:
                if f > 0: lista_floats.append(f)
            else:
                if f > 50: lista_floats.append(f) # Filtro de ruído

    if lista_floats:
        return max(lista_floats), "Maior Valor Detectado"

    return 0.0, "Valor não identificado"

def processar_lista_arquivos(arquivos: List[Path], categoria: str, log_callback) -> List[PdfItem]:
    itens = []
    total = len(arquivos)
    
    for i, arq in enumerate(arquivos):
        log_callback(f"[{i+1}/{total}] Lendo: {arq.name}...")
        
        texto, metodo_leitura = read_pdf_text(arq)
        if texto:
            valor, metodo_extracao = extrair_valor(texto)
            status = "OK" if valor > 0 else "REVISAR"
            metodo_final = f"{metodo_leitura} -> {metodo_extracao}"
        else:
            valor = 0.0
            status = "ERRO"
            metodo_final = metodo_leitura
            
        itens.append(PdfItem(arq.name, str(arq), categoria, valor, status, metodo_final))
        
    return itens

def salvar_excel(caminho: Path, itens: List[PdfItem]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    
    # Estilos
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws.append(["Arquivo", "Categoria", "Valor", "Status", "Método", "Caminho"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    total_rec = 0.0
    total_desp = 0.0

    for i in itens:
        ws.append([i.file_name, i.category, i.amount, i.status, i.method, i.file_path])
        # Formatação de moeda na célula C
        cell_val = ws.cell(row=ws.max_row, column=3)
        cell_val.number_format = '"R$ "#,##0.00'
        
        if i.status == "OK":
            if i.category == "Receita": total_rec += i.amount
            elif i.category == "Despesa": total_desp += i.amount

    # Totais
    ws.append([])
    ws.append(["RESUMO FINAL", "", "", "", "", ""])
    ws.append(["(+) RECEITAS", "", total_rec, "", "", ""])
    ws.append(["(-) DESPESAS", "", total_desp, "", "", ""])
    ws.append(["(=) SALDO", "", total_rec - total_desp, "", "", ""])
    
    # Ajuste largura
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["E"].width = 30
    
    wb.save(caminho)
    return total_rec, total_desp

# ==========================================
# 2. INTERFACE GRÁFICA MODERNA (CustomTkinter)
# ==========================================

class AppConciliador(ctk.CTkToplevel):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.title("ConciliaPDF 2.0 - Automação Financeira")
        self.geometry("900x700")
        
        # Variáveis de Estado
        self.path_rec = tk.StringVar()
        self.path_desp = tk.StringVar()
        self.status_ocr_txt = "✅ MOTOR OCR ATIVO" if OCR_ENABLED else "❌ OCR NÃO ENCONTRADO"
        self.cor_ocr = "#27ae60" if OCR_ENABLED else "#c0392b"

        self._setup_ui()

    def _setup_ui(self):
        # --- HEADER ---
        self.header_frame = ctk.CTkFrame(self, height=80, fg_color="transparent")
        self.header_frame.pack(fill="x", padx=20, pady=20)
        
        ctk.CTkLabel(self.header_frame, text="Conciliador Financeiro Inteligente", 
                     font=("Roboto", 24, "bold")).pack(side="left")
        
        # Badge do Status OCR
        self.ocr_badge = ctk.CTkLabel(self.header_frame, text=self.status_ocr_txt,
                                    fg_color=self.cor_ocr, text_color="white",
                                    corner_radius=10, font=("Roboto", 12, "bold"), padx=10)
        self.ocr_badge.pack(side="right")

        # --- ÁREA DE SELEÇÃO (GRID) ---
        self.selection_frame = ctk.CTkFrame(self)
        self.selection_frame.pack(fill="x", padx=20, pady=10)
        
        # Coluna Receitas
        self._criar_input_folder(self.selection_frame, "📂 Pasta de RECEITAS (Entrada)", 
                               self.path_rec, self.sel_rec, "green")
        
        # Separador
        ctk.CTkFrame(self.selection_frame, height=2, fg_color="gray30").pack(fill="x", padx=10, pady=10)
        
        # Coluna Despesas
        self._criar_input_folder(self.selection_frame, "📂 Pasta de DESPESAS (Saída)", 
                               self.path_desp, self.sel_desp, "red")

        # --- BOTÃO DE AÇÃO ---
        self.btn_run = ctk.CTkButton(self, text="⚡ PROCESSAR E CONCILIAR", 
                                   command=self.iniciar_thread,
                                   font=("Roboto", 16, "bold"),
                                   height=50, fg_color="#2980b9", hover_color="#3498db")
        self.btn_run.pack(fill="x", padx=40, pady=(20, 5))

        self.btn_salvar_scg = ctk.CTkButton(
            self, text="💾 SALVAR SALDO (RP) NO SCG",
            command=self._salvar_rp_scg,
            font=("Roboto", 13, "bold"),
            height=38,
            fg_color="#27ae60", hover_color="#1e8449",
            state="disabled",
        )
        self.btn_salvar_scg.pack(fill="x", padx=40, pady=(0, 15))

        # --- LOG / CONSOLE ---
        ctk.CTkLabel(self, text="Log de Processamento:", anchor="w").pack(fill="x", padx=20)
        
        self.log_box = ctk.CTkTextbox(self, height=200, font=("Consolas", 12))
        self.log_box.pack(fill="both", expand=True, padx=20, pady=(5, 20))
        self.log_message("Sistema pronto. Selecione as pastas acima.")

        # --- PROGRESS BAR ---
        self.progress = ctk.CTkProgressBar(self)
        self.progress.set(0)
        self.progress.pack(fill="x", padx=0, side="bottom")

    def _criar_input_folder(self, parent, titulo, variavel, comando, cor_borda):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(frame, text=titulo, font=("Roboto", 14, "bold"), 
                     text_color=cor_borda).pack(anchor="w")
        
        sub = ctk.CTkFrame(frame, fg_color="transparent")
        sub.pack(fill="x", pady=5)
        
        entry = ctk.CTkEntry(sub, textvariable=variavel, placeholder_text="Nenhuma pasta selecionada...", 
                           width=500, state="readonly")
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        btn = ctk.CTkButton(sub, text="Selecionar", command=comando, width=100)
        btn.pack(side="right")

    # --- LÓGICA DE INTERFACE ---
    def log_message(self, msg):
        self.log_box.insert("end", f"> {datetime.now().strftime('%H:%M:%S')} | {msg}\n")
        self.log_box.see("end")

    def sel_rec(self):
        p = filedialog.askdirectory()
        if p: self.path_rec.set(p)

    def sel_desp(self):
        p = filedialog.askdirectory()
        if p: self.path_desp.set(p)

    def iniciar_thread(self):
        if not self.path_rec.get() and not self.path_desp.get():
            messagebox.showwarning("Aviso", "Selecione pelo menos uma pasta!")
            return
            
        # Bloqueia botão para evitar duplo clique
        self.btn_run.configure(state="disabled", text="Processando... Aguarde")
        self.progress.start()
        
        # Inicia processamento em segundo plano (THREAD)
        threading.Thread(target=self.rodar_processamento, daemon=True).start()

    def rodar_processamento(self):
        try:
            p_rec = Path(self.path_rec.get()) if self.path_rec.get() else None
            p_desp = Path(self.path_desp.get()) if self.path_desp.get() else None
            
            arquivos_rec = list(p_rec.rglob("*.pdf")) if p_rec else []
            arquivos_desp = list(p_desp.rglob("*.pdf")) if p_desp else []
            
            total_files = len(arquivos_rec) + len(arquivos_desp)
            self.log_message(f"Iniciando. Total de arquivos: {total_files}")
            
            if total_files == 0:
                self.log_message("Nenhum PDF encontrado.")
                self.restaurar_interface()
                return

            # Processamento
            itens = []
            if arquivos_rec:
                self.log_message("--- Processando Receitas ---")
                itens += processar_lista_arquivos(arquivos_rec, "Receita", self.log_message)
                
            if arquivos_desp:
                self.log_message("--- Processando Despesas ---")
                itens += processar_lista_arquivos(arquivos_desp, "Despesa", self.log_message)

            # Salvar
            timestamp = datetime.now().strftime("%H%M%S")
            nome_excel = f"Conciliacao_Final_{timestamp}.xlsx"
            
            # Pergunta onde salvar (precisa ser invocado na thread principal, mas aqui simplificamos)
            # Vamos salvar na pasta de Downloads ou na pasta do script para evitar travar a thread
            caminho_final = Path(os.getcwd()) / nome_excel
            
            self.log_message("Gerando Excel...")
            tot_rec, tot_desp = salvar_excel(caminho_final, itens)
            
            self.log_message(f"CONCLUÍDO! Salvo em: {caminho_final}")
            
            # Mostra Resultado Final na Tela
            saldo = tot_rec - tot_desp
           # Ficará assim:
            msg_final = (f"PROCESSAMENTO FINALIZADO!\n\n"
                         f"Receitas: R$ {format_brl_plain(tot_rec)}\n"
                         f"Despesas: R$ {format_brl_plain(tot_desp)}\n"
                         f"----------------\n"
                         f"SALDO: R$ {format_brl_plain(saldo)}\n\n"
                         f"Relatório salvo na pasta do programa.")
            messagebox.showinfo("Sucesso", msg_final)

            # Guarda saldo para o botão Salvar no SCG
            self._ultimo_saldo_rp = saldo
            self.btn_salvar_scg.configure(state="normal")

        except Exception as e:
            self.log_message(f"ERRO CRÍTICO: {e}")
            messagebox.showerror("Erro", str(e))
        
        finally:
            self.restaurar_interface()

    def _salvar_rp_scg(self):
        """Salva o saldo RP (Receita − Despesa) no banco de consolidação SCG."""
        from tkinter import simpledialog
        from database import DatabasePMPV

        if not hasattr(self, '_ultimo_saldo_rp'):
            messagebox.showwarning("Aviso", "Execute o processamento antes de salvar.")
            return

        periodo = simpledialog.askstring(
            "Salvar RP no SCG",
            "Digite o período (ex: Dez/2025):",
            initialvalue="Dez/2025",
        )
        if not periodo:
            return

        db = DatabasePMPV()
        db.atualizar_rp(periodo, self._ultimo_saldo_rp)
        db.fechar()

        # Ficará assim:
        messagebox.showinfo(
            "RP Salvo ✅",
            f"Período : {periodo}\n"
            f"RP salvo: R$ {format_brl_plain(self._ultimo_saldo_rp)}\n\n"
            f"Acesse o módulo SCG para ver o resultado final.",
        )

    def restaurar_interface(self):
        self.progress.stop()
        self.progress.set(1)
        self.btn_run.configure(state="normal", text="⚡ PROCESSAR E CONCILIAR")

if __name__ == "__main__":
    app = AppConciliador()
    app.mainloop()