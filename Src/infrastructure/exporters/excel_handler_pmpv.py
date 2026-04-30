import os
from datetime import datetime
from typing import Dict, Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference


# ── Paleta ARPE ───────────────────────────────────────────────────────────────
_NAVY        = "1A3A5C"   # azul marinho — cabeçalhos principais
_NAVY2       = "0D2444"   # azul mais escuro
_BLUE        = "2E6DA4"   # azul médio — sub-cabeçalhos
_BLUE_LIGHT  = "BDD7EE"   # azul claro — período de apuração
_CYAN        = "17A589"   # destaque positivo
_GOLD        = "F4D03F"   # destaque amarelo — preço final
_GOLD_LIGHT  = "FEF9E7"   # fundo amarelo claro
_GREEN       = "1E8449"   # positivo
_GREEN_LIGHT = "D5F5E3"   # fundo verde claro
_RED         = "C0392B"   # negativo
_RED_LIGHT   = "FADBD8"   # fundo vermelho claro
_ORANGE      = "D35400"   # alerta
_WHITE       = "FFFFFF"
_GRAY_DARK   = "4A4A4A"
_GRAY_MID    = "7F8C8D"
_GRAY_LIGHT  = "F2F3F4"
_ROW_ALT     = "EBF5FB"   # linha alternada


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def _border(style="thin", color="B0B0B0") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom(color="CCCCCC") -> Border:
    return Border(bottom=Side(style="thin", color=color))


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _to_float(val: Any, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def _merge(ws, r, c1, c2, value, bg, fnt, align_h="center", fmt="@", row_h=None):
    if c1 != c2:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=value)
    cell.fill = _fill(bg)
    cell.font = fnt
    cell.alignment = _align(align_h, "center")
    cell.border = _border(color=bg)
    if fmt != "@":
        cell.number_format = fmt
    if row_h:
        ws.row_dimensions[r].height = row_h
    return cell


def _section_bar(ws, r, c1, c2, text, bg=_NAVY):
    _merge(ws, r, c1, c2, f"  {text}", bg,
           _font(bold=True, size=10, color=_WHITE), "left", row_h=20)


def _kpi_card(ws, start_row, start_col, label, value, fmt, bg_dark, bg_lite):
    """Card de KPI em 2 colunas × 5 linhas."""
    c2 = start_col + 1
    W = Border(
        left=Side(style="thin", color=_WHITE),
        right=Side(style="thin", color=_WHITE),
        top=Side(style="thin", color=_WHITE),
        bottom=Side(style="thin", color=_WHITE),
    )

    def _mc(r, c1, c2_, val, bg, fnt, al="center", fmt_="@"):
        if c1 != c2_:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2_)
        c = ws.cell(row=r, column=c1, value=val)
        c.fill = _fill(bg); c.font = fnt
        c.alignment = _align(al, "center"); c.border = W
        if fmt_ != "@": c.number_format = fmt_
        return c

    # Linha 0: rótulo
    _mc(start_row, start_col, c2, f"  {label}",
        bg_dark, _font(bold=True, size=9, color="C8DCF0"), "left")
    ws.row_dimensions[start_row].height = 18

    # Linha 1: espaço
    for ci in (start_col, c2):
        cc = ws.cell(row=start_row+1, column=ci)
        cc.fill = _fill(bg_dark); cc.border = W
    ws.row_dimensions[start_row+1].height = 4

    # Linha 2: valor principal
    _mc(start_row+2, start_col, c2, value,
        bg_dark, _font(bold=True, size=20, color=_WHITE), "center", fmt_)
    ws.row_dimensions[start_row+2].height = 34

    # Linha 3: espaço
    for ci in (start_col, c2):
        cc = ws.cell(row=start_row+3, column=ci)
        cc.fill = _fill(bg_lite); cc.border = W
    ws.row_dimensions[start_row+3].height = 4

    # Linha 4: sub-rótulo
    _mc(start_row+4, start_col, c2, "R$/m³" if "R$" in str(fmt) else "m³",
        bg_lite, _font(size=8, color=_GRAY_MID, italic=True), "center")
    ws.row_dimensions[start_row+4].height = 14


# ── Formatação de tabela de dados ─────────────────────────────────────────────

_BRL4 = 'R$ #,##0.0000'
_BRL2 = 'R$ #,##0.00'
_VOL  = '#,##0.000000'
_NUM  = '#,##0.00'


def _header_row(ws, row_num, labels, widths, bg=_NAVY, height=22):
    ws.row_dimensions[row_num].height = height
    for i, (lbl, w) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=row_num, column=i, value=lbl)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=_WHITE, size=10)
        cell.alignment = _align("center")
        cell.border = _border(color=bg)
        ws.column_dimensions[get_column_letter(i)].width = w


def _data_row(ws, row_num, values, fmts=None, alternate=False):
    bg = _ROW_ALT if alternate else _WHITE
    fmts = fmts or ["@"] * len(values)
    ws.row_dimensions[row_num].height = 18
    for i, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.fill = _fill(bg)
        cell.font = _font(size=10)
        cell.alignment = _align("right" if fmt != "@" else "left")
        cell.border = _border_bottom()
        if fmt != "@":
            cell.number_format = fmt


def _total_row(ws, row_num, values, fmts=None, bg=_NAVY, color=_WHITE):
    fmts = fmts or ["@"] * len(values)
    ws.row_dimensions[row_num].height = 20
    for i, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell = ws.cell(row=row_num, column=i, value=val)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=color, size=10)
        cell.alignment = _align("right" if fmt != "@" else "left")
        cell.border = _border(color=bg)
        if fmt != "@":
            cell.number_format = fmt


# ══════════════════════════════════════════════════════════════════════════════
# CLASSE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

class ExcelHandlerPMPV:

    @staticmethod
    def exportar_trimestre(dados_por_mes: Dict, resultado: Dict, nome_arquivo: str = None) -> str:
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"Relatorio_PMPV_{timestamp}.xlsx"

        # Garante nome único se arquivo estiver aberto
        nome_base = nome_arquivo.replace(".xlsx", "")
        contador = 1
        nome_final = nome_arquivo
        while True:
            try:
                with open(nome_final, "w"):
                    pass
                os.remove(nome_final)
                break
            except (PermissionError, IOError):
                nome_final = f"{nome_base}_{contador}.xlsx"
                contador += 1
                if contador > 100:
                    nome_final = f"{nome_base}_{datetime.now().strftime('%H%M%S%f')}.xlsx"
                    break

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Aba Dashboard primeiro (posição 0)
        ExcelHandlerPMPV._criar_aba_dashboard(wb, dados_por_mes, resultado)

        # Abas de dados mensais
        for nome_aba, dados in dados_por_mes.items():
            ExcelHandlerPMPV._criar_aba_mes(wb, nome_aba, dados, resultado)

        # Aba de resumo executivo
        ExcelHandlerPMPV._criar_aba_resumo(wb, dados_por_mes, resultado)

        wb.save(nome_final)
        wb.close()

        try:
            os.startfile(nome_final)
        except Exception:
            pass

        return nome_final

    # ── Dashboard ─────────────────────────────────────────────────────────────

    @staticmethod
    def _criar_aba_dashboard(wb, dados_por_mes, resultado):
        ws = wb.create_sheet("📊 Dashboard", 0)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.showRowColHeaders = False
        ws.sheet_properties.tabColor = _NAVY2
        ws.sheet_view.zoomScale = 95

        pmpv      = _to_float(resultado.get("pmpv"))
        preco_fin = _to_float(resultado.get("preco_final"))
        conta_g   = _to_float(resultado.get("conta_grafica"))
        vp        = _to_float(resultado.get("vp_mensal", resultado.get("volume_programado_mensal", 0)))
        vf        = _to_float(resultado.get("vf_total", resultado.get("volume_total", 0)))
        custo     = _to_float(resultado.get("custo_total"))
        ts        = datetime.now().strftime("%d/%m/%Y  %H:%M")

        # ── Layout de colunas: A=marg | B:C=card | D=gap | E:F=card | G=gap | H:I=card | J=marg
        col_cfg = [("A", 1.5), ("B", 16), ("C", 16), ("D", 2.5),
                   ("E", 16), ("F", 16), ("G", 2.5), ("H", 16), ("I", 16), ("J", 1.5)]
        for col_l, w in col_cfg:
            ws.column_dimensions[col_l].width = w

        FULL_C1, FULL_C2 = 2, 9   # B até I

        def _bg_row(r, h, bg):
            ws.row_dimensions[r].height = h
            for ci in range(1, 11):
                ws.cell(row=r, column=ci).fill = _fill(bg)

        # ── R1 — Banner ARPE ──────────────────────────────────────────────────
        _bg_row(1, 48, _NAVY2)
        _merge(ws, 1, FULL_C1, 6,
               "  ARPE  ·  TARIFA DE GÁS CANALIZADO  ·  RELATÓRIO PMPV",
               _NAVY2, _font(bold=True, size=16, color=_WHITE), "left", row_h=48)
        _merge(ws, 1, 7, FULL_C2,
               f"  {ts}  ",
               _GOLD, _font(bold=True, size=11, color=_NAVY2), "right")

        # ── R2 — Faixa azul médio ─────────────────────────────────────────────
        _bg_row(2, 5, _BLUE)

        # ── R3 — Subtítulo período ────────────────────────────────────────────
        _bg_row(3, 22, _BLUE)
        meses_nomes = list(dados_por_mes.keys())
        periodo_txt = "  ".join(meses_nomes) if meses_nomes else "—"
        _merge(ws, 3, FULL_C1, FULL_C2,
               f"  PERÍODO DE APURAÇÃO  ·  {periodo_txt}",
               _BLUE, _font(bold=True, size=11, color=_BLUE_LIGHT), "left")

        # ── R4 — espaço ───────────────────────────────────────────────────────
        _bg_row(4, 10, "F0F4F8")

        # ── CARDS KPI (R5–R9) ─────────────────────────────────────────────────
        R = 5
        for gap_ci in (1, 4, 7, 10):
            for rr in range(R, R + 5):
                ws.cell(row=rr, column=gap_ci).fill = _fill("F0F4F8")

        _kpi_card(ws, R, 2, "PMPV CALCULADO",      pmpv,      _BRL4, _NAVY,  _BLUE)
        _kpi_card(ws, R, 5, "PREÇO FINAL (PV)",     preco_fin, _BRL4, "0E6655", "117A65")
        _kpi_card(ws, R, 8, "VOLUME PROSPECTIVO",   vp,        '#,##0.000', "154360", "1A5276")

        # ── R10 — espaço ──────────────────────────────────────────────────────
        _bg_row(R + 5, 10, "F0F4F8")

        # ── FLUXO DE CÁLCULO (R11–R19) ────────────────────────────────────────
        RF = R + 6
        _bg_row(RF, 22, _NAVY)
        _merge(ws, RF, FULL_C1, FULL_C2,
               "  FLUXO DE CÁLCULO — DECOMPOSIÇÃO DO PREÇO FINAL",
               _NAVY, _font(bold=True, size=10, color=_WHITE), "left")

        W_BOX = Border(
            left=Side(style="medium", color=_BLUE),
            right=Side(style="medium", color=_BLUE),
            top=Side(style="medium", color=_BLUE),
            bottom=Side(style="medium", color=_BLUE),
        )

        def _flow_box(r, c1, c2, label, value, fmt, bg, val_color=_WHITE):
            ws.row_dimensions[r].height = 18
            ws.row_dimensions[r+1].height = 28
            ws.row_dimensions[r+2].height = 14

            if c1 != c2:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            lc = ws.cell(row=r, column=c1, value=f"  {label}")
            lc.fill = _fill(bg); lc.font = _font(bold=True, size=9, color="C8DCF0")
            lc.alignment = _align("left", "center"); lc.border = W_BOX

            if c1 != c2:
                ws.merge_cells(start_row=r+1, start_column=c1, end_row=r+1, end_column=c2)
            vc = ws.cell(row=r+1, column=c1, value=value)
            vc.fill = _fill(bg); vc.font = _font(bold=True, size=18, color=val_color)
            vc.alignment = _align("center", "center"); vc.border = W_BOX
            if fmt != "@": vc.number_format = fmt

            for ci in range(c1, c2 + 1):
                ws.cell(row=r+2, column=ci).fill = _fill("F0F4F8")

        RF2 = RF + 1
        _flow_box(RF2, 2, 3, "PMPV",           pmpv,      _BRL4, _NAVY)
        # seta
        ws.row_dimensions[RF2].height   = 18
        ws.row_dimensions[RF2+1].height = 28
        _merge(ws, RF2,   4, 4, "",   "F0F4F8", _font(size=14, color=_GRAY_MID), "center")
        _merge(ws, RF2+1, 4, 4, "►", "F0F4F8", _font(size=18, color=_BLUE), "center")

        _flow_box(RF2, 5, 6, "(+) CONTA GRÁFICA",
                  conta_g, _BRL4,
                  "1A5276" if conta_g >= 0 else "922B21",
                  val_color=_GOLD)
        _merge(ws, RF2,   7, 7, "",   "F0F4F8", _font(size=14), "center")
        _merge(ws, RF2+1, 7, 7, "►", "F0F4F8", _font(size=18, color=_BLUE), "center")

        _flow_box(RF2, 8, 9, "(=) PREÇO FINAL (PV)",
                  preco_fin, _BRL4, "0E6655", val_color=_GOLD)

        # ── espaço ────────────────────────────────────────────────────────────
        _bg_row(RF2 + 3, 10, "F0F4F8")

        # ── TABELA RESUMO POR MÊS (R20+) ──────────────────────────────────────
        RT = RF2 + 4
        _bg_row(RT, 22, _NAVY)
        _merge(ws, RT, FULL_C1, FULL_C2,
               "  RESUMO POR EMPRESA / MÊS",
               _NAVY, _font(bold=True, size=10, color=_WHITE), "left")

        RT += 1
        _header_row(ws, RT,
                    ["Mês", "Empresa", "Molécula (R$/m³)", "Transporte (R$/m³)",
                     "Logística (R$/m³)", "Preço Unit. (R$/m³)", "Volume (m³/dia)", "Custo (R$)"],
                    [14, 22, 18, 18, 16, 18, 16, 18], _BLUE)

        row_idx = RT + 1
        alt = False
        tot_custo = 0.0
        for mes_nome, linhas in dados_por_mes.items():
            for linha in linhas:
                if not linha.get("empresa"):
                    continue
                mol   = _to_float(linha.get("molecula"))
                trans = _to_float(linha.get("transporte"))
                log_  = _to_float(linha.get("logistica"))
                vol   = _to_float(linha.get("volume"))
                preco = mol + trans + log_
                custo_linha = preco * vol
                tot_custo  += custo_linha
                _data_row(ws, row_idx,
                          [mes_nome, linha.get("empresa", ""), mol, trans, log_, preco, vol, custo_linha],
                          ["@", "@", _BRL4, _BRL4, _BRL4, _BRL4, _VOL, _BRL2],
                          alternate=alt)
                alt = not alt
                row_idx += 1

        _total_row(ws, row_idx,
                   ["TOTAL", "", "", "", "", "", "", tot_custo],
                   ["@", "@", "@", "@", "@", "@", "@", _BRL2],
                   bg=_NAVY, color=_GOLD)

        # ── Rodapé ────────────────────────────────────────────────────────────
        foot_r = row_idx + 2
        _bg_row(foot_r, 3, _NAVY2)
        _merge(ws, foot_r, FULL_C1, FULL_C2,
               f"   PMPV = Σ(Pᵢ × Vᵢ) / ΣVᵢ   |   PV = PMPV + PR   |   "
               f"ARPE – Agência de Regulação de Pernambuco  ·  {datetime.now().year}",
               _NAVY2, _font(italic=True, size=8, color="5A7A9A"), "center")

    # ── Aba por mês ───────────────────────────────────────────────────────────

    @staticmethod
    def _criar_aba_mes(wb, nome_aba, dados, resultado):
        ws = wb.create_sheet(nome_aba)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = _BLUE

        col_cfg = [("A", 28), ("B", 14), ("C", 14), ("D", 14),
                   ("E", 16), ("F", 16), ("G", 16)]
        for col_l, w in col_cfg:
            ws.column_dimensions[col_l].width = w

        # Cabeçalho
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value = f"  ARPE  ·  DADOS MENSAIS  ·  {nome_aba.upper()}"
        c.fill = _fill(_NAVY2)
        c.font = _font(bold=True, size=13, color=_WHITE)
        c.alignment = _align("left", "center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:G2")
        c2 = ws["A2"]
        c2.value = f"  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Módulo PMPV"
        c2.fill = _fill(_BLUE)
        c2.font = _font(size=9, color=_BLUE_LIGHT, italic=True)
        c2.alignment = _align("left", "center")
        ws.row_dimensions[2].height = 16

        ws.row_dimensions[3].height = 8
        for ci in range(1, 8):
            ws.cell(row=3, column=ci).fill = _fill("F0F4F8")

        # Cabeçalho da tabela
        labels  = ["Empresa", "Molécula\n(R$/m³)", "Transporte\n(R$/m³)",
                   "Logística\n(R$/m³)", "Preço Unit.\n(R$/m³)", "Volume\n(m³/dia)", "Custo\n(R$)"]
        widths  = [28, 14, 14, 14, 16, 16, 16]
        _header_row(ws, 4, labels, widths, _NAVY, height=30)
        for ci in range(1, 8):
            ws.cell(row=4, column=ci).alignment = _align("center", "center", wrap=True)

        row_idx = 5
        alt = False
        tot_mol = tot_trans = tot_log = tot_preco_pond = tot_vol = tot_custo = 0.0

        for linha in dados:
            if not linha.get("empresa"):
                continue
            mol   = _to_float(linha.get("molecula"))
            trans = _to_float(linha.get("transporte"))
            log_  = _to_float(linha.get("logistica"))
            vol   = _to_float(linha.get("volume"))
            preco = mol + trans + log_
            custo = preco * vol

            tot_mol       += mol
            tot_trans     += trans
            tot_log       += log_
            tot_preco_pond += preco * vol
            tot_vol       += vol
            tot_custo     += custo

            _data_row(ws, row_idx,
                      [linha["empresa"], mol, trans, log_, preco, vol, custo],
                      ["@", _BRL4, _BRL4, _BRL4, _BRL4, _VOL, _BRL2],
                      alternate=alt)
            ws.cell(row=row_idx, column=1).font = _font(bold=True, size=10)
            alt = not alt
            row_idx += 1

        # Linha de totais
        preco_medio = tot_preco_pond / tot_vol if tot_vol else 0.0
        _total_row(ws, row_idx,
                   ["TOTAL / MÉDIA PONDERADA", tot_mol, tot_trans, tot_log,
                    preco_medio, tot_vol, tot_custo],
                   ["@", _BRL4, _BRL4, _BRL4, _BRL4, _VOL, _BRL2],
                   bg=_NAVY, color=_GOLD)

        # Mini-resumo de PMPV no canto
        row_idx += 2
        _section_bar(ws, row_idx, 1, 3, "RESULTADO DO CÁLCULO")
        row_idx += 1

        def _res_line(r, label, val, fmt, is_final=False):
            bg_r = _GOLD_LIGHT if is_final else "F0F4F8"
            col_r = _NAVY if is_final else _GRAY_DARK
            ws.row_dimensions[r].height = 18
            lc = ws.cell(row=r, column=1, value=f"  {label}")
            lc.fill = _fill(bg_r); lc.font = _font(bold=is_final, size=10, color=col_r)
            lc.alignment = _align("left"); lc.border = _border_bottom()

            vc = ws.cell(row=r, column=2, value=val)
            vc.fill = _fill(bg_r)
            vc.font = _font(bold=True, size=10, color=_NAVY if is_final else _NAVY)
            vc.alignment = _align("right"); vc.border = _border_bottom()
            if fmt != "@": vc.number_format = fmt

        _res_line(row_idx,     "PMPV Calculado",    resultado.get("pmpv", 0),       _BRL4)
        _res_line(row_idx + 1, "(+) Conta Gráfica", resultado.get("conta_grafica", 0), _BRL4)
        _res_line(row_idx + 2, "(=) Preço Final",   resultado.get("preco_final", 0),   _BRL4, is_final=True)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16

    # ── Aba Resumo Executivo ──────────────────────────────────────────────────

    @staticmethod
    def _criar_aba_resumo(wb, dados_por_mes, resultado):
        ws = wb.create_sheet("📋 Resumo Executivo")
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = _GOLD[:6]

        col_cfg = [("A", 36), ("B", 20), ("C", 20), ("D", 20)]
        for col_l, w in col_cfg:
            ws.column_dimensions[col_l].width = w

        pmpv      = _to_float(resultado.get("pmpv"))
        preco_fin = _to_float(resultado.get("preco_final"))
        conta_g   = _to_float(resultado.get("conta_grafica"))
        vp        = _to_float(resultado.get("vp_mensal", resultado.get("volume_programado_mensal", 0)))
        vf        = _to_float(resultado.get("vf_total", resultado.get("volume_total", 0)))
        custo     = _to_float(resultado.get("custo_total"))
        ts        = datetime.now().strftime("%d/%m/%Y  %H:%M")

        # ── Banner ────────────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        c = ws["A1"]
        c.value = "  ARPE  ·  FECHAMENTO TRIMESTRAL  ·  PMPV"
        c.fill = _fill(_NAVY2); c.font = _font(bold=True, size=16, color=_WHITE)
        c.alignment = _align("left", "center"); ws.row_dimensions[1].height = 44

        ws.merge_cells("A2:D2")
        c2 = ws["A2"]
        c2.value = f"  Relatório gerado em {ts}  |  Agência de Regulação de Pernambuco"
        c2.fill = _fill(_BLUE); c2.font = _font(size=9, color=_BLUE_LIGHT, italic=True)
        c2.alignment = _align("left"); ws.row_dimensions[2].height = 16

        ws.merge_cells("A3:D3")
        ws["A3"].fill = _fill("F0F4F8"); ws.row_dimensions[3].height = 10

        # ── Cards de resultado ────────────────────────────────────────────────
        def _card_res(r, label, val, fmt, bg_d, bg_l, unit="R$/m³"):
            ws.row_dimensions[r].height   = 20
            ws.row_dimensions[r+1].height = 36
            ws.row_dimensions[r+2].height = 16

            bdr = Border(
                left=Side(style="medium", color=bg_d),
                right=Side(style="medium", color=bg_d),
                top=Side(style="medium", color=bg_d),
                bottom=Side(style="medium", color=bg_d),
            )

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            lc = ws.cell(row=r, column=1, value=f"  {label}")
            lc.fill = _fill(bg_d); lc.font = _font(bold=True, size=10, color="C8DCF0")
            lc.alignment = _align("left"); lc.border = bdr

            ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=4)
            vc = ws.cell(row=r+1, column=1, value=val)
            vc.fill = _fill(bg_d); vc.font = _font(bold=True, size=26, color=_WHITE)
            vc.alignment = _align("center"); vc.border = bdr
            if fmt != "@": vc.number_format = fmt

            ws.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=4)
            sc = ws.cell(row=r+2, column=1, value=unit)
            sc.fill = _fill(bg_l); sc.font = _font(size=9, color=_GRAY_MID, italic=True)
            sc.alignment = _align("center"); sc.border = bdr

        row = 4
        _card_res(row, "PMPV CALCULADO", pmpv, _BRL4, _NAVY, _BLUE)
        row += 4
        ws.row_dimensions[row].height = 6

        _card_res(row + 1, "(+) CONTA GRÁFICA",
                  conta_g, _BRL4,
                  "1A5276" if conta_g >= 0 else "922B21",
                  "EBF5FB" if conta_g >= 0 else _RED_LIGHT,
                  "R$/m³ — Parcela de Recuperação")
        row += 5
        ws.row_dimensions[row].height = 6

        _card_res(row + 1, "(=) PREÇO FINAL (PV)", preco_fin, _BRL4,
                  "0E6655", _GREEN_LIGHT, "R$/m³ — Preço de Venda ao Consumidor")
        row += 5

        # ── Separador ─────────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 8
        ws[f"A{row}"].fill = _fill("F0F4F8")
        row += 1

        # ── Tabela de indicadores secundários ─────────────────────────────────
        _section_bar(ws, row, 1, 4, "INDICADORES COMPLEMENTARES")
        row += 1

        _header_row(ws, row, ["Indicador", "Valor", "Unidade", "Observação"],
                    [36, 20, 16, 20], _BLUE, height=20)
        row += 1

        indicadores = [
            ("Volume Faturado (VF) — Trimestre",        vf,     "#,##0.00",  "m³"),
            ("Volume Prospectivo (VP) — Trimestre",      vp,     "#,##0.000", "m³"),
            ("Custo Total — Trimestre",                  custo,  _BRL2,       "R$"),
            ("PMPV = Σ(Pᵢ·Vᵢ) / ΣVᵢ",                 pmpv,   _BRL4,       "R$/m³"),
            ("Conta Gráfica (PR)",                       conta_g, _BRL4,      "R$/m³"),
            ("Preço Final (PV = PMPV + PR)",             preco_fin, _BRL4,    "R$/m³"),
        ]

        for i, (ind, val, fmt, unid) in enumerate(indicadores):
            bg = _GOLD_LIGHT if "Preço Final" in ind else (_ROW_ALT if i % 2 else _WHITE)
            bold_row = "Preço Final" in ind
            ws.row_dimensions[row].height = 18
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).fill = _fill(bg)
                ws.cell(row=row, column=ci).border = _border_bottom()

            lc = ws.cell(row=row, column=1, value=f"  {ind}")
            lc.font = _font(bold=bold_row, size=10)

            vc = ws.cell(row=row, column=2, value=val)
            vc.font = _font(bold=True, size=10, color=_NAVY if bold_row else _GRAY_DARK)
            vc.alignment = _align("right")
            if fmt != "@": vc.number_format = fmt

            ws.cell(row=row, column=3, value=unid).font = _font(size=10, color=_GRAY_MID)
            row += 1

        # ── Rodapé ────────────────────────────────────────────────────────────
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 18
        fc = ws[f"A{row}"]
        fc.value = f"  PMPV = Σ(Pᵢ·Vᵢ) / ΣVᵢ  |  PV = PMPV + PR  |  ARPE · {datetime.now().year}"
        fc.fill = _fill(_NAVY2)
        fc.font = _font(italic=True, size=8, color="5A7A9A")
        fc.alignment = _align("center")
