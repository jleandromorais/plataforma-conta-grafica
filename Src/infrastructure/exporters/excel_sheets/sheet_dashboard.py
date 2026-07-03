from __future__ import annotations

from datetime import datetime

from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Border

from Src.infrastructure.exporters.excel_styles import (
    fill as _fill, font as _font, align as _align, to_float as _to_float,
)
from ._helpers import BRL, VOL, VOL4


def sheet_dashboard(
    wb,
    cons: dict | None,
    cons_periodos: list[dict],
    pr: dict | None,
    pv: dict | None,
    sr: dict | None,
    periodo: str | None,
) -> None:
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines     = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_properties.tabColor    = "0F1A2E"
    ws.sheet_view.zoomScale         = 100

    BG       = "0F1A2E"
    SURFACE  = "1A2940"
    SURFACE2 = "0B1424"
    ACCENT   = "00D9C6"
    ACCENT2  = "60E5DA"
    GOLD     = "FFD166"
    BLUE     = "4FC3F7"
    PURPLE   = "B388FF"
    GREEN    = "69F0AE"
    RED      = "FF5252"
    ORANGE   = "FFAB40"
    WHITE    = "FFFFFF"
    MUTED    = "8896B0"
    DIM      = "5A6B85"
    BORDER   = "243A5C"

    col_cfg = [
        ("A", 2.0),
        ("B", 11.0), ("C", 11.0), ("D", 11.0),
        ("E", 1.5),
        ("F", 11.0), ("G", 11.0), ("H", 11.0),
        ("I", 1.5),
        ("J", 11.0), ("K", 11.0), ("L", 11.0),
        ("M", 1.5),
        ("N", 11.0), ("O", 11.0), ("P", 11.0),
        ("Q", 2.0),
    ]
    for col_ltr, w in col_cfg:
        ws.column_dimensions[col_ltr].width = w

    CARD_STARTS = [2, 6, 10, 14]
    FULL_START  = 2
    FULL_END    = 16

    d     = cons or {}
    cgr   = _to_float(d.get("cgr"))
    cgf   = _to_float(d.get("cgf"))
    rpv   = _to_float(d.get("rpv", cgr - cgf))
    ret   = _to_float(d.get("ret"))
    rp    = _to_float(d.get("rp"))
    scg   = _to_float(d.get("scg", rpv + ret + rp))
    pr_d  = pr or {}
    sr_d  = sr or {}
    pv_d  = pv or {}
    pr_v  = _to_float(pr_d.get("pr"))
    pv_v  = _to_float(pv_d.get("pv"))
    pmpv_v= _to_float(pv_d.get("pmpv"))
    vp_v  = _to_float(pr_d.get("vp") or sr_d.get("vp"))
    sr_v  = _to_float(sr_d.get("sr"))
    saldo = scg + sr_v

    NONE_BDR = Border()

    def _no_border(r, c1, c2):
        for ci in range(c1, c2 + 1):
            ws.cell(row=r, column=ci).border = NONE_BDR

    def _rh(r, h):
        ws.row_dimensions[r].height = h

    def _bg_row(r, h, bg, c1=1, c2=17):
        _rh(r, h)
        for ci in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = _fill(bg)
            cell.border = NONE_BDR

    def _merge(r, c1, c2, value, bg, fnt, align_h="center", fmt="@", row_h=None):
        if c1 != c2:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=value)
        cell.fill = _fill(bg)
        cell.font = fnt
        cell.alignment = _align(align_h, "center")
        cell.border = NONE_BDR
        if fmt != "@":
            cell.number_format = fmt
        if row_h:
            _rh(r, row_h)
        return cell

    def _fill_range(r, c1, c2, bg):
        for ci in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = _fill(bg)
            cell.border = NONE_BDR

    def _kpi_card(row, col, icon, label, value, fmt, accent_color,
                  sub1_lbl=None, sub1_val=None, sub1_fmt=BRL,
                  sub2_lbl=None, sub2_val=None, sub2_fmt=BRL):
        c2 = col + 2
        _fill_range(row, col, c2, accent_color)
        _rh(row, 4)
        _merge(row+1, col, c2, f"   {icon}  {label}",
               SURFACE, _font(bold=True, size=10, color=MUTED), "left", "@", 22)
        _merge(row+2, col, c2, value,
               SURFACE, _font(bold=True, size=18, color=WHITE), "center", fmt, 36)
        _fill_range(row+3, col, c2, BORDER)
        _rh(row+3, 2)
        for offset, (lbl, val, vfmt) in enumerate([
            (sub1_lbl, sub1_val, sub1_fmt),
            (sub2_lbl, sub2_val, sub2_fmt),
        ]):
            rr = row + 4 + offset
            if lbl is None:
                _fill_range(rr, col, c2, SURFACE2)
                _rh(rr, 16)
                continue
            lc = ws.cell(row=rr, column=col, value=f"  {lbl}")
            _fill_range(rr, col, c2, SURFACE2)
            lc.fill = _fill(SURFACE2)
            lc.font = _font(size=9, color=MUTED)
            lc.alignment = _align("left", "center")
            lc.border = NONE_BDR
            vc = ws.cell(row=rr, column=c2, value=val)
            vc.fill = _fill(SURFACE2)
            vc.font = _font(bold=True, size=10, color=accent_color)
            vc.alignment = _align("right", "center")
            vc.border = NONE_BDR
            if val is not None and vfmt != "@":
                vc.number_format = vfmt
            _rh(rr, 16)
        _fill_range(row+6, col, c2, BG)
        _rh(row+6, 6)

    def _section_band(row, label):
        _bg_row(row, 6, BG)
        _bg_row(row+1, 28, SURFACE2)
        _merge(row+1, FULL_START, FULL_END,
               f"   {label}",
               SURFACE2, _font(bold=True, size=11, color=ACCENT), "left", "@", 28)
        _bg_row(row+2, 2, ACCENT)
        _bg_row(row+3, 8, BG)
        return row + 4

    for r in range(1, 80):
        _bg_row(r, ws.row_dimensions[r].height or 15, BG)

    _bg_row(1, 3, ACCENT)
    _bg_row(2, 56, SURFACE2)
    _merge(2, FULL_START, 11,
           "   ARPE  ·  CONTA GRÁFICA",
           SURFACE2, _font(bold=True, size=22, color=WHITE), "left")
    _merge(2, 12, FULL_END,
           f" {periodo or 'GERAL'}  ",
           SURFACE2, _font(bold=True, size=14, color=ACCENT), "right")
    _bg_row(3, 22, SURFACE)
    _merge(3, FULL_START, 11,
           "   Tarifa de Gás Canalizado · Dashboard Executivo",
           SURFACE, _font(size=10, color=MUTED, italic=True), "left")
    _merge(3, 12, FULL_END,
           f"Gerado {datetime.now().strftime('%d/%m/%Y · %H:%M')}  ",
           SURFACE, _font(size=10, color=DIM), "right")
    _bg_row(4, 18, BG)

    R = _section_band(5, "INDICADORES PRINCIPAIS")
    saldo_color = GREEN if saldo >= 0 else RED
    scg_color   = GREEN if scg   >= 0 else RED

    _kpi_card(R, CARD_STARTS[0], "💰", "SALDO A RECUPERAR",
              saldo, BRL, saldo_color,
              "SCG Atualizado",        scg,  BRL,
              "Saldo Remanescente SR", sr_v, BRL)
    _kpi_card(R, CARD_STARTS[1], "💼", "SCG — CONTA GRÁFICA",
              scg, BRL, scg_color,
              "RPV (CGR − CGF)", rpv,      BRL,
              "RET + RP",        ret + rp, BRL)
    _kpi_card(R, CARD_STARTS[2], "📈", "PARCELA DE RECUPERAÇÃO",
              pr_v, VOL4, ACCENT,
              "Volume Prosp. (m³)", vp_v, VOL,
              "Saldo / VP",         None, "@")
    _kpi_card(R, CARD_STARTS[3], "🎯", "PREÇO FINAL — PV",
              pv_v, VOL4, GOLD,
              "PMPV (R$/m³)", pmpv_v, VOL4,
              "PR (R$/m³)",   pr_v,   VOL4)

    for r in range(R, R + 8):
        for gap_col in (5, 9, 13):
            ws.cell(row=r, column=gap_col).fill = _fill(BG)
            ws.cell(row=r, column=gap_col).border = NONE_BDR
        ws.cell(row=r, column=1).fill  = _fill(BG)
        ws.cell(row=r, column=17).fill = _fill(BG)

    R2 = R + 9
    R2 = _section_band(R2, "COMPONENTES DA CONTA GRÁFICA")
    _kpi_card(R2, CARD_STARTS[0], "🔍", "CGR · AUDITORIA XML",
              cgr, BRL, BLUE,
              "Notas Fiscais (NF-e)", cgr, BRL, None, None, "@")
    _kpi_card(R2, CARD_STARTS[1], "📋", "CGF · VOLUME × PMPV",
              cgf, BRL, GOLD,
              "Volume Faturado", None, "@", "× PMPV trimestral", None, "@")
    _kpi_card(R2, CARD_STARTS[2], "⚡", "RET · ENCARGOS",
              ret, BRL, ORANGE,
              "EAT × (1 − PIS/COFINS)", None, "@", "+ Encargos Capacidade", None, "@")
    _kpi_card(R2, CARD_STARTS[3], "📄", "RP · CONCILIAÇÃO",
              rp, BRL, PURPLE,
              "Penalidades Recebidas", None, "@", "− Penalidades Aplicadas", None, "@")

    for r in range(R2, R2 + 8):
        for gap_col in (5, 9, 13):
            ws.cell(row=r, column=gap_col).fill = _fill(BG)
            ws.cell(row=r, column=gap_col).border = NONE_BDR
        ws.cell(row=r, column=1).fill  = _fill(BG)
        ws.cell(row=r, column=17).fill = _fill(BG)

    R3 = R2 + 9
    R3 = _section_band(R3, "EQUAÇÃO DA CONTA GRÁFICA")

    def _formula_box(row, col_start, col_end, label, value, fmt, color):
        _bg_row(row, 22, SURFACE)
        _merge(row, col_start, col_end, label,
               SURFACE, _font(size=9, color=MUTED), "center", "@", 22)
        _bg_row(row+1, 30, SURFACE)
        _merge(row+1, col_start, col_end, value,
               SURFACE, _font(bold=True, size=14, color=color), "center", fmt, 30)
        _fill_range(row+2, col_start, col_end, color)
        _rh(row+2, 2)

    def _formula_op(row, col_start, col_end, op):
        _bg_row(row, 22, BG)
        _bg_row(row+1, 30, BG)
        _merge(row+1, col_start, col_end, op,
               BG, _font(bold=True, size=18, color=ACCENT), "center", "@", 30)
        _bg_row(row+2, 2, BG)

    FR = R3
    _formula_box(FR, 2, 3,   "CGR",      cgr,      BRL, BLUE)
    _formula_op(FR, 4, 4, "−")
    _formula_box(FR, 5, 6,   "CGF",      cgf,      BRL, GOLD)
    _formula_op(FR, 7, 7, "=")
    _formula_box(FR, 8, 9,   "RPV",      rpv,      BRL, PURPLE)
    _formula_op(FR, 10, 10, "+")
    _formula_box(FR, 11, 12, "RET + RP", ret + rp, BRL, ORANGE)
    _formula_op(FR, 13, 13, "=")
    _formula_box(FR, 14, 16, "SCG",      scg,      BRL, scg_color)

    R4 = FR + 4
    R4 = _section_band(R4, "PARCELA DE RECUPERAÇÃO · HISTÓRICO POR PERÍODO")

    TBL_R = R4
    for ci, lbl in ((2, "Período"), (3, "PR (R$/m³)")):
        cc = ws.cell(row=TBL_R, column=ci, value=lbl)
        cc.fill = _fill(SURFACE)
        cc.font = _font(bold=True, size=8, color=DIM)
        cc.alignment = _align("center")
        cc.border = NONE_BDR
    _rh(TBL_R, 14)

    periodos_g = cons_periodos[-14:] if cons_periodos else []
    dr = TBL_R + 1
    for i, item in enumerate(periodos_g):
        p_txt = item.get("periodo", "")
        p_pr = _to_float(item.get("pr")) if item.get("pr") else (
            _to_float(item.get("scg", 0)) / max(_to_float(item.get("vp", 1)), 1)
        )
        pc = ws.cell(row=dr, column=2, value=p_txt)
        pc.fill = _fill(SURFACE if i % 2 == 0 else SURFACE2)
        pc.font = _font(size=9, color=MUTED)
        pc.alignment = _align("center")
        pc.border = NONE_BDR
        vc = ws.cell(row=dr, column=3, value=p_pr)
        vc.fill = _fill(SURFACE if i % 2 == 0 else SURFACE2)
        vc.font = _font(size=9, color=ACCENT, bold=True)
        vc.alignment = _align("center")
        vc.number_format = VOL4
        vc.border = NONE_BDR
        _rh(dr, 13)
        dr += 1

    if not periodos_g and pr_v:
        ws.cell(row=dr, column=2, value=periodo or "Atual").fill = _fill(SURFACE)
        vc = ws.cell(row=dr, column=3, value=pr_v)
        vc.fill = _fill(SURFACE)
        vc.number_format = VOL4
        dr += 1

    data_end = dr - 1

    if data_end > TBL_R + 1:
        chart = LineChart()
        chart.title  = None
        chart.style  = 2
        chart.legend = None
        chart.y_axis.numFmt           = '#,##0.0000'
        chart.y_axis.delete           = False
        chart.y_axis.majorGridlines   = None
        chart.x_axis.tickLblPos       = "low"
        chart.x_axis.delete           = False
        chart.height = 10
        chart.width  = 28
        data_ref = Reference(ws, min_col=3, min_row=TBL_R + 1, max_row=data_end)
        chart.add_data(data_ref)
        cats = Reference(ws, min_col=2, min_row=TBL_R + 1, max_row=data_end)
        chart.set_categories(cats)
        s = chart.series[0]
        s.graphicalProperties.line.solidFill        = ACCENT
        s.graphicalProperties.line.width            = 32000
        s.marker.symbol                              = "circle"
        s.marker.size                                = 8
        s.marker.graphicalProperties.solidFill      = GOLD
        s.marker.graphicalProperties.line.solidFill = ACCENT
        ws.add_chart(chart, f"E{TBL_R}")

    for r in range(TBL_R, data_end + 22):
        for ci in range(1, 18):
            cell = ws.cell(row=r, column=ci)
            if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                cell.fill = _fill(BG)

    foot_r = max(data_end + 22, FR + 6)
    _bg_row(foot_r, 8, BG)
    _bg_row(foot_r+1, 2, ACCENT)
    _bg_row(foot_r+2, 28, SURFACE2)
    _merge(foot_r+2, FULL_START, FULL_END,
           "  SCG = RPV + RET + RP    ·    RPV = CGR − CGF    ·    "
           "PR = (SCG + ΣSR) ÷ VP    ·    PV = PMPV + PR",
           SURFACE2, _font(italic=True, size=9, color=MUTED), "center", "@", 28)
    _bg_row(foot_r+3, 22, SURFACE2)
    _merge(foot_r+3, FULL_START, FULL_END,
           f"ARPE · Conta Gráfica · {datetime.now().year}",
           SURFACE2, _font(size=8, color=DIM), "center", "@", 22)
