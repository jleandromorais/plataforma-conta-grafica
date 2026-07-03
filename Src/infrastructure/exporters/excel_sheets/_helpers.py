"""Constantes e helpers de formatação compartilhados por todos os módulos de sheets."""
from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter

from Src.infrastructure.exporters.excel_styles import (
    fill as _fill, font as _font, border as _border, align as _align,
    to_float as _to_float,
)

# ── Formatos numéricos ────────────────────────────────────────────────────────
BRL  = 'R$ #,##0.00'
VOL  = '#,##0.00'
VOL4 = '#,##0.0000'
NUM  = '#,##0'

# ── Paleta de cores ───────────────────────────────────────────────────────────
NAVY       = "1A3A5C"
BLUE       = "2E86C1"
TEAL       = "0E6655"
ORANGE     = "D35400"
PURPLE     = "6C3483"
GOLD       = "B7950B"
GREEN      = "1E8449"
RED        = "C0392B"
HEADER_FG  = "FFFFFF"
ROW_ALT    = "EBF5FB"
ROW_NORM   = "FFFFFF"
SUMMARY    = "FEF9E7"


def apply_header_row(ws, row_num: int, labels: list[str],
                     widths: list[int], bg: str = NAVY) -> None:
    for col_idx, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=HEADER_FG, size=11)
        cell.alignment = _align("center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_data_row(ws, row_num: int, values: list[Any],
                   fmts: list[str] | None = None,
                   alternate: bool = False,
                   bold_last: bool = False) -> None:
    bg = ROW_ALT if alternate else ROW_NORM
    fmts = fmts or ["@"] * len(values)
    for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell_value = _to_float(val) if fmt != "@" else ("" if val is None else val)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.value = cell_value
        cell.fill = _fill(bg)
        cell.font = _font(bold=(bold_last and col_idx == len(values)))
        cell.alignment = _align("right" if fmt != "@" else "left")
        cell.number_format = fmt
        cell.border = _border()


def apply_total_row(ws, row_num: int, values: list[Any],
                    fmts: list[str] | None = None, bg: str = SUMMARY) -> None:
    fmts = fmts or ["@"] * len(values)
    for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell_value = _to_float(val) if fmt != "@" else ("" if val is None else val)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.value = cell_value
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, size=11)
        cell.alignment = _align("right" if fmt != "@" else "left")
        cell.number_format = fmt
        cell.border = _border()


def section_title(ws, row_num: int, text: str, ncols: int, bg: str) -> None:
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=ncols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.fill = _fill(bg)
    cell.font = _font(bold=True, color=HEADER_FG, size=12)
    cell.alignment = _align("left")
    cell.border = _border()
