"""
Exportador do Relatório Consolidado da Conta Gráfica.

Gera um único arquivo .xlsx com todas as etapas do processo:
  📋 Resumo Executivo  |  📊 PMPV  |  🔍 Auditoria CGR
  ⚡ RET               |  📄 Conciliação RP  |  📋 CGF
  🧾 RPV               |  🧾 SCG Final
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import DataPoint

from Src.Database.database import DatabasePMPV
from Src.common.formatting import format_brl, format_brl4, format_brl_plain
from Src.common.periodos import TRIMESTRES_FISCAIS_ABREVS, TRIMESTRES_CIVIS
from Src.infrastructure.exporters.excel_styles import (
    fill as _fill, font as _font, border as _border,
    align as _align, to_float as _to_float,
)
from Src.infrastructure.exporters.excel_sheets import (
    sheet_sr as _sheet_sr_fn,
    sheet_pr as _sheet_pr_fn,
    sheet_pv as _sheet_pv_fn,
    sheet_progresso as _sheet_progresso_fn,
    sheet_dashboard as _sheet_dashboard_fn,
)

# ── Paleta de cores ───────────────────────────────────────────────────────────
_NAVY      = "1A3A5C"   # cabeçalhos principais
_BLUE      = "2E86C1"   # módulo Auditoria
_TEAL      = "0E6655"   # módulo PMPV
_ORANGE    = "D35400"   # módulo RET
_PURPLE    = "6C3483"   # módulo Conciliação
_GOLD      = "B7950B"   # módulo CGF / SCG
_GREEN     = "1E8449"   # positivo
_RED       = "C0392B"   # negativo
_HEADER_FG = "FFFFFF"   # texto em cabeçalho escuro
_ROW_ALT   = "EBF5FB"   # linha alternada clara
_ROW_NORM  = "FFFFFF"   # linha normal
_SUMMARY   = "FEF9E7"   # linha de totais
_TITLE_BG  = "1A3A5C"   # fundo do título principal


def _vol_fmt(val: Any) -> str:
    return format_brl_plain(_to_float(val))


def _money_fmt(val: Any) -> str:
    return format_brl(_to_float(val))


def _money4_fmt(val: Any) -> str:
    return format_brl4(_to_float(val))


def _apply_header_row(ws, row_num: int, labels: list[str],
                      widths: list[int], bg: str = _NAVY):
    for col_idx, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=_HEADER_FG, size=11)
        cell.alignment = _align("center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_data_row(ws, row_num: int, values: list[Any],
                    fmts: list[str] | None = None,
                    alternate: bool = False,
                    bold_last: bool = False):
    bg = _ROW_ALT if alternate else _ROW_NORM
    fmts = fmts or ["@"] * len(values)
    for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell_value = _to_float(val) if fmt != "@" else ("" if val is None else val)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.value = cell_value
        cell.fill = _fill(bg)
        cell.font = _font(bold=(bold_last and col_idx == len(values)))
        cell.alignment = _align("right" if fmt != "@" else "left")
        cell.number_format = fmt
        cell.border = _border()


def _apply_total_row(ws, row_num: int, values: list[Any],
                     fmts: list[str] | None = None, bg: str = _SUMMARY):
    fmts = fmts or ["@"] * len(values)
    for col_idx, (val, fmt) in enumerate(zip(values, fmts), start=1):
        cell_value = _to_float(val) if fmt != "@" else ("" if val is None else val)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.value = cell_value
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, size=11)
        cell.alignment = _align("right" if fmt != "@" else "left")
        cell.number_format = fmt
        cell.border = _border()


def _section_title(ws, row_num: int, text: str, ncols: int, bg: str):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=ncols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.fill = _fill(bg)
    cell.font = _font(bold=True, color=_HEADER_FG, size=12)
    cell.alignment = _align("left")
    cell.border = _border()


_BRL = 'R$ #,##0.00'
_VOL = '#,##0.00'
_VOL4 = '#,##0.0000'

# Mapa de abreviatura → nome completo do mês (PT-BR)
_MESES_FULL: dict[str, str] = {
    "Jan": "Janeiro",  "Fev": "Fevereiro", "Mar": "Março",
    "Abr": "Abril",    "Mai": "Maio",       "Jun": "Junho",
    "Jul": "Julho",    "Ago": "Agosto",     "Set": "Setembro",
    "Out": "Outubro",  "Nov": "Novembro",   "Dez": "Dezembro",
}

# Ordem numérica de cada abreviatura
_MES_ORD: dict[str, int] = {ab: i for i, ab in enumerate(_MESES_FULL, 1)}

# Janelas de meses válidos por módulo (abreviaturas PT-BR)
_MESES_PMPV = {"Fev", "Mar", "Abr"}   # Fevereiro a Abril
_MESES_RET  = {"Jan", "Fev", "Mar"}   # Janeiro a Março
# RPV (CGR + CGF): sem restrição de mês — usa todos disponíveis

def _abrev_de_periodo(periodo: str) -> str:
    """Extrai a abreviatura do mês de 'Mmm/YYYY' → 'Mmm'."""
    return periodo.split("/")[0].capitalize()[:3] if "/" in periodo else periodo[:3].capitalize()

def _nome_mes_completo(periodo: str) -> str:
    """'Abr/2026' → 'Abril/2026'."""
    ab = _abrev_de_periodo(periodo)
    ano = periodo.split("/")[1] if "/" in periodo else ""
    nome = _MESES_FULL.get(ab, ab)
    return f"{nome}/{ano}" if ano else nome

def _chave_trimestre_civil(periodo: str) -> tuple[str, int] | None:
    """Dado 'Mmm/YYYY', retorna (ano, índice do bloco civil 0-3) — ex: 'Jan/2026' → ('2026', 0).
    Usada para agrupar meses no bloco civil correto (Jan-Mar, Abr-Jun, Jul-Set, Out-Dez),
    nunca misturando meses de anos ou trimestres civis diferentes no mesmo banner.
    """
    if not periodo or "/" not in periodo:
        return None
    abrev, ano = _abrev_de_periodo(periodo), periodo.split("/")[1]
    for idx, meses_abrev in enumerate(TRIMESTRES_CIVIS.values()):
        if abrev in meses_abrev:
            return (ano, idx)
    return None

def _agrupar_em_trimestres(meses: list[str]) -> list[list[str]]:
    """Agrupa meses ordenados em blocos de trimestre civil (Jan-Mar, Abr-Jun,
    Jul-Set, Out-Dez), nunca misturando meses de anos ou blocos diferentes no
    mesmo grupo — mesmo que a lista de entrada não venha em múltiplos de 3.
    Ex: ['Dez/25','Jan/26','Fev/26','Mar/26'] →
        [['Dez/25'], ['Jan/26','Fev/26','Mar/26']]
    """
    grupos: list[list[str]] = []
    chave_atual = None
    for m in meses:
        chave = _chave_trimestre_civil(m)
        if chave != chave_atual or not grupos:
            grupos.append([])
            chave_atual = chave
        grupos[-1].append(m)
    return grupos

def _trimestre_civil_de(periodo: str | None) -> list[str]:
    """Dado um período 'Mmm/YYYY', retorna os 3 meses do trimestre civil fixo
    a que ele pertence (Jan-Mar, Abr-Jun, Jul-Set ou Out-Dez do mesmo ano).
    """
    if not periodo or "/" not in periodo:
        return []
    abrev, ano = _abrev_de_periodo(periodo), periodo.split("/")[1]
    for meses_abrev in TRIMESTRES_CIVIS.values():
        if abrev in meses_abrev:
            return [f"{m}/{ano}" for m in meses_abrev]
    return []

_NUM = '#,##0'


# ══════════════════════════════════════════════════════════════════════════════
# CLASSE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

class ExcelConsolidado:
    """Gera o Relatório Consolidado completo a partir do banco de dados."""

    @staticmethod
    def _agregar_consolidacao(registros: list[dict]) -> dict:
        total = {
            "cgr": 0.0,
            "cgf": 0.0,
            "ret": 0.0,
            "rp": 0.0,
            "rpv": 0.0,
            "scg": 0.0,
        }
        for registro in registros:
            for chave in total:
                total[chave] += registro.get(chave, 0.0) or 0.0
        return total

    @staticmethod
    def exportar(
        periodo: str | None = None,
        nome_arquivo: str | None = None,
        periodos_trimestre: list[str] | None = None,
    ) -> str:
        """
        Gera o Excel consolidado.

        Args:
            periodo: Período de referência principal (ex: 'Abr/2025') — usado
                     para PMPV, CGF, PR, PV, SR e SCG.
            periodos_trimestre: Lista dos 3 meses do trimestre (ex:
                     ['Fev/25', 'Mar/25', 'Abr/25']). Quando fornecida,
                     Auditoria, RET e Conciliação são agregados dos 3 meses.
                     Se None, usa apenas `periodo`.
            nome_arquivo: Caminho de saída. Se None, gera automaticamente.
        """
        with DatabasePMPV() as db:
            return ExcelConsolidado._gerar(db, periodo, nome_arquivo, periodos_trimestre)

    # ── Gerador principal ────────────────────────────────────────────────────

    @staticmethod
    def _gerar(
        db: DatabasePMPV,
        periodo: str | None,
        nome_arquivo: str | None,
        periodos_trimestre: list[str] | None = None,
    ) -> str:
        # Abreviaturas de mês válidas PT-BR
        _ABREVS_VALIDAS = {"Jan","Fev","Mar","Abr","Mai","Jun",
                           "Jul","Ago","Set","Out","Nov","Dez"}

        def _mes_valido(p: str) -> bool:
            """Aceita apenas 'Mmm/YYYY' (ex: 'Jan/2026'). Rejeita 'RET/25', 'Q1/2026', etc."""
            partes = (p or "").strip().split("/")
            if len(partes) != 2:
                return False
            abrev, ano = partes[0].strip().capitalize()[:3], partes[1].strip()
            return abrev in _ABREVS_VALIDAS and len(ano) == 4 and ano.isdigit()

        # Lista de meses do trimestre para Auditoria / RET / Conciliação
        meses_raw = periodos_trimestre or ([periodo] if periodo else [])
        meses = [m for m in meses_raw if _mes_valido(m)]
        # Se filtragem removeu tudo, usa os originais (melhor exibir algo do que nada)
        if not meses and meses_raw:
            meses = meses_raw

        # Usa nomes completos de mês nos títulos (ex: "Fevereiro/2026  ·  Março/2026  ·  Abril/2026")
        label_trimestre = "  ·  ".join(_nome_mes_completo(m) for m in meses) if meses else (periodo or "completo")

        p_slug = label_trimestre.replace("/", "-").replace("  ·  ", "_")

        if nome_arquivo is None:
            nome_arquivo = f"Relatorio_ContaGrafica_{p_slug}.xlsx"

        final = str(Path(nome_arquivo))
        Path(final).parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # ── PMPV / PR / PV / SR / SCG — filtrado pelo período principal
        cons_periodos = db.listar_consolidacao_completa(periodo) if periodo else db.listar_consolidacao_completa()
        cons          = db.buscar_consolidacao(periodo) if periodo else ExcelConsolidado._agregar_consolidacao(cons_periodos)
        # PMPV: usa apenas a sessão mais recente (evita acúmulo de sessões antigas)
        _todas_sessoes = db.listar_sessoes_com_volumes()
        pmpv_sessoes   = _todas_sessoes[:1] if _todas_sessoes else []
        sr            = db.buscar_sr(periodo) if periodo else None
        sr_lista      = [sr] if periodo and sr else db.listar_sr()
        pr            = db.buscar_pr(periodo) if periodo else None
        pr_lista      = [pr] if periodo and pr else db.listar_pr()
        pv            = db.buscar_pv(periodo) if periodo else None
        pv_lista      = [pv] if periodo and pv else db.listar_pv()
        pmpv_mensal   = db.listar_pmpv_mensal()
        execucoes     = db.listar_execucoes_excel_final(periodo=periodo) if periodo else db.listar_execucoes_excel_final()

        # ── Helper: períodos válidos disponíveis em cada tabela ─────────────────
        def _ord(p):
            ab = _abrev_de_periodo(p)
            an = p.split("/")[1] if "/" in p else "0"
            return int(an) * 12 + _MES_ORD.get(ab, 0)

        def _periodos_tabela(
            tabela: str,
            col_periodo: str = "periodo",
            filtro_modulo: set[str] | None = None,
            filtro_periodos: list[str] | None = None,
        ) -> list[str]:
            """Retorna períodos Mmm/YYYY distintos da tabela.

            filtro_modulo: conjunto de abreviaturas permitidas (ex: _MESES_RET).
            filtro_periodos: lista exata de períodos do trimestre civil
                (ex: ['Jan/2026','Fev/2026','Mar/2026']) — usada por módulos
                que devem seguir a grade civil fixa. `None` = sem filtro;
                lista vazia = filtro ativo que não casa com nada (retorna []).
            Se nenhum filtro for informado, retorna todos os meses válidos do banco.
            """
            try:
                rows = [r for r in db.listar_periodos_distintos(tabela, col_periodo) if _mes_valido(r)]
            except Exception:
                return []
            if filtro_modulo:
                # Módulo com janela fixa (RET=Jan-Mar, PMPV=Fev-Abr):
                # filtra só pelas abreviaturas permitidas.
                rows = [r for r in rows if _abrev_de_periodo(r) in filtro_modulo]
            if filtro_periodos is not None:
                # Segue a grade civil fixa (Jan-Mar, Abr-Jun, Jul-Set, Out-Dez).
                rows = [r for r in rows if r in filtro_periodos]
            return sorted(rows, key=_ord)

        # Cada módulo usa apenas os meses da sua janela de negócio:
        # RET  = Janeiro–Março (fixo)  |  PMPV = Fevereiro–Abril (fixo)
        # Auditoria (CGR) e CGF: mostram TODOS os meses disponíveis no banco
        # (todos os anos/trimestres); o agrupamento em banners dentro da aba
        # respeita a grade civil (Jan-Mar, Abr-Jun, Jul-Set, Out-Dez) via
        # _agrupar_em_trimestres, nunca misturando meses de blocos diferentes.
        meses_ret   = _periodos_tabela("ret_itens",       filtro_modulo=_MESES_RET)
        meses_audit = _periodos_tabela("auditoria_itens")   # Auditoria/CGR — sem restrição
        meses_cgf   = _periodos_tabela("cgf_resumo")        # CGF — sem restrição
        meses_conc  = _periodos_tabela("concilia_itens")

        # Labels de cada módulo para o título da aba
        def _label(ms): return "  ·  ".join(ms) if ms else label_trimestre

        # Buscar dados de cada módulo nos seus próprios períodos
        audit_itens = []
        for mes in meses_audit:
            audit_itens.extend(db.listar_auditoria_itens(mes) or [])

        ret_itens = []
        for mes in meses_ret:
            ret_itens.extend(db.listar_ret_itens(mes) or [])

        conc_itens = []
        for mes in meses_conc:
            conc_itens.extend(db.listar_concilia_itens(mes) or [])

        cgf_lista = []
        for mes in meses_cgf:
            resumo = db.buscar_cgf_resumo(mes)
            if resumo:
                cgf_lista.append(resumo)

        # Fallback CGF: se não encontrou nada pelos períodos, tenta tudo
        if not cgf_lista:
            cgf_lista = db.listar_cgf_resumos()

        # CGF de referência = último mês com dados
        cgf = cgf_lista[-1] if cgf_lista else None

        # PMPV: usa os próprios meses do trimestre informado (sem filtro fixo de abreviatura)
        meses_pmpv = meses if meses else []

        # ── Sheets na ordem ideal: Dashboard primeiro, módulos por trimestre, fechamento
        ExcelConsolidado._sheet_dashboard(
            wb, cons, cons_periodos, pr if periodo else (pr_lista[0] if pr_lista else None),
            pv if periodo else (pv_lista[0] if pv_lista else None),
            sr if periodo else (sr_lista[0] if sr_lista else None),
            label_trimestre,
        )
        ExcelConsolidado._sheet_resumo(wb, cons, cons_periodos, pmpv_sessoes, cgf_lista, sr_lista, pr_lista, pv_lista, label_trimestre)
        ExcelConsolidado._sheet_pmpv(wb, db, pmpv_sessoes, meses_pmpv)
        ExcelConsolidado._sheet_auditoria(wb, audit_itens, _label(meses_audit), meses_audit)
        ExcelConsolidado._sheet_ret(wb, ret_itens, _label(meses_ret), meses_ret)
        ExcelConsolidado._sheet_concilia(wb, conc_itens, _label(meses_conc))
        ExcelConsolidado._sheet_cgf(wb, cgf_lista, _label(meses_cgf), meses_cgf, db)
        ExcelConsolidado._sheet_rpv(wb, cons_periodos, meses_audit, meses_cgf, db, periodo)
        ExcelConsolidado._sheet_scg_mensal(wb, db, cons, cons_periodos, sr if periodo else sr_lista, periodo)
        ExcelConsolidado._sheet_scg_trimestral(wb, db)
        ExcelConsolidado._sheet_sr(wb, sr if periodo else sr_lista, periodo)
        ExcelConsolidado._sheet_pr(wb, pr if periodo else pr_lista, periodo)
        ExcelConsolidado._sheet_pv(wb, pv if periodo else pv_lista, periodo)
        ExcelConsolidado._sheet_progresso(wb, execucoes, periodo)

        try:
            wb.save(final)
        except PermissionError:
            wb.close()
            from tkinter import messagebox
            nome_arquivo = Path(final).name
            messagebox.showerror(
                "Arquivo em uso",
                f"Não foi possível salvar '{nome_arquivo}'.\n\n"
                f"O arquivo está aberto no Excel. Feche-o e tente gerar o relatório novamente."
            )
            raise PermissionError(f"Feche o arquivo '{nome_arquivo}' no Excel e tente novamente.")

        wb.close()

        try:
            os.startfile(final)
        except Exception:
            pass

        return final

    # ── Sheet 1: Resumo Executivo ────────────────────────────────────────────

    @staticmethod
    def _sheet_resumo(wb, cons, cons_periodos, pmpv_sessoes, cgf_lista, sr_lista, pr_lista, pv_lista, periodo):
        ws = wb.create_sheet("📋 Resumo Executivo")
        ws.sheet_view.showGridLines     = False
        ws.sheet_view.showRowColHeaders = False
        ws.sheet_properties.tabColor    = "1A3A5C"

        # ── Paleta consistente com Dashboard ──────────────────────────────────
        BG       = "F8FAFC"   # fundo claro (papel)
        SURFACE  = "FFFFFF"
        ACCENT   = "0E7C7B"   # teal escuro
        NAVY     = "1A3A5C"
        GOLD     = "B7950B"
        BLUE     = "2E86C1"
        PURPLE   = "6C3483"
        ORANGE   = "D35400"
        GREEN    = "1E8449"
        RED      = "C0392B"
        MUTED    = "5A6B85"
        DIM      = "8896B0"
        BORDER   = "DDE3EC"
        WHITE    = "FFFFFF"

        # Larguras
        ws.column_dimensions["A"].width = 1.5
        for col in "BCDEFGH":
            ws.column_dimensions[col].width = 16
        ws.column_dimensions["I"].width = 1.5

        FULL_START = 2
        FULL_END   = 8

        NONE_BDR = Border()

        def _rh(r, h):
            ws.row_dimensions[r].height = h

        def _bg_row(r, h, bg, c1=1, c2=9):
            _rh(r, h)
            for ci in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=ci)
                cell.fill = _fill(bg)
                cell.border = NONE_BDR

        def _merge(r, c1, c2, value, bg, fnt, align_h="center", fmt="@", row_h=None):
            if c1 != c2:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            cell = ws.cell(row=r, column=c1, value=value)
            cell.fill = _fill(bg)
            cell.font = fnt
            cell.alignment = _align(align_h, "center")
            cell.border = NONE_BDR
            if fmt != "@":
                cell.number_format = fmt
            if row_h:
                _rh(r, row_h)
            return cell

        def _fill_range(r, c1, c2, bg):
            for ci in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=ci)
                cell.fill = _fill(bg)
                cell.border = NONE_BDR

        # Fundo
        for r in range(1, 60):
            _bg_row(r, ws.row_dimensions[r].height or 15, BG)

        # ══════════════════════════════════════════════════════════════════════
        # HEADER
        # ══════════════════════════════════════════════════════════════════════
        _bg_row(1, 4, ACCENT)

        _bg_row(2, 50, NAVY)
        _merge(2, FULL_START, 5,
               "   RELATÓRIO CONSOLIDADO",
               NAVY, _font(bold=True, size=18, color=WHITE), "left")
        _merge(2, 6, FULL_END,
               f"{periodo or 'Todos os Períodos'}  ",
               NAVY, _font(bold=True, size=13, color="A8D5E2"), "right")

        _bg_row(3, 22, "2E4057")
        _merge(3, FULL_START, FULL_END,
               f"   Conta Gráfica · Tarifa de Gás Canalizado · "
               f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
               "2E4057", _font(size=10, color="CCDDEE", italic=True), "left")

        _bg_row(4, 14, BG)

        # ══════════════════════════════════════════════════════════════════════
        # CARDS DE RESUMO
        # ══════════════════════════════════════════════════════════════════════
        def _resumo_card(row, c1, c2, icon, label, value, fmt, color):
            """Card com label/valor em layout vertical clean."""
            # Top stripe
            _fill_range(row, c1, c2, color)
            _rh(row, 3)

            # Label
            _merge(row+1, c1, c2, f" {icon}  {label}",
                   SURFACE, _font(bold=True, size=10, color=MUTED), "left", "@", 22)
            # Borders esquerda/direita
            for cc in (ws.cell(row=row+1, column=c1), ws.cell(row=row+1, column=c2)):
                cc.border = Border(
                    left=Side(style="thin", color=BORDER),
                    right=Side(style="thin", color=BORDER))

            # Valor
            _merge(row+2, c1, c2, value,
                   SURFACE, _font(bold=True, size=15, color=color), "center", fmt, 32)
            for cc in (ws.cell(row=row+2, column=c1), ws.cell(row=row+2, column=c2)):
                cc.border = Border(
                    left=Side(style="thin", color=BORDER),
                    right=Side(style="thin", color=BORDER))

            # Bottom border
            _fill_range(row+3, c1, c2, BORDER)
            _rh(row+3, 1)

        sr_total = (sr_lista[0] if periodo and sr_lista else None) or None

        cgr_val = (cons or {}).get("cgr", 0.0)
        cgf_val = (cons or {}).get("cgf", 0.0)
        rpv_val = (cons or {}).get("rpv", cgr_val - cgf_val)
        ret_val = (cons or {}).get("ret", 0.0)
        rp_val  = (cons or {}).get("rp", 0.0)
        scg_val = (cons or {}).get("scg", 0.0)
        sr_val  = sum((item or {}).get("sr", 0.0) or 0.0 for item in sr_lista) if not periodo else _to_float((sr_total or {}).get("sr", 0.0))
        pr_ref  = (pr_lista[0] if pr_lista else {}) or {}
        pv_ref  = (pv_lista[0] if pv_lista else {}) or {}

        # Linha 1: CGR, CGF, RPV
        R = 5
        _resumo_card(R, 2, 3, "🔍", "CGR · Auditoria",         cgr_val, _BRL, BLUE)
        _resumo_card(R, 4, 5, "📋", "CGF · Volume × PMPV",     cgf_val, _BRL, GOLD)
        _resumo_card(R, 6, 7, "🧾", "RPV · CGR − CGF",         rpv_val, _BRL, PURPLE)

        _bg_row(R+4, 8, BG)
        # Linha 2: RET, RP, SR
        R = R + 5
        _resumo_card(R, 2, 3, "⚡", "RET · Encargos",          ret_val, _BRL, ORANGE)
        _resumo_card(R, 4, 5, "📄", "RP · Conciliação",        rp_val,  _BRL, ACCENT)
        _resumo_card(R, 6, 7, "📈", "SR · Saldo Remanesc.",    sr_val,  _BRL, NAVY)

        # Espaço
        _bg_row(R+4, 16, BG)

        # ══════════════════════════════════════════════════════════════════════
        # SCG FINAL — destaque grande
        # ══════════════════════════════════════════════════════════════════════
        R = R + 5
        scg_color = GREEN if scg_val >= 0 else RED

        _bg_row(R, 4, scg_color)
        _bg_row(R+1, 24, NAVY)
        _merge(R+1, FULL_START, FULL_END,
               "   💼  SCG FINAL  =  RPV  +  RET  +  RP",
               NAVY, _font(bold=True, size=12, color=WHITE), "left", "@", 24)

        _bg_row(R+2, 50, scg_color)
        _merge(R+2, FULL_START, FULL_END, scg_val,
               scg_color, _font(bold=True, size=24, color=WHITE), "center", _BRL, 50)

        _bg_row(R+3, 8, BG)

        # ══════════════════════════════════════════════════════════════════════
        # PR e PV
        # ══════════════════════════════════════════════════════════════════════
        R = R + 5
        _resumo_card(R, 2, 4, "💡", "PR FINAL · (SCG + ΣSR) ÷ VP",
                     _to_float(pr_ref.get("pr")), _VOL4, ACCENT)
        _resumo_card(R, 5, 7, "💰", "PV FINAL · PMPV + PR",
                     _to_float(pv_ref.get("pv")), _VOL4, GREEN)

        # ══════════════════════════════════════════════════════════════════════
        # PMPV — sessões salvas
        # ══════════════════════════════════════════════════════════════════════
        R = R + 5
        _bg_row(R, 4, BG)
        _bg_row(R+1, 26, ACCENT)
        _merge(R+1, FULL_START, FULL_END,
               "   📊  PMPV — Sessões Salvas",
               ACCENT, _font(bold=True, size=11, color=WHITE), "left", "@", 26)

        if pmpv_sessoes:
            R = R + 2
            _apply_header_row(ws, R,
                [None, "Sessão", "Data", "VP (m³)", "VF (m³)", "PMPV", "Final", None],
                [1, 25, 16, 14, 14, 12, 12, 1], NAVY)
            for i, s in enumerate(pmpv_sessoes[:8]):
                R += 1
                row_bg = SURFACE if i % 2 == 0 else "F0F4F8"
                vals = ["", s.get("nome", ""), s.get("data_criacao", ""),
                        _to_float(s.get("vp")), _to_float(s.get("vf")), "", "", ""]
                fmts = ["@", "@", "@", _VOL, _VOL, _BRL, _BRL, "@"]
                for ci, (v, f) in enumerate(zip(vals, fmts), start=1):
                    cell = ws.cell(row=R, column=ci, value=v)
                    cell.fill = _fill(row_bg)
                    cell.font = _font(size=9, color=MUTED)
                    cell.alignment = _align("right" if f != "@" else "left")
                    cell.border = NONE_BDR
                    if f != "@":
                        cell.number_format = f
                _rh(R, 18)
        else:
            R = R + 2
            _bg_row(R, 22, SURFACE)
            _merge(R, FULL_START, FULL_END,
                   "   Nenhuma sessão PMPV salva ainda.",
                   SURFACE, _font(italic=True, size=10, color=DIM), "left", "@", 22)

        # ══════════════════════════════════════════════════════════════════════
        # CONSOLIDAÇÃO POR PERÍODO (se múltiplos períodos)
        # ══════════════════════════════════════════════════════════════════════
        if not periodo and cons_periodos:
            R = R + 2
            _bg_row(R, 12, BG)
            R += 1
            _bg_row(R, 26, NAVY)
            _merge(R, FULL_START, FULL_END,
                   "   📚  Consolidação por Período",
                   NAVY, _font(bold=True, size=11, color=WHITE), "left", "@", 26)
            R += 1

            _apply_header_row(ws, R,
                [None, "Período", "CGR", "CGF", "RPV", "RET", "RP", "SCG"],
                [1, 14, 14, 14, 14, 14, 14, 14], NAVY)

            for i, item in enumerate(cons_periodos):
                R += 1
                row_bg = SURFACE if i % 2 == 0 else "F0F4F8"
                scg_item = _to_float(item.get("scg"))
                vals = ["", item.get("periodo", ""),
                        _to_float(item.get("cgr")), _to_float(item.get("cgf")),
                        _to_float(item.get("rpv")), _to_float(item.get("ret")),
                        _to_float(item.get("rp")),  scg_item]
                fmts = ["@", "@", _BRL, _BRL, _BRL, _BRL, _BRL, _BRL]
                for ci, (v, f) in enumerate(zip(vals, fmts), start=1):
                    cell = ws.cell(row=R, column=ci, value=v)
                    cell.fill = _fill(row_bg)
                    cell.font = _font(
                        size=9,
                        color=(GREEN if (ci == 8 and scg_item >= 0)
                               else RED if (ci == 8 and scg_item < 0) else MUTED),
                        bold=(ci == 8))
                    cell.alignment = _align("right" if f != "@" else "left")
                    cell.border = NONE_BDR
                    if f != "@":
                        cell.number_format = f
                _rh(R, 18)

        # Rodapé
        R = R + 3
        _bg_row(R, 6, BG)
        _bg_row(R+1, 2, ACCENT)
        _bg_row(R+2, 22, NAVY)
        _merge(R+2, FULL_START, FULL_END,
               f"   ARPE · Conta Gráfica · {datetime.now().year}",
               NAVY, _font(italic=True, size=9, color="CCDDEE"), "center", "@", 22)

    # ── Sheet 2: PMPV ────────────────────────────────────────────────────────

    @staticmethod
    def _sheet_pmpv(wb, db: DatabasePMPV, sessoes, meses_trimestre: list[str] | None = None):
        # Se não veio trimestre explícito, tenta o trimestre ativo salvo no banco
        if not meses_trimestre:
            meses_trimestre = db.buscar_trimestre_ativo() or []

        def _nome_mes(mes_num: int) -> str:
            """Converte posição 1/2/3 no nome real do mês do trimestre (ex: 'Fevereiro/2026')."""
            if meses_trimestre and mes_num <= len(meses_trimestre):
                return _nome_mes_completo(meses_trimestre[mes_num - 1])
            # fallback: mostra a posição numerada apenas se não há trimestre definido
            return f"Mês {mes_num} (trimestre não definido)"

        ws = wb.create_sheet("📊 PMPV")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = "GESTÃO PMPV — VOLUMES E PREÇOS POR EMPRESA / MÊS"
        t.fill = _fill(_TEAL)
        t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("center")
        ws.row_dimensions[1].height = 30

        row = 3

        if not sessoes:
            ws.cell(row=row, column=1, value="Nenhuma sessão PMPV salva no banco de dados.")
            return

        _NCOLS_PMPV = 7
        _COLS_PMPV  = ["Empresa", "Molécula", "Transporte", "Logística",
                       "Preço Unit.", "Volume (m³/dia)", "Subtotal (R$)"]
        _WIDTHS_PMPV = [28, 14, 14, 14, 14, 16, 16]
        _FMTS_PMPV   = ["@", _VOL4, _VOL4, _VOL4, _VOL4, _VOL, _BRL]

        for sessao in sessoes:
            sid  = sessao["id"]
            nome = sessao.get("nome", f"Sessão {sid}")
            data_raw = sessao.get("data_criacao", "")
            # formata '2026-06-12 10:36:57' → '12/06/2026 10:36'
            try:
                from datetime import datetime as _dt
                data = _dt.strptime(data_raw[:16], "%Y-%m-%d %H:%M").strftime("%d/%m/%Y %H:%M")
            except Exception:
                data = data_raw

            # ── Cabeçalho da sessão ───────────────────────────────────────────
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS_PMPV)
            sh = ws.cell(row=row, column=1,
                value=f"  Sessão: {nome}  |  Criada em: {data}  |  "
                      f"VP: {_vol_fmt(sessao.get('vp', 0))} m³  |  "
                      f"VF: {_vol_fmt(sessao.get('vf', 0))} m³")
            sh.fill = _fill(_TEAL); sh.font = _font(bold=True, size=12, color=_HEADER_FG)
            sh.alignment = _align("left"); sh.border = _border()
            ws.row_dimensions[row].height = 22; row += 1
            ws.row_dimensions[row].height = 4; row += 1

            grand_vol = grand_sub = 0.0

            for mes_num in [1, 2, 3]:
                linhas = db.carregar_dados_mes(sid, mes_num)
                if not linhas:
                    continue

                nome_mes = _nome_mes(mes_num)

                # ── Divisor do mês (igual ao RET) ─────────────────────────────
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=_NCOLS_PMPV)
                sec = ws.cell(row=row, column=1,
                              value=f"  ── {nome_mes} " + "─" * 42)
                sec.fill = _fill("1A5276"); sec.font = _font(bold=True, color=_HEADER_FG, size=11)
                sec.alignment = _align("left"); sec.border = _border()
                ws.row_dimensions[row].height = 20; row += 1

                # ── Cabeçalho de colunas do mês ───────────────────────────────
                _apply_header_row(ws, row, _COLS_PMPV, _WIDTHS_PMPV, "2E86C1")
                row += 1

                sub_vol = sub_preco = 0.0
                for i, l in enumerate(linhas):
                    mol   = l.get("molecula", 0.0)
                    trans = l.get("transporte", 0.0)
                    log   = l.get("logistica", 0.0)
                    vol   = l.get("volume", 0.0)
                    preco = mol + trans + log
                    subtotal = preco * vol
                    sub_vol += vol; sub_preco += subtotal
                    _apply_data_row(ws, row,
                        [l.get("empresa", ""), mol, trans, log, preco, vol, subtotal],
                        _FMTS_PMPV, alternate=(i % 2 == 1))
                    row += 1

                # ── Subtotal do mês ───────────────────────────────────────────
                for ci in range(1, _NCOLS_PMPV + 1):
                    c = ws.cell(row=row, column=ci)
                    c.fill = _fill("D6EAF8"); c.border = _border()
                    c.font = _font(bold=True, size=11, color="1A5276")
                ws.cell(row=row, column=1,
                        value=f"SUBTOTAL {nome_mes.upper()}").alignment = _align("left")
                c6 = ws.cell(row=row, column=6, value=sub_vol)
                c6.number_format = _VOL; c6.alignment = _align("right")
                c6.fill = _fill("D6EAF8"); c6.border = _border()
                c6.font = _font(bold=True, size=11, color="1A5276")
                c7 = ws.cell(row=row, column=7, value=sub_preco)
                c7.number_format = _BRL; c7.alignment = _align("right")
                c7.fill = _fill("D6EAF8"); c7.border = _border()
                c7.font = _font(bold=True, size=11, color="1A5276")
                ws.row_dimensions[row].height = 20; row += 1
                ws.row_dimensions[row].height = 6; row += 1

                grand_vol += sub_vol; grand_sub += sub_preco

            # ── Total geral da sessão ─────────────────────────────────────────
            res_row_db = db.buscar_resultado_sessao(sid)
            pmpv_val = res_row_db.get("pmpv_trimestral", 0) if res_row_db else 0

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            tl = ws.cell(row=row, column=1,
                         value=f"  ══ TOTAL GERAL DA SESSÃO  |  PMPV: {_money4_fmt(pmpv_val)} /m³")
            tl.fill = _fill(_TEAL); tl.font = _font(bold=True, size=13, color=_HEADER_FG)
            tl.alignment = _align("left"); tl.border = _border()
            for ci in [2,3,4,5]:
                ws.cell(row=row,column=ci).fill=_fill(_TEAL); ws.cell(row=row,column=ci).border=_border()
            c6t = ws.cell(row=row, column=6, value=grand_vol)
            c6t.number_format=_VOL; c6t.alignment=_align("right")
            c6t.fill=_fill(_TEAL); c6t.border=_border(); c6t.font=_font(bold=True,size=13,color=_HEADER_FG)
            c7t = ws.cell(row=row, column=7, value=grand_sub)
            c7t.number_format=_BRL; c7t.alignment=_align("right")
            c7t.fill=_fill(_TEAL); c7t.border=_border(); c7t.font=_font(bold=True,size=13,color=_HEADER_FG)
            ws.row_dimensions[row].height = 26; row += 3

        for i, w in enumerate(_WIDTHS_PMPV, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Auditoria CGR ────────────────────────────────────────────────

    @staticmethod
    def _sheet_auditoria(wb, itens: list[dict], label_trimestre: str | None,
                         meses: list[str] | None = None):
        _NCOLS  = 10
        _COLS   = ["Empresa","Tipo","Número","Valor Bruto (R$)",
                   "ICMS (R$)","PIS (R$)","COFINS (R$)","Volume (m³)","CGR Líquido (R$)","Status"]
        _WIDTHS = [22, 8, 14, 18, 14, 14, 14, 14, 18, 10]
        _FMTS   = ["@","@","@",_BRL,_BRL,_BRL,_BRL,_VOL,_BRL,"@"]

        _BLUE_DARK  = "1A5276"
        _BLUE_MED   = "2E86C1"
        _BLUE_LIGHT = "D6EAF8"
        _BLUE_XL    = "EBF5FB"

        ws = wb.create_sheet("🔍 Auditoria CGR")
        ws.sheet_view.showGridLines = False

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOLS)
        t = ws.cell(row=1, column=1,
                    value=f"  AUDITORIA CGR — NF-e e CT-e  |  Trimestre: {label_trimestre or 'N/D'}")
        t.fill = _fill(_BLUE_DARK); t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("left"); ws.row_dimensions[1].height = 34

        ws.row_dimensions[2].height = 6
        for c in range(1, _NCOLS + 1):
            ws.cell(row=2, column=c).fill = _fill("A9CCE3")

        if not itens:
            ws.cell(row=4, column=1, value="Nenhum item de Auditoria CGR registrado para este trimestre.")
            return

        from collections import defaultdict
        grupos: dict[str, list] = defaultdict(list)
        for it in itens:
            grupos[it.get("periodo", "Sem Período")].append(it)

        ordem = list(meses or [])
        for p in grupos:
            if p not in ordem:
                ordem.append(p)

        grand = {"bruto":0.0,"icms":0.0,"pis":0.0,"cofins":0.0,"vol":0.0,"cgr":0.0}
        row = 3

        trimestres = _agrupar_em_trimestres(ordem)
        meses_por_trimestre = {m: i for i, tri in enumerate(trimestres) for m in tri}

        tri_atual = -1
        for periodo_mes in ordem:
            its = grupos.get(periodo_mes)
            if not its:
                continue

            # Banner de trimestre a cada novo grupo de 3 meses
            tri_idx = meses_por_trimestre.get(periodo_mes, -1)
            if tri_idx != tri_atual:
                tri_atual = tri_idx
                tri_meses = trimestres[tri_idx]
                label_tri = "  🗓  TRIMESTRE:  " + "  ·  ".join(tri_meses)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
                bt = ws.cell(row=row, column=1, value=label_tri)
                bt.fill = _fill(_BLUE_DARK); bt.font = _font(bold=True, color="F0F8FF", size=12)
                bt.alignment = _align("left"); bt.border = _border()
                ws.row_dimensions[row].height = 24; row += 1

            ano_mes  = periodo_mes.split("/")[1] if "/" in periodo_mes else ""
            mes_full = _nome_mes_completo(periodo_mes).split("/")[0]
            label_sec = f"  ── {mes_full}{'/'+ano_mes if ano_mes else ''} " + "─"*40

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
            sec = ws.cell(row=row, column=1, value=label_sec)
            sec.fill = _fill(_BLUE_MED); sec.font = _font(bold=True, color=_HEADER_FG, size=11)
            sec.alignment = _align("left"); sec.border = _border()
            ws.row_dimensions[row].height = 20; row += 1

            _apply_header_row(ws, row, _COLS, _WIDTHS, _BLUE_MED); row += 1

            sub = {"bruto":0.0,"icms":0.0,"pis":0.0,"cofins":0.0,"vol":0.0,"cgr":0.0}
            for i, it in enumerate(its):
                vb  = _to_float(it.get("valor_total"))
                ic  = _to_float(it.get("icms"))
                ps  = _to_float(it.get("pis"))
                cf  = _to_float(it.get("cofins"))
                vol = _to_float(it.get("volume_total"))
                cgr = _to_float(it.get("cgr_liquido"))
                sub["bruto"]+=vb; sub["icms"]+=ic; sub["pis"]+=ps
                sub["cofins"]+=cf; sub["vol"]+=vol; sub["cgr"]+=cgr
                _apply_data_row(ws, row,
                    [it.get("empresa",""),it.get("tipo",""),it.get("numero",""),
                     vb,ic,ps,cf,vol,cgr,"OK"], _FMTS, alternate=(i%2==1))
                row += 1

            # Subtotal do mês
            label_sub = f"SUBTOTAL {mes_full.upper()}{'/'+ano_mes if ano_mes else ''}"
            for ci in range(1, _NCOLS+1):
                c = ws.cell(row=row, column=ci)
                c.fill = _fill(_BLUE_LIGHT); c.border = _border()
                c.font = _font(bold=True, size=11, color=_BLUE_DARK)
            ws.cell(row=row, column=1, value=label_sub).alignment = _align("left")
            for ci, key, fmt in [(4,"bruto",_BRL),(5,"icms",_BRL),(6,"pis",_BRL),
                                  (7,"cofins",_BRL),(8,"vol",_VOL),(9,"cgr",_BRL)]:
                c = ws.cell(row=row, column=ci, value=sub[key])
                c.number_format=fmt; c.alignment=_align("right")
                c.fill=_fill(_BLUE_LIGHT); c.border=_border()
                c.font=_font(bold=True,size=11,color=_BLUE_DARK)
            for k in sub: grand[k] += sub[k]
            ws.row_dimensions[row].height = 20; row += 1
            ws.row_dimensions[row].height = 8; row += 1

        # Total Geral
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        tl = ws.cell(row=row, column=1, value="  ══ TOTAL GERAL DO TRIMESTRE ══")
        tl.fill=_fill(_BLUE_DARK); tl.font=_font(bold=True,size=13,color=_HEADER_FG)
        tl.alignment=_align("left"); tl.border=_border()
        for ci in [2,3]: ws.cell(row=row,column=ci).fill=_fill(_BLUE_DARK); ws.cell(row=row,column=ci).border=_border()
        for ci, key, fmt in [(4,"bruto",_BRL),(5,"icms",_BRL),(6,"pis",_BRL),
                              (7,"cofins",_BRL),(8,"vol",_VOL),(9,"cgr",_BRL)]:
            c = ws.cell(row=row, column=ci, value=grand[key])
            c.fill=_fill(_BLUE_DARK); c.font=_font(bold=True,size=13,color=_HEADER_FG)
            c.number_format=fmt; c.alignment=_align("right"); c.border=_border()
        for ci in [10]: ws.cell(row=row,column=ci).fill=_fill(_BLUE_DARK); ws.cell(row=row,column=ci).border=_border()
        ws.row_dimensions[row].height = 26

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["C"].width = 14
        for i, w in enumerate(_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: RET ────────────────────────────────────────────────────────

    @staticmethod
    def _sheet_ret(wb, itens: list[dict], label_trimestre: str | None,
                   meses: list[str] | None = None):
        _NCOLS = 9
        _COLS  = ["Arquivo", "Tipo Encargo", "Empresa", "Tipo Nota",
                  "Nº ND", "Vencimento", "Valor Total (R$)", "Moeda", "Contrib. EC"]
        _WIDTHS = [34, 16, 18, 12, 12, 14, 20, 8, 14]
        _FMTS   = ["@", "@", "@", "@", "@", "@", _BRL, "@", "@"]

        # Cores de destaque do módulo RET
        _ORANGE_DARK  = "A04000"   # título principal
        _ORANGE_MED   = "D35400"   # cabeçalho de mês / header de tabela
        _ORANGE_LIGHT = "FAD7A0"   # subtotal / total
        _ORANGE_XL    = "FEF5E7"   # resumo por tipo

        ws = wb.create_sheet("⚡ RET")
        ws.sheet_view.showGridLines = False

        # ── Linha 1: Título do trimestre ──────────────────────────────────────
        trimestre_label = label_trimestre or "N/D"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOLS)
        t = ws.cell(row=1, column=1,
                    value=f"  SISTEMA RET — ENCARGOS E DOCUMENTOS  |  Trimestre: {trimestre_label}")
        t.fill = _fill(_ORANGE_DARK)
        t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("left")
        ws.row_dimensions[1].height = 34

        # Linha 2: espaço
        ws.row_dimensions[2].height = 6
        for c in range(1, _NCOLS + 1):
            ws.cell(row=2, column=c).fill = _fill("F5CBA7")

        if not itens:
            ws.cell(row=4, column=1, value="Nenhum item RET registrado para este trimestre.")
            return

        # ── Agrupar itens por período (na ordem de meses do trimestre) ─────────
        from collections import defaultdict
        grupos: dict[str, list] = defaultdict(list)
        for it in itens:
            grupos[it.get("periodo", "Sem Período")].append(it)

        # Ordem: segue a lista de meses do trimestre; itens sem período vão ao fim
        ordem = list(meses or [])
        for p in grupos:
            if p not in ordem:
                ordem.append(p)

        tipos_total: dict[str, float] = {}
        grand_total = 0.0
        row = 3

        trimestres = _agrupar_em_trimestres(ordem)
        meses_por_trimestre = {m: i for i, tri in enumerate(trimestres) for m in tri}
        tri_atual = -1

        for periodo_mes in ordem:
            itens_mes = grupos.get(periodo_mes)
            if not itens_mes:
                continue

            # Banner de trimestre a cada novo grupo de 3 meses
            tri_idx = meses_por_trimestre.get(periodo_mes, -1)
            if tri_idx != tri_atual:
                tri_atual = tri_idx
                tri_meses = trimestres[tri_idx]
                label_tri = "  🗓  TRIMESTRE:  " + "  ·  ".join(tri_meses)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
                bt = ws.cell(row=row, column=1, value=label_tri)
                bt.fill = _fill(_ORANGE_DARK); bt.font = _font(bold=True, color="FFF8F0", size=12)
                bt.alignment = _align("left"); bt.border = _border()
                ws.row_dimensions[row].height = 24; row += 1

            # Nome completo do mês para o cabeçalho da seção
            ano_mes  = periodo_mes.split("/")[1] if "/" in periodo_mes else ""
            mes_full = _nome_mes_completo(periodo_mes).split("/")[0]
            label_mes = f"  ── {mes_full}{'/'+ano_mes if ano_mes else ''} " + "─" * 40

            # Cabeçalho da seção do mês (laranja médio)
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=_NCOLS)
            sec = ws.cell(row=row, column=1, value=label_mes)
            sec.fill = _fill(_ORANGE_MED)
            sec.font = _font(bold=True, color=_HEADER_FG, size=11)
            sec.alignment = _align("left")
            sec.border = _border()
            ws.row_dimensions[row].height = 20
            row += 1

            # Header das colunas
            _apply_header_row(ws, row, _COLS, _WIDTHS, _ORANGE_MED)
            row += 1

            _PIS_COFINS = 0.0925
            sub_eat = 0.0   # soma bruta EAT do mês
            sub_ec  = 0.0   # soma bruta EC_docs do mês
            for i, it in enumerate(itens_mes):
                vt    = _to_float(it.get("valor_total"))
                t_enc = it.get("tipo_encargo", "")
                if t_enc == "EAT":
                    sub_eat += vt
                elif t_enc == "EC":
                    sub_ec += vt
                tipos_total[t_enc] = tipos_total.get(t_enc, 0.0) + vt

                _apply_data_row(ws, row, [
                    it.get("arquivo", ""),
                    t_enc,
                    it.get("empresa", ""),
                    it.get("nota_tipo", ""),
                    it.get("numero_nd", ""),
                    it.get("data_vencimento", ""),
                    vt,
                    it.get("moeda", "BRL"),
                    it.get("contrib_ec", ""),
                ], _FMTS, alternate=(i % 2 == 1))
                row += 1

            # EC líquido do mês = EAT × (1 − PIS/COFINS) + EC_docs
            ec_liquido_mes = sub_eat * (1.0 - _PIS_COFINS) + sub_ec

            # Subtotal do mês — mostra EC líquido
            label_sub = f"SUBTOTAL EC LÍQUIDO — {mes_full.upper()}{'/'+ano_mes if ano_mes else ''}"
            for col_i in range(1, _NCOLS + 1):
                c = ws.cell(row=row, column=col_i)
                c.fill = _fill(_ORANGE_LIGHT)
                c.font = _font(bold=True, size=11, color=_ORANGE_DARK)
                c.border = _border()
                if col_i == 1:
                    c.value = label_sub
                    c.alignment = _align("left")
                elif col_i == 7:
                    c.value = ec_liquido_mes
                    c.number_format = _BRL
                    c.alignment = _align("right")
            ws.row_dimensions[row].height = 20
            grand_total += ec_liquido_mes
            row += 1

            # Espaço entre meses
            ws.row_dimensions[row].height = 8
            row += 1

        # ── Total Geral do Trimestre ──────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        tot_lbl = ws.cell(row=row, column=1, value="  ══ EC LÍQUIDO TOTAL DO TRIMESTRE (RET) ══")
        tot_lbl.fill = _fill(_ORANGE_DARK)
        tot_lbl.font = _font(bold=True, size=13, color=_HEADER_FG)
        tot_lbl.alignment = _align("left")
        tot_lbl.border = _border()
        tot_val = ws.cell(row=row, column=7, value=grand_total)
        tot_val.fill = _fill(_ORANGE_DARK)
        tot_val.font = _font(bold=True, size=13, color=_HEADER_FG)
        tot_val.number_format = _BRL
        tot_val.alignment = _align("right")
        tot_val.border = _border()
        for c in [2, 3, 4, 5, 6, 8, 9]:
            ws.cell(row=row, column=c).fill = _fill(_ORANGE_DARK)
            ws.cell(row=row, column=c).border = _border()
        ws.row_dimensions[row].height = 26
        row += 2

        # ── Resumo por Tipo de Encargo ────────────────────────────────────────
        _section_title(ws, row, "  ── RESUMO POR TIPO DE ENCARGO (TRIMESTRE)", _NCOLS, _ORANGE_MED)
        ws.row_dimensions[row].height = 20
        row += 1
        _apply_header_row(ws, row, ["Tipo de Encargo", "Total (R$)"], [24, 20], _ORANGE_MED)
        row += 1
        for tp, val in sorted(tipos_total.items()):
            for c in range(1, _NCOLS + 1):
                ws.cell(row=row, column=c).fill = _fill(_ORANGE_XL)
            ws.cell(row=row, column=1, value=tp).font = _font(bold=True, size=11)
            ws.cell(row=row, column=1).alignment = _align("left")
            ws.cell(row=row, column=1).border = _border()
            v_cell = ws.cell(row=row, column=2, value=val)
            v_cell.number_format = _BRL
            v_cell.font = _font(bold=True, size=11)
            v_cell.alignment = _align("right")
            v_cell.border = _border()
            row += 1

        # ── Larguras das colunas ──────────────────────────────────────────────
        for i, w in enumerate(_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 5: Conciliação RP ──────────────────────────────────────────────

    @staticmethod
    def _sheet_concilia(wb, itens: list[dict], periodo: str | None):
        ws = wb.create_sheet("📄 Conciliação RP")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = f"CONCILIAÇÃO RP — RECEITAS E DESPESAS (PDFs)  |  Período: {periodo or 'N/D'}"
        t.fill = _fill(_PURPLE)
        t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("center")
        ws.row_dimensions[1].height = 30

        row = 3
        mostrar_periodo = periodo is None
        cols   = (["Período"] if mostrar_periodo else []) + ["Arquivo", "Categoria", "Valor (R$)", "Status", "Método"]
        widths = ([14] if mostrar_periodo else []) + [40, 12, 18, 10, 30]
        _apply_header_row(ws, row, cols, widths, _PURPLE)
        row += 1

        if not itens:
            ws.cell(row=row, column=1,
                    value=f"Nenhum item de Conciliação registrado para o período '{periodo}'.")
            return

        tot_rec = tot_desp = 0.0
        for i, it in enumerate(itens):
            val  = it.get("valor", 0.0)
            cat  = it.get("categoria", "")
            if cat == "Receita":
                tot_rec += val
            else:
                tot_desp += val

            _apply_data_row(ws, row,
                ([it.get("periodo", "")] if mostrar_periodo else []) + [it.get("arquivo", ""), cat, val,
                 it.get("status", ""), it.get("metodo", "")],
                (["@"] if mostrar_periodo else []) + ["@", "@", _BRL, "@", "@"],
                alternate=(i % 2 == 1))
            row += 1

        saldo = tot_rec - tot_desp
        _apply_total_row(ws, row,
            (["TOTAL RECEITA"] if mostrar_periodo else []) + ["TOTAL RECEITA", "Receita", tot_rec, "", ""],
            bg="D5F5E3")
        row += 1
        _apply_total_row(ws, row,
            (["TOTAL DESPESA"] if mostrar_periodo else []) + ["TOTAL DESPESA", "Despesa", tot_desp, "", ""],
            bg="FADBD8")
        row += 1
        bg_saldo = _GREEN if saldo >= 0 else _RED
        c = ws.cell(row=row, column=1, value="SALDO  RP  =  Receita − Despesa")
        c.fill = _fill(bg_saldo)
        c.font = _font(bold=True, color=_HEADER_FG, size=12)
        c.border = _border()
        v = ws.cell(row=row, column=3, value=saldo)
        v.number_format = _BRL
        v.fill = _fill(bg_saldo)
        v.font = _font(bold=True, color=_HEADER_FG, size=12)
        v.border = _border()
        ws.cell(row=row, column=2).fill = _fill(bg_saldo)
        ws.cell(row=row, column=2).border = _border()

        ws.column_dimensions["A"].width = 40
        for col in "BCDEF":
            ws.column_dimensions[col].width = 18

    # ── Sheet 6: CGF ────────────────────────────────────────────────────────

    @staticmethod
    def _sheet_cgf(wb, cgf_lista: list[dict] | None, label_trimestre: str | None,
                   meses: list[str] | None = None, db=None):
        _GOLD_DARK  = "7D6608"
        _GOLD_MED   = "B7950B"
        _GOLD_LIGHT = "FCF3CF"
        _GOLD_XL    = "FEFDE7"
        _NCOLS = 5  # adicionou coluna CGF R$

        ws = wb.create_sheet("📋 CGF")
        ws.sheet_view.showGridLines = False

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOLS)
        t = ws.cell(row=1, column=1,
                    value=f"  VOLUME CGF — CONTA GRÁFICA DE FATURAMENTO  |  Trimestre: {label_trimestre or 'N/D'}")
        t.fill = _fill(_GOLD_DARK); t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("left"); ws.row_dimensions[1].height = 34

        ws.row_dimensions[2].height = 6
        for c in range(1, _NCOLS + 1):
            ws.cell(row=2, column=c).fill = _fill("F9E79F")

        registros = cgf_lista or []
        if not registros:
            ws.cell(row=4, column=1, value="Nenhum dado CGF registrado para este trimestre.")
            return

        # Indexar por período para seguir a ordem de meses
        idx = {r.get("periodo", ""): r for r in registros}
        ordem = list(meses or [])
        for p in idx:
            if p not in ordem:
                ordem.append(p)

        grand_vf  = 0.0
        grand_rs  = 0.0
        row = 3

        def _cgf_cell(r, c, val, bg, bold=False, fmt="@", align_h="left"):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = _fill(bg); cell.border = _border()
            cell.font = _font(bold=bold, size=11)
            cell.alignment = _align(align_h, "center")
            if fmt != "@": cell.number_format = fmt
            return cell

        trimestres = _agrupar_em_trimestres(ordem)
        meses_por_trimestre = {m: i for i, tri in enumerate(trimestres) for m in tri}
        tri_atual = -1

        for periodo_mes in ordem:
            data = idx.get(periodo_mes)
            if not data:
                continue

            # Banner de trimestre a cada novo grupo de 3 meses
            tri_idx = meses_por_trimestre.get(periodo_mes, -1)
            if tri_idx != tri_atual:
                tri_atual = tri_idx
                tri_meses = trimestres[tri_idx]
                label_tri = "  🗓  TRIMESTRE:  " + "  ·  ".join(tri_meses)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
                bt = ws.cell(row=row, column=1, value=label_tri)
                bt.fill = _fill(_GOLD_DARK); bt.font = _font(bold=True, color="FFFFF0", size=12)
                bt.alignment = _align("left"); bt.border = _border()
                ws.row_dimensions[row].height = 24; row += 1

            ano_mes  = periodo_mes.split("/")[1] if "/" in periodo_mes else ""
            mes_full = _nome_mes_completo(periodo_mes).split("/")[0]

            # Busca CGF em R$ da tabela consolidacao para este período
            cgf_rs_mes = 0.0
            pmpv_mes   = 0.0
            if db:
                try:
                    cons_mes = db.buscar_consolidacao(periodo_mes)
                    cgf_rs_mes = _to_float((cons_mes or {}).get("cgf"))
                    pmpv_row = db.buscar_pmpv_mensal(periodo_mes) if hasattr(db, "buscar_pmpv_mensal") else None
                    if pmpv_row is None:
                        pmpv_mes = db.buscar_pmpv_por_periodo(periodo_mes)
                    else:
                        pmpv_mes = _to_float(pmpv_row) if isinstance(pmpv_row, (int, float)) else 0.0
                except Exception:
                    pass

            # Cabeçalho do mês
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_NCOLS)
            sec = ws.cell(row=row, column=1,
                          value=f"  ── {mes_full}{'/'+ano_mes if ano_mes else ''} " + "─"*40)
            sec.fill = _fill(_GOLD_MED); sec.font = _font(bold=True, color=_HEADER_FG, size=11)
            sec.alignment = _align("left"); sec.border = _border()
            ws.row_dimensions[row].height = 20; row += 1

            # Header de colunas — agora com CGF R$
            _apply_header_row(ws, row,
                ["Componente", "Volume (m³)", "Sinal", "PMPV (R$/m³)", "CGF (R$)"],
                [38, 20, 8, 16, 20], _GOLD_MED); row += 1

            campos = [
                ("(+) Volume Faturado (s/ cons. próprio)", data.get("volume_faturado", 0.0),         _ROW_NORM, "+"),
                ("   ↳ Consumo Próprio excluído",          data.get("volume_consumo_proprio", 0.0),  _ROW_ALT,  ""),
                ("(−) Volume Canceladas / Denegadas",      data.get("volume_canceladas", 0.0),       "FADBD8",  "(−)"),
                ("(−) Volume Devoluções",                  data.get("volume_devolucoes", 0.0),       "FADBD8",  "(−)"),
            ]
            for label, val, bg, sinal in campos:
                _cgf_cell(row, 1, label, bg)
                _cgf_cell(row, 2, val,   bg, fmt=_VOL,  align_h="right")
                _cgf_cell(row, 3, sinal, bg, align_h="center")
                _cgf_cell(row, 4, "",    bg)
                _cgf_cell(row, 5, "",    bg)
                row += 1

            # Volume Final + CGF em R$ do mês
            vf = _to_float(data.get("volume_final"))
            grand_vf += vf
            grand_rs += cgf_rs_mes

            _cgf_cell(row, 1, f"(=)  VOLUME FINAL CGF — {mes_full.upper()}", "D5F5E3", bold=True)
            _cgf_cell(row, 2, vf,         "D5F5E3", bold=True, fmt=_VOL, align_h="right")
            _cgf_cell(row, 3, "=",        "D5F5E3", align_h="center")
            _cgf_cell(row, 4, pmpv_mes if pmpv_mes else "", "D5F5E3",
                      fmt="#,##0.0000" if pmpv_mes else "@", align_h="right")
            _cgf_cell(row, 5, cgf_rs_mes if cgf_rs_mes else "", "D5F5E3",
                      bold=True, fmt=_BRL if cgf_rs_mes else "@", align_h="right")
            ws.row_dimensions[row].height = 22; row += 1
            ws.row_dimensions[row].height = 8;  row += 1

        # Total Geral
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        tl = ws.cell(row=row, column=1, value="  ══ TOTAL DO TRIMESTRE ══")
        tl.fill=_fill(_GOLD_DARK); tl.font=_font(bold=True,size=13,color=_HEADER_FG)
        tl.alignment=_align("left"); tl.border=_border()
        for ci in [2, 3]:
            ws.cell(row=row, column=ci).fill = _fill(_GOLD_DARK)
            ws.cell(row=row, column=ci).border = _border()
        tv = ws.cell(row=row, column=4, value=grand_vf)
        tv.fill=_fill(_GOLD_DARK); tv.font=_font(bold=True,size=13,color=_HEADER_FG)
        tv.number_format=_VOL; tv.alignment=_align("right"); tv.border=_border()
        tr = ws.cell(row=row, column=5, value=grand_rs if grand_rs else "")
        tr.fill=_fill(_GOLD_DARK); tr.font=_font(bold=True,size=13,color=_HEADER_FG)
        tr.number_format=_BRL; tr.alignment=_align("right"); tr.border=_border()
        ws.row_dimensions[row].height = 26

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 22

    # ── Sheet 7: RPV ─────────────────────────────────────────────────────────

    @staticmethod
    def _sheet_rpv(wb, cons_periodos: list[dict], meses_audit: list[str],
                   meses_cgf: list[str], db: "DatabasePMPV", periodo: str | None):
        _RPV_COLOR = "7D3C98"   # roxo RPV

        ws = wb.create_sheet("🧾 RPV")
        ws.sheet_view.showGridLines = False

        # Título
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = f"RPV  =  CGR  −  CGF   |   Período: {periodo or 'Todos'}"
        t.fill = _fill(_RPV_COLOR)
        t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("center")
        ws.row_dimensions[1].height = 30

        row = 3

        # Monta lista de períodos com CGR e CGF
        meses_todos = sorted(
            set(meses_audit + meses_cgf),
            key=lambda p: (
                int(p.split("/")[1]) * 12 + {
                    "Jan": 1, "Fev": 2, "Mar": 3, "Abr": 4, "Mai": 5, "Jun": 6,
                    "Jul": 7, "Ago": 8, "Set": 9, "Out": 10, "Nov": 11, "Dez": 12,
                }.get(p.split("/")[0], 0) if "/" in p else 0
            )
        )

        registros = []
        for mes in meses_todos:
            # CGR considera apenas NF-e (compra de gás) — CT-e é frete/transporte.
            cgr = sum(
                _to_float(i.get("cgr_liquido"))
                for i in (db.listar_auditoria_itens(mes) or [])
                if i.get("tipo") == "NF-e"
            )
            cons_m = db.buscar_consolidacao(mes) or {}
            cgf = _to_float(cons_m.get("cgf"))
            rpv = cgr - cgf
            registros.append({"periodo": mes, "cgr": cgr, "cgf": cgf, "rpv": rpv})

        # Cabeçalho
        _apply_header_row(ws, row,
            ["Período", "CGR (R$)", "CGF (R$)", "RPV = CGR − CGF (R$)"],
            [18, 22, 22, 26], _RPV_COLOR)
        row += 1

        _BRL = 'R$ #,##0.00'
        total_cgr = total_cgf = total_rpv = 0.0

        for i, item in enumerate(registros):
            cgr_v = item["cgr"]
            cgf_v = item["cgf"]
            rpv_v = item["rpv"]
            total_cgr += cgr_v
            total_cgf += cgf_v
            total_rpv += rpv_v

            bg = _ROW_ALT if i % 2 else _ROW_NORM
            rpv_color = _GREEN if rpv_v >= 0 else _RED

            for ci, (val, fmt, bold, color) in enumerate([
                (item["periodo"], "@",   False, "000000"),
                (cgr_v,          _BRL,  False, "000000"),
                (cgf_v,          _BRL,  False, "000000"),
                (rpv_v,          _BRL,  True,  rpv_color),
            ], start=1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = _fill(bg)
                c.font = _font(bold=bold, color=color)
                c.border = _border()
                c.alignment = _align("right" if fmt != "@" else "left")
                if fmt != "@":
                    c.number_format = fmt
            ws.row_dimensions[row].height = 18
            row += 1

        # Linha de total
        if registros:
            row += 1
            rpv_tot_color = _GREEN if total_rpv >= 0 else _RED
            for ci, (val, fmt, color) in enumerate([
                ("TOTAL",      "@",  _HEADER_FG),
                (total_cgr,   _BRL, _HEADER_FG),
                (total_cgf,   _BRL, _HEADER_FG),
                (total_rpv,   _BRL, _HEADER_FG),
            ], start=1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = _fill(rpv_tot_color if ci == 4 else _RPV_COLOR)
                c.font = _font(bold=True, color=color)
                c.border = _border()
                c.alignment = _align("right" if fmt != "@" else "left")
                if fmt != "@":
                    c.number_format = fmt
            ws.row_dimensions[row].height = 22

            # Card resultado RPV total
            row += 2
            ws.merge_cells(f"A{row}:D{row}")
            lbl = ws.cell(row=row, column=1, value="🧾  RPV TOTAL  =  CGR  −  CGF")
            bg_card = _GREEN if total_rpv >= 0 else _RED
            lbl.fill = _fill(bg_card)
            lbl.font = _font(bold=True, color=_HEADER_FG, size=13)
            lbl.alignment = _align("left")
            lbl.border = _border()
            ws.row_dimensions[row].height = 28
            row += 1

            ws.merge_cells(f"A{row}:D{row}")
            v = ws.cell(row=row, column=1, value=total_rpv)
            v.number_format = _BRL
            v.fill = _fill(bg_card)
            v.font = _font(bold=True, color=_HEADER_FG, size=22)
            v.alignment = _align("center")
            v.border = _border()
            ws.row_dimensions[row].height = 44

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 26

    # ── Sheet 7: SCG Final ───────────────────────────────────────────────────

    @staticmethod
    def _sheet_scg_mensal(wb, db, cons: dict | None, cons_periodos: list[dict], sr: dict | list[dict] | None, periodo: str | None):
        ws = wb.create_sheet("🧾 SCG Mensal")
        ws.sheet_view.showGridLines = False

        # ── Ordenação cronológica local ──────────────────────────────────────
        def _ord(p):
            ab = _abrev_de_periodo(p)
            an = p.split("/")[1] if "/" in p else "0"
            try:
                return int(an) * 12 + _MES_ORD.get(ab, 0)
            except Exception:
                return 0

        # ── Coleta os meses disponíveis nas tabelas de origem ────────────────
        def _meses_de(tabela):
            try:
                return sorted({r for r in db.listar_periodos_distintos(tabela) if r}, key=_ord)
            except Exception:
                return []

        meses_audit = _meses_de("auditoria_itens")
        meses_cgf   = _meses_de("consolidacao")   # CGF em R$ vem de consolidacao
        meses_ret   = _meses_de("ret_itens")
        meses_conc  = _meses_de("concilia_itens")
        meses_todos = sorted(set(meses_audit + meses_cgf + meses_ret + meses_conc), key=_ord)

        # ── Agrega cada componente por mês ────────────────────────────────────
        def _cgr_mes(m):
            # CGR considera apenas NF-e (compra de gás) — CT-e é frete/transporte.
            itens = db.listar_auditoria_itens(m) or []
            return sum(_to_float(i.get("cgr_liquido")) for i in itens if i.get("tipo") == "NF-e")

        def _cgf_mes(m):
            # CGF em R$ = volume × PMPV, salvo na tabela consolidacao pelo módulo CGF
            cons_m = db.buscar_consolidacao(m) or {}
            return _to_float(cons_m.get("cgf"))

        def _ret_mes(m):
            # RET já calculado (EAT × (1-PIS/COFINS) + EC) e salvo na consolidacao
            cons_m = db.buscar_consolidacao(m) or {}
            return _to_float(cons_m.get("ret"))

        def _rp_mes(m):
            itens = db.listar_concilia_itens(m) or []
            return sum(_to_float(i.get("valor")) for i in itens)

        # ── Título ────────────────────────────────────────────────────────────
        n_cols = 2 + len(meses_todos) + 1   # DADOS + UNIDADE + meses + TOTAL
        span = f"A1:{chr(64 + n_cols)}1"
        ws.merge_cells(span)
        t = ws["A1"]
        t.value = "APURAÇÃO MENSAL DO SALDO DA CONTA GRÁFICA (SCG)"
        t.fill = _fill(_NAVY)
        t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("center")
        ws.row_dimensions[1].height = 30

        row = 3

        # ── Cabeçalho de colunas ──────────────────────────────────────────────
        cabecalhos = ["DADOS", "UNIDADE"] + meses_todos + ["TOTAL"]
        larguras   = [28, 12] + [14] * len(meses_todos) + [16]
        _apply_header_row(ws, row, cabecalhos, larguras, _NAVY)
        row += 1

        # ── Linhas de dados ───────────────────────────────────────────────────
        def _linha(label, unidade, valores_mes: dict, fmt, alt=False):
            nonlocal row
            bg = _ROW_ALT if alt else _ROW_NORM
            total = sum(valores_mes.get(m, 0.0) for m in meses_todos)
            valores = [label, unidade] + [valores_mes.get(m, 0.0) for m in meses_todos] + [total]
            fmts    = ["@", "@"] + [fmt] * len(meses_todos) + [fmt]
            for col_idx, (val, nf) in enumerate(zip(valores, fmts), start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.fill   = _fill(bg)
                c.border = _border()
                c.font   = _font()
                if nf != "@":
                    c.number_format = nf
                    c.alignment = _align("right")
                else:
                    c.alignment = _align("left")
            row += 1

        cgr_por_mes = {m: _cgr_mes(m) for m in meses_todos}
        cgf_por_mes = {m: _cgf_mes(m) for m in meses_todos}
        ret_por_mes = {m: _ret_mes(m) for m in meses_todos}
        rp_por_mes  = {m: _rp_mes(m)  for m in meses_todos}
        rpv_por_mes = {m: cgr_por_mes[m] - cgf_por_mes[m] for m in meses_todos}
        scg_por_mes = {m: rpv_por_mes[m] + ret_por_mes[m] + rp_por_mes[m] for m in meses_todos}

        _linha("CGR  (Auditoria)",      "R$", cgr_por_mes, _BRL, alt=False)
        _linha("CGF  (Volume Faturado × PMPV)", "R$", cgf_por_mes, _BRL, alt=True)
        _linha("RPV  = CGR − CGF",          "R$", rpv_por_mes, _BRL, alt=False)
        _linha("RET  (EAT + EC)",            "R$", ret_por_mes, _BRL, alt=True)
        _linha("RP   (Penalidades)",         "R$", rp_por_mes,  _BRL, alt=False)

        # ── Linha SCG destacada ───────────────────────────────────────────────
        row += 1
        scg_total = sum(scg_por_mes.values())
        bg_scg = _GREEN if scg_total >= 0 else _RED
        scg_vals = ["SCG  =  RPV + RET + RP", "R$"] + [scg_por_mes.get(m, 0.0) for m in meses_todos] + [scg_total]
        scg_fmts = ["@", "@"] + [_BRL] * len(meses_todos) + [_BRL]
        for col_idx, (val, nf) in enumerate(zip(scg_vals, scg_fmts), start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill   = _fill(bg_scg)
            c.border = _border()
            c.font   = _font(bold=True, color=_HEADER_FG)
            if nf != "@":
                c.number_format = nf
                c.alignment = _align("right")
            else:
                c.alignment = _align("left")
        ws.row_dimensions[row].height = 24
        row += 2

        # ── SR (se disponível) ────────────────────────────────────────────────
        sr_lista = sr if isinstance(sr, list) else ([sr] if sr else [])
        sr_lista = [r for r in sr_lista if r]
        if sr_lista:
            _section_title(ws, row, "  📈  SR por Período", n_cols, _NAVY)
            row += 1
            _apply_header_row(ws, row,
                ["Período", "VP (m³)", "VF (m³)", "SR (R$)"],
                [18, 18, 18, 18], _NAVY)
            row += 1
            for i, item in enumerate(sr_lista):
                _apply_data_row(ws, row,
                    [item.get("periodo", ""), item.get("vp", 0.0),
                     item.get("vf", 0.0), item.get("sr", 0.0)],
                    ["@", _VOL, _VOL, _BRL],
                    alternate=(i % 2 == 1))
                row += 1

        # ── Ajusta larguras ───────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        for i, m in enumerate(meses_todos, start=3):
            ws.column_dimensions[chr(64 + i)].width = 16
        ws.column_dimensions[chr(64 + 2 + len(meses_todos) + 1)].width = 18

    # ── Sheet SCG Trimestral ─────────────────────────────────────────────────

    @staticmethod
    def _sheet_scg_trimestral(wb, db):
        """Aba SCG agrupada por trimestre fiscal (Nov–Jan, Fev–Abr, Mai–Jul, Ago–Out)."""

        TRIMESTRES_DEF = TRIMESTRES_FISCAIS_ABREVS
        _GOLD_SCG  = "B7950B"
        _GOLD_DARK = "7D6608"
        _NAVY_SCG  = "1A3A5C"

        ws = wb.create_sheet("📅 SCG Trimestral")
        ws.sheet_view.showGridLines = False

        # ── helpers internos ────────────────────────────────────────────────
        def _ord(p):
            ab = _abrev_de_periodo(p)
            an = p.split("/")[1] if "/" in p else "0"
            try:
                return int(an) * 12 + _MES_ORD.get(ab, 0)
            except Exception:
                return 0

        def _meses_de(tabela):
            try:
                return sorted({r for r in db.listar_periodos_distintos(tabela) if r}, key=_ord)
            except Exception:
                return []

        def _cgr_mes(m):
            # CGR considera apenas NF-e (compra de gás) — CT-e é frete/transporte.
            return sum(_to_float(i.get("cgr_liquido"))
                       for i in (db.listar_auditoria_itens(m) or [])
                       if i.get("tipo") == "NF-e")

        def _cgf_mes(m):
            return _to_float((db.buscar_consolidacao(m) or {}).get("cgf"))

        def _ret_mes(m):
            return _to_float((db.buscar_consolidacao(m) or {}).get("ret"))

        def _rp_mes(m):
            return sum(_to_float(i.get("valor"))
                       for i in (db.listar_concilia_itens(m) or []))

        # Todos os meses disponíveis no banco
        meses_banco = sorted(
            set(_meses_de("auditoria_itens") +
                _meses_de("consolidacao") +
                _meses_de("ret_itens") +
                _meses_de("concilia_itens")),
            key=_ord,
        )

        # Descobre os anos presentes
        anos = sorted({m.split("/")[1] for m in meses_banco if "/" in m})

        # ── Título ──────────────────────────────────────────────────────────
        ws.merge_cells("A1:H1")
        t = ws["A1"]
        t.value = "APURAÇÃO TRIMESTRAL DO SALDO DA CONTA GRÁFICA (SCG)"
        t.fill = _fill(_NAVY_SCG)
        t.font = _font(bold=True, size=14, color=_HEADER_FG)
        t.alignment = _align("center")
        ws.row_dimensions[1].height = 34

        # Larguras fixas: col A=etiqueta, B=unidade, C–F=3 meses, G=total trimestre
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 8
        for col in "CDEFG":
            ws.column_dimensions[col].width = 18
        ws.column_dimensions["H"].width = 20

        CAMPOS = [
            ("CGR  (Auditoria)",           "cgr"),
            ("CGF  (Volume × PMPV)",           "cgf"),
            ("RPV  = CGR − CGF",               "rpv"),
            ("RET  (EAT + EC)",                "ret"),
            ("RP   (Penalidades)",              "rp"),
        ]

        row = 3
        grand_scg_total = 0.0

        for ano in anos:
            # Cabeçalho do ano
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ac = ws.cell(row=row, column=1, value=f"  ANO: {ano}")
            ac.fill = _fill(_NAVY_SCG)
            ac.font = _font(bold=True, size=13, color=_HEADER_FG)
            ac.alignment = _align("left")
            ac.border = _border()
            ws.row_dimensions[row].height = 26
            row += 1

            ano_scg_total = 0.0

            for nome_tri, abrevs in TRIMESTRES_DEF:
                # Constrói a lista de períodos deste trimestre neste ano
                # Nov-Jan: Nov/ano-1, Dez/ano-1, Jan/ano
                if nome_tri == "Nov - Jan":
                    ano_anterior = str(int(ano) - 1)
                    periodos_tri = [f"Nov/{ano_anterior}", f"Dez/{ano_anterior}", f"Jan/{ano}"]
                else:
                    periodos_tri = [f"{ab}/{ano}" for ab in abrevs]

                # Só exibe trimestres que têm ao menos 1 mês com dados no banco
                periodos_com_dados = [p for p in periodos_tri if p in meses_banco]
                if not periodos_com_dados:
                    continue

                # Calcula valores por mês
                dados_mes = {}
                for p in periodos_tri:
                    cgr = _cgr_mes(p)
                    cgf = _cgf_mes(p)
                    ret = _ret_mes(p)
                    rp  = _rp_mes(p)
                    rpv = cgr - cgf
                    scg = rpv + ret + rp
                    dados_mes[p] = {"cgr": cgr, "cgf": cgf, "rpv": rpv,
                                    "ret": ret, "rp": rp, "scg": scg}

                # Banner do trimestre
                label_tri = f"  🗓  {nome_tri} / {ano}"
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                bt = ws.cell(row=row, column=1, value=label_tri)
                bt.fill = _fill(_GOLD_DARK)
                bt.font = _font(bold=True, size=12, color="FFFFF0")
                bt.alignment = _align("left")
                bt.border = _border()
                ws.row_dimensions[row].height = 22
                row += 1

                # Cabeçalho de colunas: DADOS | UNI | mês1 | mês2 | mês3 | TOTAL TRIMESTRE
                cab = ["DADOS", "UNI"] + periodos_tri + ["TOTAL TRIMESTRE"]
                _apply_header_row(ws, row, cab,
                                  [30, 8, 18, 18, 18, 20], _NAVY_SCG)
                row += 1

                # Linhas CGR, CGF, RPV, RET, RP
                for li, (label, key) in enumerate(CAMPOS):
                    bg = _ROW_ALT if li % 2 == 1 else _ROW_NORM
                    tot = sum(dados_mes[p][key] for p in periodos_tri)
                    vals = ([label, "R$"] +
                            [dados_mes[p][key] for p in periodos_tri] +
                            [tot])
                    fmts = ["@", "@"] + [_BRL] * 3 + [_BRL]
                    for ci, (v, nf) in enumerate(zip(vals, fmts), start=1):
                        c = ws.cell(row=row, column=ci, value=v)
                        c.fill = _fill(bg)
                        c.border = _border()
                        c.font = _font()
                        if nf != "@":
                            c.number_format = nf
                            c.alignment = _align("right")
                        else:
                            c.alignment = _align("left")
                    row += 1

                # Linha SCG (destaque)
                scg_total_tri = sum(dados_mes[p]["scg"] for p in periodos_tri)
                ano_scg_total += scg_total_tri
                bg_scg = _GREEN if scg_total_tri >= 0 else _RED
                scg_vals = (["SCG  =  RPV + RET + RP", "R$"] +
                            [dados_mes[p]["scg"] for p in periodos_tri] +
                            [scg_total_tri])
                scg_fmts = ["@", "@"] + [_BRL] * 3 + [_BRL]
                for ci, (v, nf) in enumerate(zip(scg_vals, scg_fmts), start=1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.fill = _fill(bg_scg)
                    c.border = _border()
                    c.font = _font(bold=True, color=_HEADER_FG)
                    if nf != "@":
                        c.number_format = nf
                        c.alignment = _align("right")
                    else:
                        c.alignment = _align("left")
                ws.row_dimensions[row].height = 22
                row += 2

            # Total do ano
            grand_scg_total += ano_scg_total
            bg_ano = _GREEN if ano_scg_total >= 0 else _RED
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ta = ws.cell(row=row, column=1,
                         value=f"  ══ TOTAL SCG ANO {ano} ══")
            ta.fill = _fill(bg_ano)
            ta.font = _font(bold=True, size=13, color=_HEADER_FG)
            ta.alignment = _align("left")
            ta.border = _border()
            tv = ws.cell(row=row, column=8, value=ano_scg_total)
            tv.fill = _fill(bg_ano)
            tv.font = _font(bold=True, size=13, color=_HEADER_FG)
            tv.number_format = _BRL
            tv.alignment = _align("right")
            tv.border = _border()
            for ci in range(2, 8):
                ws.cell(row=row, column=ci).fill = _fill(bg_ano)
                ws.cell(row=row, column=ci).border = _border()
            ws.row_dimensions[row].height = 26
            row += 2

        if not anos:
            ws.cell(row=row, column=1,
                    value="Nenhum dado SCG encontrado no banco.")

    # ── Sheet SR: Saldo Remanescente ─────────────────────────────────────────

    @staticmethod
    def _sheet_sr(wb, sr: dict | list[dict] | None, periodo: str | None):
        _sheet_sr_fn(wb, sr, periodo)

    # ── Sheet 8: PR Final ────────────────────────────────────────────────────

    @staticmethod
    def _sheet_pr(wb, pr: dict | list[dict] | None, periodo: str | None):
        _sheet_pr_fn(wb, pr, periodo)

    # ── Sheet 9: Progresso por Etapa ───────────────────────────────────────

    @staticmethod
    def _sheet_pv(wb, pv: dict | list[dict] | None, periodo: str | None):
        _sheet_pv_fn(wb, pv, periodo)

    # ── Sheet 10: Progresso por Etapa ──────────────────────────────────────

    @staticmethod
    def _sheet_progresso(wb, execucoes: list[dict], periodo: str | None):
        _sheet_progresso_fn(wb, execucoes, periodo)

    # ── Sheet 11: Dashboard Visual ───────────────────────────────────────────

    @staticmethod
    def _sheet_dashboard(
        wb,
        cons: dict | None,
        cons_periodos: list[dict],
        pr: dict | None,
        pv: dict | None,
        sr: dict | None,
        periodo: str | None,
    ):
        _sheet_dashboard_fn(wb, cons, cons_periodos, pr, pv, sr, periodo)

