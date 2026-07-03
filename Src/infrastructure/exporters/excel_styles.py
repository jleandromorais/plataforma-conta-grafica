"""Helpers de estilo compartilhados entre todos os exporters Excel."""
from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, size=11, color="000000", italic=False, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def border(style="thin", color="B0B0B0") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def border_bottom(color="CCCCCC") -> Border:
    return Border(bottom=Side(style="thin", color=color))


def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def to_float(val: Any, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default
