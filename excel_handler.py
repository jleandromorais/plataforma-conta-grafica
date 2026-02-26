import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date as date_type
from typing import Dict
import os

class ExcelHandlerPMPV:
    @staticmethod
    def exportar_trimestre(dados_por_mes: Dict, resultado: Dict, nome_arquivo: str = None) -> str:
        if nome_arquivo is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_arquivo = f"Relatorio_PMPV_{timestamp}.xlsx"
        
        # Evita sobrescrever arquivo já aberto - adiciona número incremental
        nome_base = nome_arquivo.replace('.xlsx', '')
        contador = 1
        nome_final = nome_arquivo
        
        while True:
            try:
                # Tenta criar/abrir o arquivo para verificar se está disponível
                with open(nome_final, 'w') as f:
                    pass
                os.remove(nome_final)
                break
            except (PermissionError, IOError):
                # Arquivo em uso, tenta próximo número
                nome_final = f"{nome_base}_{contador}.xlsx"
                contador += 1
                if contador > 100:  # Segurança para evitar loop infinito
                    nome_final = f"{nome_base}_{datetime.now().strftime('%H%M%S%f')}.xlsx"
                    break
        
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
        
        # Criar abas mensais
        for nome_aba, dados in dados_por_mes.items():
            ExcelHandlerPMPV._criar_aba_mes(wb, nome_aba, dados)
        
        # Criar aba de resumo
        ExcelHandlerPMPV._criar_aba_resumo(wb, dados_por_mes, resultado)
        
        wb.save(nome_final)
        wb.close()  # Fecha o workbook antes de tentar abrir
        
        # Tenta abrir o arquivo, mas não falha se houver erro
        try:
            os.startfile(nome_final)
        except Exception as e:
            print(f"Aviso: Não foi possível abrir o arquivo automaticamente: {e}")
        
        return nome_final

    @staticmethod
    def _criar_aba_mes(wb, nome_aba, dados):
        ws = wb.create_sheet(nome_aba)
        
        # Estilos
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Cabeçalhos
        headers = ["Empresa", "Molécula", "Transporte", "Logística", "Preço Unit.", "Volume (QDC)", "Custo Total"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for linha in dados:
            if not linha.get("empresa"): continue
            
            mol = float(linha.get('molecula', 0))
            trans = float(linha.get('transporte', 0))
            log = float(linha.get('logistica', 0))
            vol = float(linha.get('volume', 0))
            preco = mol + trans + log
            total = preco * vol
            
            ws.append([linha['empresa'], mol, trans, log, preco, vol, total])
            
        # Formatação
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.border = border
            for idx in [1, 2, 3, 4, 6]: row[idx].number_format = '#,##0.0000'
            row[5].number_format = '#,##0'

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["G"].width = 20

    @staticmethod
    def _criar_aba_resumo(wb, dados_por_mes, resultado):
        ws = wb.create_sheet("Resumo Executivo", 0)
        
        ws["A1"] = "FECHAMENTO TRIMESTRAL - PMPV"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")
        
        row = 3
        def write_res(label, val, fmt, bold=False, color="000000", bg=None):
            nonlocal row
            ws[f"A{row}"] = label
            ws[f"B{row}"] = val
            ws[f"B{row}"].number_format = fmt
            if bold: 
                ws[f"A{row}"].font = Font(bold=True, size=12)
                ws[f"B{row}"].font = Font(bold=True, size=12, color=color)
            if bg:
                fill = PatternFill("solid", fgColor=bg)
                ws[f"A{row}"].fill = fill
                ws[f"B{row}"].fill = fill
            row += 1

        write_res("Volume Total (Trimestre):", resultado['volume_total'], '#,##0')
        write_res("Custo Total (Trimestre):", resultado['custo_total'], 'R$ #,##0.00')
        row += 1
        write_res("PMPV Calculado:", resultado['pmpv'], 'R$ 0.0000', bold=True)
        write_res("(+) Conta Gráfica:", resultado.get('conta_grafica', 0), 'R$ 0.0000')
        row += 1
        write_res("(=) PREÇO FINAL (PV):", resultado.get('preco_final', 0), 'R$ 0.0000', bold=True, bg="F1C40F")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    # ─────────────────────────────────────────────────────────────────
    # Exportação no formato "Memória de Cálculo" (reimportável)
    # ─────────────────────────────────────────────────────────────────
    @staticmethod
    def exportar_memoria_calculo(
        dados_por_mes: Dict,
        resultado: Dict,
        dias_config: Dict,
        nome_arquivo: str = None,
    ) -> str:
        """
        Gera 'Memória de Cálculo.xlsx' compatível com o botão 'Importar MC'.

        Estrutura da aba PMPV (reimportável):
          Col B vazia + Col C = empresa + Col D/E/F = datas  →  cabeçalho empresa
          Col B = A/B/C/E    + Col C = descrição  + Col D/E/F = valores  →  dados
        """
        _MES_NUM = {
            "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4,
            "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8,
            "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12,
        }

        if nome_arquivo is None:
            nome_arquivo = "Memória de Cálculo.xlsx"

        # Evita sobrescrever arquivo aberto
        base = nome_arquivo.replace(".xlsx", "")
        nome_final = nome_arquivo
        for cnt in range(1, 101):
            try:
                with open(nome_final, "w"):
                    pass
                os.remove(nome_final)
                break
            except (PermissionError, IOError):
                nome_final = f"{base}_{cnt}.xlsx"
        else:
            nome_final = f"{base}_{datetime.now().strftime('%H%M%S%f')}.xlsx"

        meses_nomes = list(dados_por_mes.keys())          # ex: ['Outubro', 'Novembro', 'Dezembro']
        n = len(meses_nomes)
        dias = [dias_config.get(f"Mês {i+1}", 30) for i in range(n)]

        # Datas para as colunas (necessário para importação detectar meses)
        ano_base = datetime.now().year
        datas = []
        for i, mes in enumerate(meses_nomes):
            m_num = _MES_NUM.get(mes, i + 1)
            ano = ano_base
            if i > 0:
                prev_m = _MES_NUM.get(meses_nomes[i - 1], 1)
                if m_num < prev_m:
                    ano = ano_base + 1
            datas.append(datetime(ano, m_num, 1))

        # Agrupar por empresa preservando ordem de aparição
        empresas_ord: list[str] = []
        emp_data: dict[str, dict[int, dict]] = {}
        for i, (mes_nome, linhas) in enumerate(dados_por_mes.items()):
            for l in linhas:
                nome_emp = l.get("empresa", "").strip()
                if not nome_emp:
                    continue
                if nome_emp not in emp_data:
                    emp_data[nome_emp] = {}
                    empresas_ord.append(nome_emp)
                emp_data[nome_emp][i] = l

        # ── Workbook ──────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # ── Aba PMPV (importável) ──────────────────────────────────────
        ws = wb.create_sheet("PMPV")

        # Estilos
        emp_fill  = PatternFill("solid", fgColor="1e3a5f")
        emp_font  = Font(bold=True, color="FFFFFF", size=11)
        lbl_fill  = PatternFill("solid", fgColor="ecf0f1")
        bold_font = Font(bold=True, size=10)
        sum_font  = Font(bold=True, size=12, color="1e40af")
        border    = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        F4 = "#,##0.0000"
        F2 = "#,##0.00"
        F0 = "#,##0"

        def _cell(row, col, val=None, fmt=None, font=None, fill=None, align_center=False, bord=False):
            c = ws.cell(row, col, val)
            if fmt:           c.number_format = fmt
            if font:          c.font          = font
            if fill:          c.fill          = fill
            if align_center:  c.alignment     = Alignment(horizontal="center")
            if bord:          c.border        = border
            return c

        row = 1

        for emp_nome in empresas_ord:
            mes_dict = emp_data[emp_nome]

            # Preços e volumes por mês (0 se empresa ausente naquele mês)
            mols   = [mes_dict.get(i, {}).get("molecula",    0.0) for i in range(n)]
            transs = [mes_dict.get(i, {}).get("transporte",  0.0) for i in range(n)]
            logs   = [mes_dict.get(i, {}).get("logistica",   0.0) for i in range(n)]
            # vol = QDC (diário) — importação lê E direto como QDC
            vols   = [mes_dict.get(i, {}).get("volume",      0.0) for i in range(n)]
            totals = [m + t + l for m, t, l in zip(mols, transs, logs)]
            vol_m  = [v * d for v, d in zip(vols, dias)]   # volume mensal (para F)
            custos = [t * v for t, v in zip(totals, vol_m)]

            # ── Cabeçalho da empresa (detectado pelo importador) ──────
            # Col B vazia, Col C = nome empresa, Col D/E/F = datetime
            _cell(row, 2, val=None)                            # B: vazio (import exige)
            _cell(row, 3, emp_nome, font=emp_font, fill=emp_fill)
            for i, dt in enumerate(datas):
                c = _cell(row, 4 + i, dt, fmt="MMM/AA", align_center=True)
                c.fill = emp_fill
                c.font = Font(bold=True, color="FFFFFF")
            row += 1

            # ── A: Molécula ───────────────────────────────────────────
            _cell(row, 2, "A", font=bold_font, fill=lbl_fill, bord=True)
            _cell(row, 3, "Parcela da Molécula", bord=True)
            for i, v in enumerate(mols):
                _cell(row, 4 + i, v, fmt=F4, bord=True)
            row += 1

            # ── B: Transporte ─────────────────────────────────────────
            _cell(row, 2, "B", font=bold_font, fill=lbl_fill, bord=True)
            _cell(row, 3, "Parcela de Transporte", bord=True)
            for i, v in enumerate(transs):
                _cell(row, 4 + i, v, fmt=F4, bord=True)
            row += 1

            # ── C: Logística ──────────────────────────────────────────
            _cell(row, 2, "C", font=bold_font, fill=lbl_fill, bord=True)
            _cell(row, 3, "Parcela da Logística", bord=True)
            for i, v in enumerate(logs):
                _cell(row, 4 + i, v, fmt=F4, bord=True)
            row += 1

            # ── D = A + B + C ─────────────────────────────────────────
            _cell(row, 2, "D = A + B + C", font=bold_font, fill=PatternFill("solid", fgColor="d5e8d4"), bord=True)
            _cell(row, 3, "Total (R$/m³)", bord=True)
            for i, v in enumerate(totals):
                _cell(row, 4 + i, v, fmt=F4, bord=True)
            row += 2  # pula uma linha

            # ── E: Volume (QDC diário — IMPORTÁVEL) ───────────────────
            _cell(row, 2, "E", font=bold_font, fill=lbl_fill, bord=True)
            _cell(row, 3, "QDC — Volume diário previsto (m³/dia)", bord=True)
            for i, v in enumerate(vols):
                _cell(row, 4 + i, v, fmt=F0, bord=True)
            row += 2  # pula uma linha

            # ── F = D x E x dias ──────────────────────────────────────
            _cell(row, 2, "F = D × E × dias", font=bold_font, fill=PatternFill("solid", fgColor="dae8fc"), bord=True)
            _cell(row, 3, "Custo total mensal ex impostos (R$)", bord=True)
            for i, v in enumerate(custos):
                _cell(row, 4 + i, v, fmt=F2, bord=True)
            row += 2  # espaço entre empresas

        # ── Totalizadores G / H / I ────────────────────────────────────
        def _soma_mes(idx_mes):
            v_tot = h_tot = 0.0
            for emp in empresas_ord:
                md = emp_data[emp]
                qdc   = md.get(idx_mes, {}).get("volume",    0.0)
                mol   = md.get(idx_mes, {}).get("molecula",  0.0)
                trans = md.get(idx_mes, {}).get("transporte",0.0)
                log   = md.get(idx_mes, {}).get("logistica", 0.0)
                v_mes = qdc * dias[idx_mes]
                v_tot += v_mes
                h_tot += (mol + trans + log) * v_mes
            return v_tot, h_tot

        g_mes = [_soma_mes(i)[0] for i in range(n)]
        h_mes = [_soma_mes(i)[1] for i in range(n)]
        g_tot = sum(g_mes)
        h_tot = sum(h_mes)
        pmpv  = resultado.get("pmpv", 0.0)
        cg    = resultado.get("conta_grafica", 0.0)
        pv    = resultado.get("preco_final", 0.0)

        tot_fill = PatternFill("solid", fgColor="2c3e50")
        tot_font = Font(bold=True, color="FFFFFF", size=11)

        _cell(row, 2, "G", font=tot_font, fill=tot_fill)
        _cell(row, 3, "Total de volume (m³)", fill=tot_fill, font=Font(color="FFFFFF"))
        for i, v in enumerate(g_mes):
            _cell(row, 4 + i, v, fmt=F0, bord=True)
        _cell(row, 4 + n, g_tot, fmt=F0, font=Font(bold=True), bord=True)
        row += 1

        _cell(row, 2, "H", font=tot_font, fill=tot_fill)
        _cell(row, 3, "Custo total ex impostos (R$)", fill=tot_fill, font=Font(color="FFFFFF"))
        for i, v in enumerate(h_mes):
            _cell(row, 4 + i, v, fmt=F2, bord=True)
        _cell(row, 4 + n, h_tot, fmt=F2, font=Font(bold=True), bord=True)
        row += 2

        _cell(row, 2, "I = H ÷ G", font=sum_font)
        _cell(row, 3, "Custo médio ponderado s/ conta gráfica (PMPV)  [R$/m³]", font=Font(bold=True))
        _cell(row, 4 + n, pmpv, fmt=F4, font=sum_font)
        row += 1

        _cell(row, 2, "(+) Conta Gráfica")
        _cell(row, 4 + n, cg, fmt=F4)
        row += 1

        pv_fill = PatternFill("solid", fgColor="f1c40f")
        _cell(row, 2, "(=) Preço de Venda (PV)", font=Font(bold=True, size=12), fill=pv_fill)
        _cell(row, 4 + n, pv, fmt=F4, font=Font(bold=True, size=12, color="1e3a5f"), fill=pv_fill)
        row += 2

        # VP — Volume Prospecto Total (para uso no módulo SR)
        vp_fill = PatternFill("solid", fgColor="d6eaf8")
        _cell(row, 2, "VP", font=Font(bold=True, color="1a5276"), fill=vp_fill)
        _cell(row, 3, "Volume Prospecto Total (m³)  [uso SR]", fill=vp_fill)
        _cell(row, 4 + n, g_tot, fmt=F0,
              font=Font(bold=True, color="1a5276"), fill=vp_fill)

        # Larguras das colunas
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 48
        for i in range(n + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 17

        # ── Aba Composição (resumo visual) ─────────────────────────────
        ws2 = wb.create_sheet("Composição", 0)

        ws2["A1"] = "Memória de Cálculo — Custo Médio Ponderado de Venda"
        ws2["A1"].font = Font(bold=True, size=14, color="1e3a5f")
        ws2.merge_cells(f"A1:{openpyxl.utils.get_column_letter(2 + n)}1")
        ws2.row_dimensions[1].height = 22

        # Sub-cabeçalho de meses
        ws2.cell(2, 2, "Componente").font = Font(bold=True)
        for i, (dt, d) in enumerate(zip(datas, dias)):
            c = ws2.cell(2, 3 + i, dt)
            c.number_format = "MMM/AA"
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            ws2.cell(2, 3 + i).fill = PatternFill("solid", fgColor="2c3e50")
            ws2.cell(2, 3 + i).font = Font(bold=True, color="FFFFFF")
            # dias na linha 3
            ws2.cell(3, 3 + i, f"{d} dias").alignment = Alignment(horizontal="center")
            ws2.cell(3, 3 + i).font = Font(italic=True, size=9)

        sec_fill = PatternFill("solid", fgColor="2c3e50")
        sec_font = Font(bold=True, color="FFFFFF")

        def _sec(r, label):
            ws2.cell(r, 1, label).font = sec_font
            ws2.cell(r, 1).fill = sec_fill
            ws2.merge_cells(f"A{r}:{openpyxl.utils.get_column_letter(2 + n)}{r}")

        r = 4
        _sec(r, "VOLUMES (QDC — m³/dia)")
        r += 1
        for emp in empresas_ord:
            ws2.cell(r, 2, emp)
            for i in range(n):
                v = emp_data[emp].get(i, {}).get("volume", 0.0)
                ws2.cell(r, 3 + i, v).number_format = F0
            r += 1

        r += 1
        _sec(r, "PREÇOS — MOLÉCULA (R$/m³)")
        r += 1
        for emp in empresas_ord:
            ws2.cell(r, 2, emp)
            for i in range(n):
                v = emp_data[emp].get(i, {}).get("molecula", 0.0)
                if v:
                    ws2.cell(r, 3 + i, v).number_format = F4
            r += 1

        r += 1
        _sec(r, "PREÇOS — TRANSPORTE (R$/m³)")
        r += 1
        for emp in empresas_ord:
            ws2.cell(r, 2, emp)
            for i in range(n):
                v = emp_data[emp].get(i, {}).get("transporte", 0.0)
                if v:
                    ws2.cell(r, 3 + i, v).number_format = F4
            r += 1

        r += 1
        _sec(r, "PREÇOS — LOGÍSTICA (R$/m³)")
        r += 1
        for emp in empresas_ord:
            ws2.cell(r, 2, emp)
            for i in range(n):
                v = emp_data[emp].get(i, {}).get("logistica", 0.0)
                if v:
                    ws2.cell(r, 3 + i, v).number_format = F4
            r += 1

        r += 2
        res_fill = PatternFill("solid", fgColor="f1c40f")
        ws2.cell(r, 1, "PMPV Calculado (R$/m³)").font = Font(bold=True, size=12)
        ws2.cell(r, 3 + n - 1, pmpv).number_format = F4
        ws2.cell(r, 3 + n - 1).font = Font(bold=True, size=12, color="1e3a5f")
        ws2.cell(r, 3 + n - 1).fill = res_fill
        r += 1
        ws2.cell(r, 1, "(+) Conta Gráfica").font = Font(italic=True)
        ws2.cell(r, 3 + n - 1, cg).number_format = F4
        r += 1
        ws2.cell(r, 1, "(=) Preço de Venda (PV)").font = Font(bold=True, size=12)
        ws2.cell(r, 3 + n - 1, pv).number_format = F4
        ws2.cell(r, 3 + n - 1).font = Font(bold=True, size=12, color="1e3a5f")
        ws2.cell(r, 3 + n - 1).fill = res_fill
        r += 2

        # VP — Volume Prospecto Total
        vp_fill2 = PatternFill("solid", fgColor="d6eaf8")
        ws2.cell(r, 1, "VP — Volume Prospecto Total (m³)").font = Font(bold=True, color="1a5276")
        ws2.cell(r, 1).fill = vp_fill2
        ws2.cell(r, 3 + n - 1, g_tot).number_format = F0
        ws2.cell(r, 3 + n - 1).font = Font(bold=True, color="1a5276")
        ws2.cell(r, 3 + n - 1).fill = vp_fill2

        ws2.column_dimensions["A"].width = 38
        ws2.column_dimensions["B"].width = 30
        for i in range(n):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(3 + i)].width = 16

        # ── Salvar ────────────────────────────────────────────────────
        wb.save(nome_final)
        wb.close()
        try:
            os.startfile(nome_final)
        except Exception as e:
            print(f"Aviso: não foi possível abrir automaticamente: {e}")

        return nome_final